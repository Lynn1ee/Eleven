"""排班引擎 — 从 generate_schedule.py 提取，完全参数化"""
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import random
import io
import json
import calendar
import math
import urllib.request

# 节假日 API
HOLIDAY_API = "https://timor.tech/api/holiday/year"
# 内存缓存：{year: {month: {day: {"holiday": bool, "name": str}}}}
_holiday_cache = {}

# ============================================================
# 班次名称常量（不随月份变化）
# ============================================================
XI = '休'
CHANG_BAI = '长白班'
XING_ZHENG = '行政班'
ZAO_ZAO_IM = '早早班\n(IM)'
WAN_YI_IM = '晚一\n(IM)'
XING_ZHENG_JD = '行政班\n(机动)'
ZAO_BAN = '早班'
ZAO_SAN = '早三'
BAI_BAN = '白班'
ZHONG_SAN = '中三'
ZHONG_SI = '中四'
WAN_YI = '晚一'
WAN_ER = '晚二'
DA_YE = '大夜'
ZAO_ZAO_2 = '早早2'
TIAN_DI = '天地班'
QING_JIA = '请假'
CHANG_BAI_ZAO = '长白-早'

LATE_SHIFTS = {ZHONG_SAN, ZHONG_SI, WAN_YI, WAN_ER, WAN_YI_IM}
EARLY_SHIFTS = {ZAO_BAN, ZAO_SAN, XING_ZHENG, BAI_BAN, ZAO_ZAO_IM, XING_ZHENG_JD, ZAO_ZAO_2}


# ── 全员工号库（唯一数据源）──
STAFF_DB = {
    '李玲': '62416', '何丹妮': '63032', '曾瑾': '75103',
    '周妙笛': '32054', '李珂': '63171', '陈佳慧': '32014',
    '王颖': '75112', '王欣怡': '63898', '张玉婕': '63885', '魏溪': '75105',
    '舒丹': '63661',
    '盛坤坤': '63715', '褚婉婷': '63810', '余晓芬': '63574', '胡文思': '63802',
    '余飞滔': '63610', '周雨襄': '63704', '杜梦薇': '63891', '罗小婷': '63359',
    '李艾洛': '63882', '李蒙': '63367', '贺亚兰': '63809', '邵今令': '63849',
    '马宇杰': '63805', '林雪薇': '63388', '李娜': '63655', '陈文萍': '63874',
    '张怡顺': '63837', '张慧琳': '63680', '葛宇': '63872', '龚格': '63480',
    '富惟佳': '63845', '陈瑞楠': '63819', '范敏': '63914', '徐学宇': '63894',
    '邹秋霞': '32007',
}

# 夜班分组（固定）
NIGHT_SHIFT_GROUPS = {'王颖': 'B', '王欣怡': 'B', '张玉婕': 'A', '魏溪': 'A'}


class ScheduleEngine:
    """排班引擎，每次生成创建一个实例，线程安全"""

    def __init__(self, year, month, num_days=None, start_weekday=None, daily_targets=None,
                 banzhang=None, fenxiao=None, night_shift=None, buru_ban=None,
                 buru_support=None, changbingjia=None, online_staff=None,
                 theoretical_targets=None, prev_month_data=None, rest_requests=None):
        self.year = year
        self.month = month
        # 自动计算天数和首日星期
        self.num_days = calendar.monthrange(year, month)[1]
        self.start_weekday = date(year, month, 1).weekday()  # 0=周一, 6=周日
        # 加载节假日
        self._holidays = self._fetch_holidays(year)
        self.daily_targets = daily_targets or []
        self.theoretical_targets = theoretical_targets  # 未取整的理论目标值（可能为None）
        self.prev_month_data = prev_month_data or {}  # 跨月衔接数据
        self.rest_requests = rest_requests or {}  # 员工提前请休需求 {"张三": [3,5,15]}
        self.banzhang = banzhang if banzhang else self._build_banzhang(month)
        self.fenxiao = self._resolve_staff(fenxiao or [], 'fenxiao')
        self.night_shift = self._resolve_staff(night_shift or [], 'night_shift')
        self.buru_ban = self._resolve_staff(buru_ban or [], 'buru_ban')
        self.buru_support = self._resolve_staff(buru_support or [], 'buru_support')
        self.changbingjia = self._resolve_staff(changbingjia or [], 'changbingjia')
        if online_staff:
            self.online_staff = self._resolve_staff(online_staff, 'online')
        else:
            excluded = set()
            for g in [self.banzhang, self.fenxiao, self.night_shift, self.buru_ban, self.buru_support, self.changbingjia]:
                for p in g:
                    excluded.add(p['name'])
            self.online_staff = [{'name': n, 'id': STAFF_DB[n]} for n in STAFF_DB if n not in excluded]

    @staticmethod
    def _resolve_staff(data, role):
        """将姓名或字典列表解析为完整人员字典（自动补全工号）"""
        result = []
        for item in data:
            if isinstance(item, str):
                name = item
                entry = {'name': name, 'id': STAFF_DB.get(name, '')}
            else:
                name = item.get('name', '')
                entry = {'name': name, 'id': item.get('id') or STAFF_DB.get(name, '')}
                if 'primary' in item:
                    entry['primary'] = item['primary']
            if role == 'night_shift':
                entry['group'] = item.get('group') if isinstance(item, dict) and item.get('group') else NIGHT_SHIFT_GROUPS.get(name, 'A')
            result.append(entry)
        return result

    @staticmethod
    def load_prev_month_data(excel_path, num_days):
        """从 Excel 读取上月班表，从最后一天往前数到状态改变，提取连上/连休/最后班次"""
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        people = {}
        roles = {}
        for row in range(2, ws.max_row + 1):
            name = str(ws.cell(row=row, column=1).value or '').strip()
            role = str(ws.cell(row=row, column=3).value or '').strip()
            if not name or role in ('', '合计') or name.startswith('='):
                continue

            shifts = []
            for day in range(1, num_days + 1):
                v = str(ws.cell(row=row, column=3 + day).value or '')
                shifts.append(v)
            people[name] = shifts
            roles[name] = role

        work_streak = {}
        rest_streak = {}
        last_shift = {}
        second_last_shift = {}
        for name, shifts in people.items():
            last_shift[name] = shifts[-1] if shifts else None
            second_last_shift[name] = shifts[-2] if len(shifts) >= 2 else None
            # 从最后一天往前数，直到遇到休息日为止
            ws_count = 0
            for s in reversed(shifts):
                if s and s not in (XI, QING_JIA):
                    ws_count += 1
                else:
                    break
            work_streak[name] = ws_count
            # 从最后一天往前数，直到遇到非休息日为止
            rs_count = 0
            for s in reversed(shifts):
                if s in (XI, QING_JIA):
                    rs_count += 1
                else:
                    break
            rest_streak[name] = rs_count

        return {
            'last_shift': last_shift, 'work_streak': work_streak,
            'rest_streak': rest_streak,
            'second_last_shift': second_last_shift, 'roles': roles,
        }

    # 班长轮换池：6月李玲→7月何丹妮→8月曾瑾→循环
    BANZHANG_POOL = [
        {'name': '李玲', 'id': '62416'},
        {'name': '何丹妮', 'id': '63032'},
        {'name': '曾瑾', 'id': '75103'},
    ]

    @staticmethod
    def _build_banzhang(month):
        """根据月份自动轮换班长角色，每月一行政两长白"""
        idx = (month - 6) % 3
        result = []
        for i, p in enumerate(ScheduleEngine.BANZHANG_POOL):
            result.append({**p, 'primary': '行政' if i == idx else '长白'})
        return result

    @staticmethod
    def _fetch_holidays(year):
        """获取国务院节假日数据，内存缓存"""
        if year in _holiday_cache:
            return _holiday_cache[year]

        try:
            req = urllib.request.Request(
                f"{HOLIDAY_API}/{year}",
                headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
            )
            resp = urllib.request.urlopen(req, timeout=10)
            data = json.loads(resp.read())
            if data.get("code") == 0 and data.get("holiday"):
                # 转换为 {month: {day: {"holiday": bool, "name": str, "wage": int}}}
                parsed = {}
                for date_str, info in data["holiday"].items():
                    # date_str 格式: "01-01"
                    m, d = date_str.split("-")
                    m_int, d_int = int(m), int(d)
                    if m_int not in parsed:
                        parsed[m_int] = {}
                    parsed[m_int][d_int] = {
                        "holiday": info.get("holiday", False),
                        "name": info.get("name", ""),
                        "wage": info.get("wage", 1),
                    }
                _holiday_cache[year] = parsed
                return parsed
        except Exception:
            pass
        _holiday_cache[year] = {}
        return {}

    # ── 工具方法 ──

    def is_weekend(self, day):
        """判断是否为周六/周日（用于 Excel 红色字体等显示用途）"""
        return (self.start_weekday + day - 1) % 7 >= 5

    def is_rest_day(self, day):
        """判断是否为休息日（周末 + 法定假日 - 调休补班）"""
        dt = date(self.year, self.month, day)
        weekday = dt.weekday()
        is_weekend = weekday >= 5
        month_holidays = self._holidays.get(self.month, {})

        if day in month_holidays:
            info = month_holidays[day]
            if info["holiday"]:
                return True   # 法定假日 → 休息
            else:
                return False  # 调休补班 → 上班（即使周末）
        return is_weekend      # 普通日子按周末/工作日

    def is_workday(self, day):
        """判断是否为工作日（含调休补班）"""
        return not self.is_rest_day(day)

    def is_saturday(self, day):
        return (self.start_weekday + day - 1) % 7 == 5

    def is_sunday(self, day):
        return (self.start_weekday + day - 1) % 7 == 6

    def get_holiday_name(self, day):
        """获取节假日名称，非节假日返回空字符串"""
        month_holidays = self._holidays.get(self.month, {})
        if day in month_holidays:
            return month_holidays[day].get("name", "")
        return ""

    def get_holiday_info(self, day):
        """获取节假日完整信息 {name, wage, holiday}，非节假日返回 None"""
        month_holidays = self._holidays.get(self.month, {})
        if day in month_holidays:
            return month_holidays[day]
        return None

    def get_all_holidays(self):
        """获取当月所有节假日列表 [{day, name, wage, holiday}]"""
        result = []
        month_holidays = self._holidays.get(self.month, {})
        for day, info in sorted(month_holidays.items()):
            result.append({
                "day": day,
                "name": info.get("name", ""),
                "wage": info.get("wage", 1),
                "holiday": info.get("holiday", False),
            })
        return result

    def _find_rest_blocks(self, rest_days_set):
        rest_list = sorted(rest_days_set)
        if not rest_list:
            return []
        blocks = []
        block_start = rest_list[0]
        prev = rest_list[0]
        for d in rest_list[1:]:
            if d == prev + 1:
                prev = d
            else:
                blocks.append(list(range(block_start, prev + 1)))
                block_start = d
                prev = d
        blocks.append(list(range(block_start, prev + 1)))
        return blocks

    def _max_consecutive(self, work_mask):
        mx = cur = 0
        for v in work_mask:
            if v:
                cur += 1
                mx = max(mx, cur)
            else:
                cur = 0
        return mx

    def _streak_stats(self, sched):
        max_w = max_r = cur_w = cur_r = 0
        for s in sched:
            if s != XI:
                cur_w += 1
                max_w = max(max_w, cur_w)
                cur_r = 0
            else:
                cur_r += 1
                max_r = max(max_r, cur_r)
                cur_w = 0
        return max_w, max_r

    # ── 1. 班长排班 ──

    def _banzhang_valid(self, sched):
        """检查班长排班是否满足约束：连上≤6，连休≤2"""
        mw, mr = self._streak_stats(sched)
        return mw <= 6 and mr <= 2

    def _calc_banzhang_hours(self, sched):
        """计算班长排班的总工时"""
        h = 0.0
        for d in range(1, self.num_days + 1):
            shift = sched[d - 1]
            if shift == CHANG_BAI:
                h += 11.5
            elif shift == CHANG_BAI_ZAO:
                h += 12
            elif shift == XING_ZHENG:
                h += 8
        return h

    def _generate_banzhang(self):
        # 按 primary 分类：行政班 + 长白班
        xingzheng_people = [p for p in self.banzhang if p.get('primary') == '行政']
        changbai_people = [p for p in self.banzhang if p.get('primary') == '长白']
        pm = self.prev_month_data

        # ── 1. 确定跨月衔接：上月各长白位置在月末的上2休2相位 ──
        # 4天周期: 上(0)→上(1)→休(2)→休(3)
        prev_month = self.month - 1 if self.month > 1 else 12
        prev_idx = (prev_month - 6) % 3
        # 上月长白位置的人员（按 BANZHANG_POOL 顺序，跳过行政角色）
        prev_changbai_names = []
        for i, p in enumerate(ScheduleEngine.BANZHANG_POOL):
            if i != prev_idx:
                prev_changbai_names.append(p['name'])

        # 为每个长白位置确定 Day1 在周期中的相位偏移
        position_offset = []  # offset ∈ {0,1,2,3}
        for pos_idx in range(2):
            prev_name = prev_changbai_names[pos_idx]
            if pm and prev_name:
                ws_val = pm.get('work_streak', {}).get(prev_name, 0)
                rs_val = pm.get('rest_streak', {}).get(prev_name, 0)
                if ws_val > 0:
                    # 上月月末在上班：ws=1→位置0(上班第1天), ws=2→位置1(上班第2天)
                    # offset = 下一位置 = ws_val (1→1, 2→2)
                    offset = ws_val
                elif rs_val > 0:
                    # 上月月末在休息：rs=1→位置2(休息第1天), rs=2→位置3(休息第2天)
                    # offset = 下一位置 = (2 + rs_val) % 4 (1→3, 2→0)
                    offset = (2 + rs_val) % 4
                else:
                    offset = 0 if pos_idx == 0 else 2
            else:
                # 无上月数据，回退默认：长白1从上班开始(offset=0)，长白2从休息开始(offset=2)
                offset = 0 if pos_idx == 0 else 2
            position_offset.append(offset)

        # ── 2. 生成基础排班（上2休2 + 跨月衔接）──
        schedules = {p['name']: [None] * self.num_days for p in self.banzhang}

        # 行政班：工作日行政，休息日休
        for p in xingzheng_people:
            for day in range(1, self.num_days + 1):
                schedules[p['name']][day - 1] = XING_ZHENG if self.is_workday(day) else XI

        # 长白班：按上月衔接相位延续上2休2
        for pos_idx, person in enumerate(changbai_people):
            name = person['name']
            offset = position_offset[pos_idx]
            for day in range(1, self.num_days + 1):
                pos = (day - 1 + offset) % 4
                # pos 0,1 = 长白班（上班）, pos 2,3 = 休
                schedules[name][day - 1] = CHANG_BAI if pos < 2 else XI

        # ── 3. 动态工时调整 ──
        workday_count = sum(1 for d in range(1, self.num_days + 1) if self.is_workday(d))
        standard_hours = workday_count * 8

        for person in changbai_people:
            name = person['name']
            sched = schedules[name]
            current_hours = self._calc_banzhang_hours(sched)
            deficit = standard_hours - current_hours  # >0 工时不足，<0 工时超出

            if abs(deficit) <= 2:
                continue  # 偏差在2h内，接受

            if deficit > 0:
                # 工时不足 → 补班（只补工作日，不补周末节假日）
                self._adjust_banzhang_add(sched, deficit)
            else:
                # 工时超出 → 减班
                self._adjust_banzhang_reduce(sched, -deficit)

        return schedules

    def _adjust_banzhang_add(self, sched, deficit):
        """工时不足时补班：休息→行政(+8h) / 休息→长白(+11.5h) / 长白→长白早(+0.5h)。
        补班只补工作日（is_workday=True），周末节假日不补。"""
        workday_rest = sorted([d for d in range(1, self.num_days + 1)
                               if sched[d - 1] == XI and self.is_workday(d)])
        changbai_days = sorted([d for d in range(1, self.num_days + 1)
                                if sched[d - 1] == CHANG_BAI])

        # 迭代：每次选最优的单次操作，直到偏差≤2h或无有效操作
        while abs(deficit) > 2:
            best_action = None  # (type, day, new_deficit)
            best_remaining = abs(deficit)

            # 尝试 休息→长白 (+11.5h)
            for d in workday_rest:
                test = list(sched)
                test[d - 1] = CHANG_BAI
                if self._banzhang_valid(test):
                    nd = deficit - 11.5
                    if abs(nd) < best_remaining:
                        best_remaining = abs(nd)
                        best_action = ('rest_to_changbai', d, nd)

            # 尝试 休息→行政 (+8h)
            for d in workday_rest:
                test = list(sched)
                test[d - 1] = XING_ZHENG
                if self._banzhang_valid(test):
                    nd = deficit - 8
                    if abs(nd) < best_remaining:
                        best_remaining = abs(nd)
                        best_action = ('rest_to_xingzheng', d, nd)

            # 尝试 长白→长白早 (+0.5h)
            for d in changbai_days:
                test = list(sched)
                test[d - 1] = CHANG_BAI_ZAO
                if self._banzhang_valid(test):
                    nd = deficit - 0.5
                    if abs(nd) < best_remaining:
                        best_remaining = abs(nd)
                        best_action = ('changbai_to_zao', d, nd)

            if best_action is None:
                break  # 无有效操作，接受当前偏差

            action_type, d, deficit = best_action
            if action_type == 'rest_to_changbai':
                sched[d - 1] = CHANG_BAI
                workday_rest.remove(d)
                changbai_days.append(d)
            elif action_type == 'rest_to_xingzheng':
                sched[d - 1] = XING_ZHENG
                workday_rest.remove(d)
            elif action_type == 'changbai_to_zao':
                sched[d - 1] = CHANG_BAI_ZAO
                changbai_days.remove(d)

    def _adjust_banzhang_reduce(self, sched, excess):
        """工时超出时减班：长白→行政(-3.5h) / 长白→休息(-11.5h) / 长白早→长白(-0.5h)。"""
        changbai_days = sorted([d for d in range(1, self.num_days + 1)
                                if sched[d - 1] == CHANG_BAI], reverse=True)
        changbai_zao_days = sorted([d for d in range(1, self.num_days + 1)
                                     if sched[d - 1] == CHANG_BAI_ZAO], reverse=True)

        while abs(excess) > 2:
            best_action = None
            best_remaining = abs(excess)

            # 尝试 长白→行政 (-3.5h)
            for d in changbai_days:
                test = list(sched)
                test[d - 1] = XING_ZHENG
                if self._banzhang_valid(test):
                    nd = excess - 3.5
                    if abs(nd) < best_remaining:
                        best_remaining = abs(nd)
                        best_action = ('changbai_to_xingzheng', d, nd)

            # 尝试 长白早→长白 (-0.5h)
            for d in changbai_zao_days:
                test = list(sched)
                test[d - 1] = CHANG_BAI
                if self._banzhang_valid(test):
                    nd = excess - 0.5
                    if abs(nd) < best_remaining:
                        best_remaining = abs(nd)
                        best_action = ('zao_to_changbai', d, nd)

            # 尝试 长白→休息 (-11.5h)
            for d in changbai_days:
                test = list(sched)
                test[d - 1] = XI
                if self._banzhang_valid(test):
                    nd = excess - 11.5
                    if abs(nd) < best_remaining:
                        best_remaining = abs(nd)
                        best_action = ('changbai_to_rest', d, nd)

            if best_action is None:
                break

            action_type, d, excess = best_action
            if action_type == 'changbai_to_xingzheng':
                sched[d - 1] = XING_ZHENG
                changbai_days.remove(d)
            elif action_type == 'zao_to_changbai':
                sched[d - 1] = CHANG_BAI
                changbai_zao_days.remove(d)
                changbai_days.append(d)
            elif action_type == 'changbai_to_rest':
                sched[d - 1] = XI
                changbai_days.remove(d)

    # ── 2. 专职分销排班 ──

    def _generate_fenxiao(self):
        names = [p['name'] for p in self.fenxiao]
        offsets = [0, 2, 4]
        rest_days = {}
        pm = self.prev_month_data
        forced_fx = {}  # name → set of days，跨月强制休息不可删除
        for i, name in enumerate(names):
            rest_days[name] = set()
            forced_fx[name] = set()
            offset = offsets[i]
            for day in range(1, self.num_days + 1):
                pos = (day - 1 + offset) % 6
                if pos >= 4:
                    rest_days[name].add(day)
            # 跨月强制Day1休息
            if pm.get('last_shift', {}).get(name) == DA_YE:
                rest_days[name].add(1)
                forced_fx[name].add(1)
            elif pm.get('work_streak', {}).get(name, 0) >= 5:
                rest_days[name].add(1)
                forced_fx[name].add(1)

        # 分销Day1最少2人在岗（IM早晚班需要配对），不够则回退非大夜的休息
        day1_working = [n for n in names if 1 not in rest_days[n]]
        if len(day1_working) < 2:
            def _fx_rest_priority(n):
                if pm.get('last_shift', {}).get(n) == DA_YE:
                    return 0  # 大夜强制休，不退
                elif pm.get('work_streak', {}).get(n, 0) >= 5:
                    return 1  # 连上≥5，次优先退
                return 2  # 轮转休息，优先退
            day1_resting = sorted([n for n in names if 1 in rest_days[n]],
                                 key=_fx_rest_priority, reverse=True)
            for n in day1_resting:
                if _fx_rest_priority(n) == 0:
                    continue
                if len([x for x in names if 1 not in rest_days[x]]) >= 2:
                    break
                rest_days[n].discard(1)
                forced_fx[n].discard(1)

        # 跨月连上≥5但Day1被回退的，Day2必须补休（前提：不影响Day2≥2人在岗）
        for name in names:
            if pm.get('work_streak', {}).get(name, 0) >= 5 and 1 not in rest_days[name]:
                day2_working_now = [n for n in names if 2 not in rest_days[n]]
                if len(day2_working_now) >= 3:  # 此人加休后至少还有2人
                    rest_days[name].add(2)
                    forced_fx[name].add(2)

        # 跨月早期交换：对连上≥4的人，在第一个休息日前就近换休，避免超长连上
        for name in names:
            pm_ws = pm.get('work_streak', {}).get(name, 0)
            if pm_ws < 4:
                continue
            first_rest = None
            for d in range(1, self.num_days + 1):
                if d in rest_days[name]:
                    first_rest = d
                    break
            if first_rest and first_rest > 2:
                for swap_day in range(2, min(first_rest, 8)):
                    if swap_day in forced_fx.get(name, set()):
                        continue
                    resters = [n for n in names if swap_day in rest_days[n]
                               and swap_day not in forced_fx.get(n, set())]
                    for r in resters:
                        pm_r_ws = pm.get('work_streak', {}).get(r, 0)
                        if pm_r_ws >= pm_ws:
                            continue  # 对方连上不比我低，换了没意义
                        rest_days[name].add(swap_day)
                        rest_days[r].discard(swap_day)
                        break
                    if swap_day in rest_days[name]:
                        break

        rest_target = self.num_days - sum(1 for d in range(1, self.num_days + 1) if self.is_workday(d))
        for name in names:
            to_remove = len(rest_days[name]) - rest_target
            for _ in range(to_remove):
                blocks = self._find_rest_blocks(rest_days[name])
                best_day, best_score = None, -1
                for blk in blocks:
                    for d in blk:
                        if d in forced_fx.get(name, set()):
                            continue
                        test_rest = rest_days[name] - {d}
                        wm = [day not in test_rest for day in range(1, self.num_days + 1)]
                        mw = self._max_consecutive(wm)
                        score = 100 - mw * 10 + (5 if len(blk) == 1 else 0)
                        if mw <= 5:
                            score += 50
                        if score > best_score:
                            best_score, best_day = score, d
                if best_day:
                    rest_days[name].remove(best_day)
            while len(rest_days[name]) < rest_target:
                best_d, best_score = None, -1
                for d in range(self.num_days, 0, -1):
                    if d in rest_days[name]:
                        continue
                    test_rest = rest_days[name] | {d}
                    wm = [day not in test_rest for day in range(1, self.num_days + 1)]
                    rm = [day in test_rest for day in range(1, self.num_days + 1)]
                    if self._max_consecutive(wm) <= 5:
                        rest_days[name].add(d)
                        best_d = None
                        break
                    mw = self._max_consecutive(wm)
                    mr = self._max_consecutive(rm)
                    score = 200 - mw * 20 - mr * 30
                    if score > best_score:
                        best_score, best_d = score, d
                if best_d and best_score > -1:
                    rest_days[name].add(best_d)
                else:
                    break

        # 3人全上班的天 → 连上最久的人休息（机动转给在线组）
        ws = {n: pm.get('work_streak', {}).get(n, 0) for n in names}
        for day in range(1, self.num_days + 1):
            working_today = [n for n in names if day not in rest_days[n]]
            if len(working_today) == 3:
                high_w = max(working_today, key=lambda n: ws[n])
                rest_days[high_w].add(day)
                ws[high_w] = 0
                for n in working_today:
                    if n != high_w:
                        ws[n] += 1
                for rd in range(self.num_days, 0, -1):
                    if rd in rest_days[high_w] and rd not in forced_fx.get(high_w, set()):
                        rest_days[high_w].discard(rd)
                        break
            else:
                for n in names:
                    ws[n] = 0 if day in rest_days[n] else ws[n] + 1

        # 跨月连上最终检查：确保跨月合并后无人超标
        for name in names:
            pm_ws = pm.get('work_streak', {}).get(name, 0)
            if pm_ws == 0:
                continue
            wm = [day not in rest_days[name] for day in range(1, self.num_days + 1)]
            max_allowed = 6 if name in self.rest_requests else 5
            cur = pm_ws
            for i, w in enumerate(wm):
                if w:
                    cur += 1
                    if cur > max_allowed:
                        need_day = i + 1
                        if need_day not in rest_days[name]:
                            # 加休前检查：当天是否还有≥2人上班
                            day_working = [n for n in names if need_day not in rest_days[n]]
                            if len(day_working) >= 3:  # 此人加休后至少还有2人
                                rest_days[name].add(need_day)
                                # 从月底移除一个非强制休息日以保持总数
                                for rd in range(self.num_days, 0, -1):
                                    if rd in rest_days[name] and rd not in forced_fx.get(name, set()):
                                        rest_days[name].discard(rd)
                                        break
                        cur = 0
                else:
                    cur = 0

        schedules = {n: [XI] * self.num_days for n in names}
        zao_im_cnt = {n: 0 for n in names}
        wan_im_cnt = {n: 0 for n in names}
        jidong_cnt = {n: 0 for n in names}

        for day in range(1, self.num_days + 1):
            resting = [n for n in names if day in rest_days[n]]
            working = [n for n in names if day not in rest_days[n]]
            prev_shifts = {n: schedules[n][day - 2] if day > 1 else pm.get('last_shift', {}).get(n) for n in names}

            if len(working) == 0:
                continue  # 所有分销人员都在休息，这天不安排
            if len(resting) == 1:
                r = resting[0]
                schedules[r][day - 1] = XI
                if zao_im_cnt[working[0]] <= zao_im_cnt[working[1]]:
                    zao_cand, wan_cand = working[0], working[1]
                else:
                    zao_cand, wan_cand = working[1], working[0]
                if prev_shifts[zao_cand] == WAN_YI_IM:
                    zao_cand, wan_cand = wan_cand, zao_cand
                schedules[zao_cand][day - 1] = ZAO_ZAO_IM
                schedules[wan_cand][day - 1] = WAN_YI_IM
                zao_im_cnt[zao_cand] += 1
                wan_im_cnt[wan_cand] += 1
            elif len(working) >= 2:
                offset_rot = day % 3
                rotated = working[offset_rot:] + working[:offset_rot]
                rotated.sort(key=lambda n: jidong_cnt[n])
                jidong_name = rotated[0]
                others = [n for n in working if n != jidong_name]
                if prev_shifts[jidong_name] == WAN_YI_IM:
                    for alt in others:
                        if prev_shifts[alt] != WAN_YI_IM:
                            jidong_name = alt
                            others = [n for n in working if n != jidong_name]
                            break
                if zao_im_cnt[others[0]] <= zao_im_cnt[others[1]]:
                    zao_cand, wan_cand = others[0], others[1]
                else:
                    zao_cand, wan_cand = others[1], others[0]
                if prev_shifts[zao_cand] == WAN_YI_IM:
                    zao_cand, wan_cand = wan_cand, zao_cand
                schedules[jidong_name][day - 1] = XING_ZHENG_JD
                schedules[zao_cand][day - 1] = ZAO_ZAO_IM
                schedules[wan_cand][day - 1] = WAN_YI_IM
                jidong_cnt[jidong_name] += 1
                zao_im_cnt[zao_cand] += 1
                wan_im_cnt[wan_cand] += 1

        return schedules, rest_days

    # ── 3. 专职夜班排班 ──

    def _generate_night_shift(self):
        pm = self.prev_month_data
        schedules = {p['name']: [] for p in self.night_shift}
        prev_roles = pm.get('roles', {})

        # 第一轮：计算延顺人员的起始相位，记录每组相位
        group_phase = {}  # 'A'/'B' → starting_phase
        person_phase = {}  # name → starting_phase
        for p in self.night_shift:
            name = p['name']
            grp = p.get('group', 'A')
            if prev_roles.get(name) == '专职大夜':
                last = pm.get('last_shift', {}).get(name)
                second = pm.get('second_last_shift', {}).get(name)
                if last == DA_YE and second == XI:
                    sp = 1
                elif last == DA_YE and second == DA_YE:
                    sp = 2
                elif last == XI and second == DA_YE:
                    sp = 3
                elif last == XI and second == XI:
                    sp = 0
                else:
                    sp = 0 if grp == 'B' else 2
                person_phase[name] = sp
                group_phase[grp] = sp  # 同组延顺人员会覆盖，取最后一个（同组内应该一致）

        # 第二轮：为每个人生成排班，新人跟同组相位对齐
        for p in self.night_shift:
            name = p['name']
            grp = p.get('group', 'A')
            if name in person_phase:
                sp = person_phase[name]
            elif grp in group_phase:
                sp = group_phase[grp]  # 新人跟同组延顺人员对齐
            else:
                # 该组无延顺人员，用默认相位
                last = pm.get('last_shift', {}).get(name)
                if last == DA_YE:
                    sp = 2  # 大夜后休息两天
                else:
                    sp = 0 if grp == 'B' else 2
            for day in range(1, self.num_days + 1):
                phase = (sp + day - 1) % 4
                s = DA_YE if phase < 2 else XI
                schedules[name].append(s)
        return schedules

    # ── 4. 行政班排班 ──

    def _generate_buru_ban(self):
        schedules = {}
        for p in self.buru_ban:
            s = []
            for day in range(1, self.num_days + 1):
                s.append(XING_ZHENG if self.is_workday(day) else XI)
            schedules[p['name']] = s
        support_schedules = {}
        for p in self.buru_support:
            s = []
            for day in range(1, self.num_days + 1):
                s.append(XING_ZHENG if self.is_workday(day) else XI)
            support_schedules[p['name']] = s
        # 跨月强制Day1休息
        pm = self.prev_month_data
        for sched_dict in (schedules, support_schedules):
            for name in sched_dict:
                if pm.get('last_shift', {}).get(name) == DA_YE:
                    sched_dict[name][0] = XI
                elif pm.get('work_streak', {}).get(name, 0) >= 5:
                    sched_dict[name][0] = XI
        return schedules, support_schedules

    # ── 5. 长病假排班 ──

    def _generate_changbingjia(self):
        schedules = {}
        for p in self.changbingjia:
            schedules[p['name']] = [XI] * self.num_days
        return schedules

    # ── 6. 在线轮转排班 ──

    @staticmethod
    def _pick_balanced_shift(pool, shift_type_counts, name, rng):
        """从班次池中选择该员工做得最少的班次类型，尽量均衡分配"""
        if not pool:
            return None
        if shift_type_counts is None:
            # 无均衡数据时回退到随机 pop
            return pool.pop()
        person_counts = shift_type_counts.get(name, {})
        unique_types = list(set(pool))
        rng.shuffle(unique_types)
        unique_types.sort(key=lambda s: person_counts.get(s, 0))
        chosen = unique_types[0]
        pool.remove(chosen)
        person_counts[chosen] = person_counts.get(chosen, 0) + 1
        shift_type_counts[name] = person_counts
        return chosen

    def _assign_shifts_for_day(self, working, day, prev_shift, rest_days, fenxiao_rest_days,
                                schedules, balance=None, jd_count=None, wan2_count=None,
                                zaozao2_count=None, shift_type_counts=None):
        rng_day = random.Random(day * 137 + 42)
        n_workers = len(working)

        base_shifts_wd = {ZAO_BAN: 3, ZAO_SAN: 2, ZAO_ZAO_2: 1, XING_ZHENG: 1,
                          BAI_BAN: 2, ZHONG_SAN: 1, ZHONG_SI: 1, WAN_YI: 4, WAN_ER: 2}
        base_shifts_we = {ZAO_BAN: 3, ZAO_SAN: 2, ZAO_ZAO_2: 1, XING_ZHENG: 1,
                          BAI_BAN: 1, ZHONG_SAN: 1, ZHONG_SI: 1, WAN_YI: 3, WAN_ER: 2}

        base = base_shifts_wd if self.is_workday(day) else base_shifts_we
        base_total = sum(base.values())
        day_shifts = []
        for shift, count in base.items():
            scaled = max(1, round(count * n_workers / base_total))
            day_shifts.extend([shift] * scaled)

        while len(day_shifts) > n_workers:
            removed = False
            for s in [XING_ZHENG, BAI_BAN, WAN_YI, ZAO_BAN,
                      ZAO_SAN, ZHONG_SAN, ZHONG_SI, ZAO_ZAO_2, WAN_ER]:
                if s in day_shifts:
                    day_shifts.remove(s)
                    removed = True
                    break
            if not removed:
                day_shifts = day_shifts[:n_workers]
                break
        while len(day_shifts) < n_workers:
            day_shifts.append(XING_ZHENG)

        late_pool = [s for s in day_shifts if s in LATE_SHIFTS]
        early_pool = [s for s in day_shifts if s not in LATE_SHIFTS]

        restricted = [p for p in working if prev_shift.get(p['name']) in LATE_SHIFTS]
        free = [p for p in working if prev_shift.get(p['name']) not in LATE_SHIFTS]

        rng_day.shuffle(restricted)
        rng_day.shuffle(late_pool)
        rng_day.shuffle(early_pool)

        if balance:
            def _late_ratio(name):
                b = balance.get(name, {'early': 0, 'late': 0})
                return b['late'] / max(1, b['early'] + b['late'])
            free.sort(key=lambda p: _late_ratio(p['name']))
        else:
            rng_day.shuffle(free)

        assignments = {}
        for p in restricted:
            if late_pool:
                assignments[p['name']] = self._pick_balanced_shift(late_pool, shift_type_counts, p['name'], rng_day)
            else:
                free.append(p)

        n_late = min(len(late_pool), len(free))
        for p in free[:n_late]:
            if late_pool:
                assignments[p['name']] = self._pick_balanced_shift(late_pool, shift_type_counts, p['name'], rng_day)

        for p in free[n_late:]:
            if early_pool:
                assignments[p['name']] = self._pick_balanced_shift(early_pool, shift_type_counts, p['name'], rng_day)
            elif late_pool:
                assignments[p['name']] = self._pick_balanced_shift(late_pool, shift_type_counts, p['name'], rng_day)

        if wan2_count is not None:
            late_assigned = [(p, assignments[p['name']]) for p in working
                             if p['name'] in assignments and assignments[p['name']] in LATE_SHIFTS]
            wan_er_names = set(p['name'] for p, s in late_assigned if s == WAN_ER)
            non_wan_er = [(p, s) for p, s in late_assigned if s != WAN_ER]
            for wp_name in list(wan_er_names):
                wp = next(p for p in working if p['name'] == wp_name)
                better = [(p, s) for p, s in non_wan_er
                          if wan2_count[p['name']] < wan2_count[wp_name]]
                if better:
                    swap_p, swap_s = min(better, key=lambda x: wan2_count[x[0]['name']])
                    assignments[wp_name] = swap_s
                    assignments[swap_p['name']] = WAN_ER
                    non_wan_er = [(p, s) for p, s in non_wan_er if p['name'] != swap_p['name']]
                    non_wan_er.append((wp, swap_s))
                    wan_er_names.discard(wp_name)
                    wan_er_names.add(swap_p['name'])
            for p in working:
                if assignments.get(p['name']) == WAN_ER:
                    wan2_count[p['name']] += 1

        # 早早2均衡：与晚二类似的交换逻辑
        if zaozao2_count is not None:
            early_assigned = [(p, assignments[p['name']]) for p in working
                             if p['name'] in assignments and assignments[p['name']] in EARLY_SHIFTS]
            zaozao2_names = set(p['name'] for p, s in early_assigned if s == ZAO_ZAO_2)
            non_zaozao2 = [(p, s) for p, s in early_assigned if s != ZAO_ZAO_2]
            for zp_name in list(zaozao2_names):
                zp = next(p for p in working if p['name'] == zp_name)
                better = [(p, s) for p, s in non_zaozao2
                          if zaozao2_count[p['name']] < zaozao2_count[zp_name]]
                if better:
                    swap_p, swap_s = min(better, key=lambda x: zaozao2_count[x[0]['name']])
                    assignments[zp_name] = swap_s
                    assignments[swap_p['name']] = ZAO_ZAO_2
                    non_zaozao2 = [(p, s) for p, s in non_zaozao2 if p['name'] != swap_p['name']]
                    non_zaozao2.append((zp, swap_s))
                    zaozao2_names.discard(zp_name)
                    zaozao2_names.add(swap_p['name'])
            for p in working:
                if assignments.get(p['name']) == ZAO_ZAO_2:
                    zaozao2_count[p['name']] += 1

        unassigned = [p for p in working if p['name'] not in assignments]
        leftover = early_pool + late_pool
        for p in unassigned:
            if leftover:
                assignments[p['name']] = self._pick_balanced_shift(leftover, shift_type_counts, p['name'], rng_day)
            else:
                assignments[p['name']] = XING_ZHENG

        fenxiao_all_work = False
        if fenxiao_rest_days:
            fx_names = [p['name'] for p in self.fenxiao]
            wc = sum(1 for n in fx_names
                     if day not in fenxiao_rest_days.get(n, set()))
            fenxiao_all_work = (wc == 3)

        # 工作日周妙笛固定上行政班（机动），轮转人员不再需要
        if not fenxiao_all_work and not self.is_workday(day):
            candidates = [p for p in working
                          if p['name'] in assignments
                          and prev_shift.get(p['name']) not in LATE_SHIFTS
                          and assignments[p['name']] in EARLY_SHIFTS]
            if candidates:
                if jd_count is not None:
                    min_jd = min(jd_count[p['name']] for p in candidates)
                    candidates = [p for p in candidates if jd_count[p['name']] == min_jd]
                    chosen = rng_day.choice(candidates)
                    jd_count[chosen['name']] += 1
                else:
                    chosen = rng_day.choice(candidates)
                assignments[chosen['name']] = XING_ZHENG_JD

        for p in working:
            if p['name'] in assignments:
                schedules[p['name']][day - 1] = assignments[p['name']]
            else:
                schedules[p['name']][day - 1] = XING_ZHENG

        if balance:
            for p in working:
                name = p['name']
                shift = assignments.get(name, XING_ZHENG)
                if shift in LATE_SHIFTS:
                    balance[name]['late'] += 1
                else:
                    balance[name]['early'] += 1

    def _daily_workers(self, rest_days):
        return [sum(1 for p in self.online_staff if d not in rest_days[p['name']])
                for d in range(1, self.num_days + 1)]

    def _remove_rest_from_overstaffed(self, name, rest_days, forced_rest_days, effective_target):
        """从超员日优先移除一个非强制休息日，不破坏连上约束。返回 True 表示成功移除。"""
        dw = self._daily_workers(rest_days)
        # 找该人的非强制休息日，优先选超员最多的
        candidates = []
        for d in rest_days[name]:
            if d in forced_rest_days.get(name, set()):
                continue
            over = dw[d - 1] - effective_target[d - 1]
            if over > 0:
                candidates.append((over, d))
        if candidates:
            candidates.sort(key=lambda x: -x[0])
            for _, d in candidates:
                test_rest = rest_days[name] - {d}
                test_wm = [day not in test_rest for day in range(1, self.num_days + 1)]
                if self._max_consecutive(test_wm) <= 5:
                    rest_days[name].discard(d)
                    return True
        # 没有超员日可选 → 宁可保留连上超标，也不破坏偏差=0
        return False

    def _generate_online(self, fenxiao_rest_days=None):
        import time as _time
        _t0 = _time.time()
        def _tick(msg):
            nonlocal _t0
            t = _time.time()
            print(f"  [online] {msg} ({t - _t0:.2f}s)", flush=True)
            _t0 = t

        rng = random.Random(42)
        an_rng = random.Random(67890)

        # 周妙笛按行政班作息：工作日行政班（机动），非工作日休息
        zhou_name = '周妙笛'
        zhou_sched = None
        original_online_staff = self.online_staff
        if any(p['name'] == zhou_name for p in self.online_staff):
            zhou_sched = []
            for day in range(1, self.num_days + 1):
                zhou_sched.append(XING_ZHENG_JD if self.is_workday(day) else XI)
            pm = self.prev_month_data
            if pm.get('last_shift', {}).get(zhou_name) == DA_YE:
                zhou_sched[0] = XI
            elif pm.get('work_streak', {}).get(zhou_name, 0) >= 5:
                zhou_sched[0] = XI
            self.online_staff = [p for p in self.online_staff if p['name'] != zhou_name]

        rest_target = self.num_days - sum(1 for d in range(1, self.num_days + 1) if self.is_workday(d))
        n = len(self.online_staff)
        _tick(f"初始化: {n}人, {self.num_days}天, rest_target={rest_target}")

        # 计算有效目标（在线人员需要承担的人数）
        # Step 1: 人天容量平摊 — 如果目标总人天超过可提供人天，缺口均摊到每天
        workday_count = sum(1 for d in range(1, self.num_days + 1) if self.is_workday(d))
        buru_online_count = len(self.buru_ban)  # 正常行政班人数（支援行政班不计入）
        zhou_workdays = sum(1 for s in zhou_sched if s != XI) if zhou_sched else 0
        total_capacity = (n + buru_online_count) * workday_count + zhou_workdays
        total_target = sum(self.daily_targets)
        if total_target > total_capacity:
            scale = total_capacity / total_target  # 按比例缩放，保持每日忙闲比不变
            # 优先使用理论值（有小数精度，能保留每日忙闲差异）
            src = self.theoretical_targets
            if not src or len(src) != self.num_days:
                src = self.daily_targets
            capacity_adjusted = []
            err = 0.0
            for t in src:
                raw = t * scale + err
                rounded = max(1, round(raw))
                err = raw - rounded
                capacity_adjusted.append(rounded)
            # 修正四舍五入导致的总和偏差
            diff = total_capacity - sum(capacity_adjusted)
            if diff != 0:
                fracs = [(src[i] * scale - capacity_adjusted[i], i)
                         for i in range(self.num_days)]
                if diff > 0:
                    fracs.sort(key=lambda x: -x[0])
                    for i in range(diff):
                        capacity_adjusted[fracs[i][1]] += 1
                else:
                    fracs.sort(key=lambda x: x[0])
                    for i in range(-diff):
                        capacity_adjusted[fracs[i][1]] = max(1, capacity_adjusted[fracs[i][1]] - 1)
            print(f"  [diag] 容量按比例缩放: 目标{total_target} > 容量{total_capacity}, "
                  f"比例={scale:.3f}", flush=True)
            self._diag = {
                'capacity_adjusted': True,
                'total_target': total_target,
                'total_capacity': total_capacity,
                'scale': round(scale, 4),
            }
        else:
            capacity_adjusted = list(self.daily_targets)
            self._diag = {
                'capacity_adjusted': False,
                'total_target': total_target,
                'total_capacity': total_capacity,
            }
        self.capacity_adjusted = capacity_adjusted  # 保存供 verify 使用

        # Step 2: 扣除 buru_ban 每日贡献（工作日1人，非工作日0人）
        buru_working = [1 if self.is_workday(d) else 0 for d in range(1, self.num_days + 1)]
        effective_target = [max(0, capacity_adjusted[d - 1] - buru_working[d - 1])
                           for d in range(1, self.num_days + 1)]

        # Step 2b: 扣除周妙笛固定行政班贡献（工作日1人，非工作日0人）
        if zhou_sched:
            zhou_working = [1 if s != XI else 0 for s in zhou_sched]
            effective_target = [max(0, effective_target[d] - zhou_working[d])
                               for d in range(self.num_days)]

        # 补充诊断信息
        self._diag['n_online'] = n
        self._diag['workday_count'] = workday_count
        self._diag['buru_online_count'] = buru_online_count
        self._diag['buru_workdays'] = sum(buru_working)
        self._diag['effective_sum'] = sum(effective_target)
        self._diag['original_targets'] = list(self.daily_targets)
        self._diag['effective_targets'] = effective_target

        cum_rest = 0
        expected_rest = {}
        for d in range(1, self.num_days + 1):
            cum_rest += n - effective_target[d - 1]
            expected_rest[d] = cum_rest / n

        _tick("Phase 1 开始")
        # Phase 1: Greedy forward with progressive cap
        rest_days = {p['name']: set() for p in self.online_staff}
        pm = self.prev_month_data
        work_streak = {}
        rest_streak = {}
        daye_names = []
        forced_rest_days = {}  # name → set of days，跨月强制休息，后续阶段不可删除
        init_pm_ws = {}  # 月初初始连上（上月遗留），用于月初休息日分布软约束
        for p in self.online_staff:
            name = p['name']
            forced_rest_days[name] = set()
            work_streak[name] = pm.get('work_streak', {}).get(name, 0)
            init_pm_ws[name] = work_streak[name]
            rest_streak[name] = pm.get('rest_streak', {}).get(name, 0)
            # 上月最后一天是大夜 → 至少休1天
            if pm.get('last_shift', {}).get(name) == DA_YE:
                rest_days[name].add(1)
                forced_rest_days[name].add(1)
                work_streak[name] = 0
                rest_streak[name] = 1
                daye_names.append(name)
            # 跨月连上≥5 → Day 1 强制休
            elif work_streak[name] >= 5:
                rest_days[name].add(1)
                forced_rest_days[name].add(1)
                work_streak[name] = 0
                rest_streak[name] = 1

        # 大夜人员优先休2天：day 2 产能允许则加休
        if daye_names:
            d2_max_rest = n - effective_target[1]  # day 2 最多可休息人数
            d2_already = sum(1 for p in self.online_staff if 2 in rest_days[p['name']])
            if d2_already + len(daye_names) <= d2_max_rest:
                for name in daye_names:
                    rest_days[name].add(2)
                    forced_rest_days[name].add(2)

        for day in range(1, self.num_days + 1):
            target_rest = n - effective_target[day - 1]
            max_rest_by_now = min(rest_target, int(expected_rest[day]) + 2)

            scored = []
            for p in self.online_staff:
                name = p['name']
                if len(rest_days[name]) >= rest_target:
                    continue
                if rest_streak[name] >= 2:
                    continue
                if len(rest_days[name]) >= max_rest_by_now:
                    continue

                wstreak = work_streak[name]
                must_rest = wstreak >= 5
                deficit = expected_rest[day] - len(rest_days[name])
                block_bonus = 200 if (day > 1 and (day - 1) in rest_days[name]) else 0
                ahead = len(rest_days[name]) - expected_rest[day]
                ahead_penalty = int(max(0, ahead - 0.0) * 500)

                score = (100000 if must_rest else 0)
                score += int(deficit * 400)
                if wstreak >= 4:
                    score += (2 ** wstreak) * 10
                else:
                    score += wstreak * 40
                score += block_bonus
                score -= ahead_penalty
                # 月初分布软偏好：上月连上高的人（跨月风险高），月初优先获得休息日
                if init_pm_ws.get(name, 0) >= 3 and day <= 5 and 1 not in forced_rest_days.get(name, set()):
                    score += 80
                score += rng.randint(0, 20)
                scored.append((score, name))

            scored.sort(key=lambda x: -x[0])

            day_rest = set()
            for _, name in scored:
                if len(day_rest) >= target_rest:
                    break
                day_rest.add(name)

            # 补足：如果打分后仍不足目标休息人数
            if len(day_rest) < target_rest:
                remaining = sorted(
                    [p for p in self.online_staff
                     if p['name'] not in day_rest
                     and len(rest_days[p['name']]) < rest_target
                     and rest_streak[p['name']] < 2
                     and len(rest_days[p['name']]) < max_rest_by_now],
                    key=lambda p: work_streak[p['name']],
                    reverse=True
                )
                for p in remaining:
                    if len(day_rest) >= target_rest:
                        break
                    day_rest.add(p['name'])

            for p in self.online_staff:
                name = p['name']
                if name in day_rest or day in rest_days[name]:
                    rest_days[name].add(day)
                    work_streak[name] = 0
                    rest_streak[name] += 1
                else:
                    work_streak[name] += 1
                    rest_streak[name] = 0

        # Phase 1.5: 修复跨月连上超标 — 用 1:1 交换替代插入+补偿
        for p in self.online_staff:
            name = p['name']
            pm_ws = pm.get('work_streak', {}).get(name, 0)
            if pm_ws == 0:
                continue
            wm = [d not in rest_days[name] for d in range(1, self.num_days + 1)]
            cur = pm_ws
            for i, w in enumerate(wm):
                if w:
                    cur += 1
                    day = i + 1
                    if cur > 5:
                        # 跨月连上超标，尝试 1:1 交换
                        need_day = day
                        s_non_forced = sorted([d for d in rest_days[name]
                                               if d not in forced_rest_days.get(name, set())],
                                              reverse=True)
                        found = False
                        for fname in sorted([p2['name'] for p2 in self.online_staff if p2['name'] != name],
                                           key=lambda n: pm.get('work_streak', {}).get(n, 0)):
                            f_rests = rest_days[fname]
                            f_pm = pm.get('work_streak', {}).get(fname, 0)
                            for early_day in range(1, need_day + 1):
                                if found: break
                                if early_day not in f_rests: continue
                                if early_day in forced_rest_days.get(fname, set()): continue
                                # 预检：fname 交出 early_day 后月内是否超标
                                tf_pre = f_rests - {early_day}
                                tf_pre_wm = [d not in tf_pre for d in range(1, self.num_days + 1)]
                                if self._max_consecutive(tf_pre_wm) > 5:
                                    continue
                                for give_day in s_non_forced:
                                    if found: break
                                    if give_day <= early_day: continue
                                    if give_day in f_rests: continue
                                    if give_day in wm and not wm[give_day - 1]: continue  # f 这天已休息
                                    f_wm_full = [d not in f_rests for d in range(1, self.num_days + 1)]
                                    if not f_wm_full[give_day - 1]: continue
                                    # 尝试交换
                                    ts = (rest_days[name] | {early_day}) - {give_day}
                                    tf = (f_rests | {give_day}) - {early_day}
                                    ts_wm = [d not in ts for d in range(1, self.num_days + 1)]
                                    tf_wm = [d not in tf for d in range(1, self.num_days + 1)]
                                    ts_rm = [d in ts for d in range(1, self.num_days + 1)]
                                    tf_rm = [d in tf for d in range(1, self.num_days + 1)]
                                    s_cons = sum(1 for w2 in ts_wm if w2)
                                    s_start_cons = 0
                                    for w2 in ts_wm:
                                        if w2: s_start_cons += 1
                                        else: break
                                    f_start_cons = 0
                                    for w2 in tf_wm:
                                        if w2: f_start_cons += 1
                                        else: break
                                    new_s_cross = pm_ws + s_start_cons
                                    new_f_cross = f_pm + f_start_cons if f_pm > 0 else 0
                                    if (self._max_consecutive(ts_wm) <= 5 and self._max_consecutive(ts_rm) <= 2 and
                                        self._max_consecutive(tf_wm) <= 5 and self._max_consecutive(tf_rm) <= 2 and
                                        new_s_cross <= 5 and new_f_cross <= 5):
                                        rest_days[name] = ts
                                        rest_days[fname] = tf
                                        forced_rest_days.setdefault(name, set()).add(early_day)
                                        found = True
                        if found:
                            # 已修复此人的跨月连上，重新计算 wm 继续检查
                            wm = [d not in rest_days[name] for d in range(1, self.num_days + 1)]
                            cur = pm_ws
                            for i2, w2 in enumerate(wm):
                                if w2:
                                    cur += 1
                                    if cur > 5:
                                        pass  # 重新进入修复流程
                                else:
                                    cur = 0
                        cur = 0  # 当前超标已处理（无论成功与否）
                else:
                    cur = 0

        for p in self.online_staff:
            name = p['name']
            wm = [d not in rest_days[name] for d in range(1, self.num_days + 1)]
            cur = 0
            worst_start, worst_end, worst_len = 0, 0, 0
            worst_severity = 0
            for i, w in enumerate(wm):
                if w:
                    cur += 1
                else:
                    cur = 0
                day = i + 1
                max_allowed = 5
                if cur > max_allowed:
                    severity = cur - max_allowed
                    if severity > worst_severity:
                        worst_severity = severity
                        worst_start = day - cur + 1
                        worst_end = day
                        worst_len = cur
            if worst_severity == 0:
                continue
            rest_before = [d for d in sorted(rest_days[name])
                          if d < worst_start and d not in forced_rest_days.get(name, set())]
            if rest_before:
                rest_days[name].remove(rest_before[-1])

        _tick("Phase 1 完成, Phase 2")
        # 诊断：Phase 1 后 day 12 状态
        dw_diag = self._daily_workers(rest_days)
        print(f"  [diag] Phase1后 day12 到岗={dw_diag[11]}, 目标={effective_target[11]}, "
              f"休息={sum(1 for rd in rest_days.values() if 12 in rd)}", flush=True)
        # Phase 2: Fix rest counts
        dw_all = self._daily_workers(rest_days)
        for p in self.online_staff:
            name = p['name']
            # 删除多余休息：优先删缺员日（删休息=加人，补到缺员日）
            while len(rest_days[name]) > rest_target:
                blocks = self._find_rest_blocks(rest_days[name])
                candidates = []
                for blk in blocks:
                    for d in blk:
                        if d in forced_rest_days.get(name, set()):
                            continue
                        test_rest = rest_days[name] - {d}
                        wm = [day not in test_rest for day in range(1, self.num_days + 1)]
                        mw = self._max_consecutive(wm)
                        gap = effective_target[d - 1] - dw_all[d - 1]
                        candidates.append((-gap, mw, d))
                if candidates:
                    candidates.sort()
                    d_sel = candidates[0][2]
                    rest_days[name].remove(d_sel)
                    dw_all[d_sel - 1] += 1  # 少一人休息 = 多一人到岗

            # 动态计算单日缺口上限（基于全月容量均摊）
            workday_cnt = self.num_days - rest_target
            total_shortage = max(0, sum(effective_target) - n * workday_cnt)
            min_gap = -math.ceil(total_shortage / self.num_days) if total_shortage > 0 else -1

            while len(rest_days[name]) < rest_target:
                candidates = []
                for d in range(1, self.num_days + 1):
                    if d in rest_days[name]:
                        continue
                    gap = dw_all[d - 1] - effective_target[d - 1]
                    if gap <= min_gap:
                        continue  # 该日缺口已达容量上限，不再恶化
                    test_rest = rest_days[name] | {d}
                    wm = [day not in test_rest for day in range(1, self.num_days + 1)]
                    mw = self._max_consecutive(wm)
                    max_mw = 6
                    mw_penalty = max(0, mw - max_mw)
                    score = (mw_penalty, -gap, mw, d)
                    candidates.append(score)
                if not candidates:
                    break  # 无日可选，放弃补充此人的休息
                candidates.sort()
                d_sel = candidates[0][3]
                gap_sel = dw_all[d_sel - 1] - effective_target[d_sel - 1]
                if gap_sel < 0:
                    print(f"  [diag] Phase2 {name} 被迫在 day{d_sel} 加休息 "
                          f"(gap={gap_sel}, mw={candidates[0][2]}, penalty={candidates[0][0]})", flush=True)
                rest_days[name].add(d_sel)
                dw_all[d_sel - 1] -= 1

        # Phase 2.5: Push rest into last week
        w5_start = self.num_days - 4
        w5_days = list(range(w5_start, self.num_days + 1))
        for p in self.online_staff:
            name = p['name']
            w5_rest = len(rest_days[name] & set(w5_days))
            if w5_rest > 0:
                continue
            early_rest = sorted((rest_days[name] - set(w5_days)) - forced_rest_days.get(name, set()))
            if not early_rest:
                continue
            last_rest = early_rest[-1]
            if last_rest <= w5_start - 2:
                for w5d in [w5_start + 1, w5_start + 2, w5_start, w5_start + 3, w5_start + 4]:
                    if w5d > self.num_days:
                        continue
                    if w5d in rest_days[name]:
                        continue
                    current_rest_w5d = sum(1 for pp in self.online_staff if w5d in rest_days[pp['name']])
                    target_rest_w5d = n - effective_target[w5d - 1]
                    if current_rest_w5d < target_rest_w5d + 3:
                        test_rest = (rest_days[name] - {last_rest}) | {w5d}
                        rm = [d in test_rest for d in range(1, self.num_days + 1)]
                        wm = [d not in test_rest for d in range(1, self.num_days + 1)]
                        if self._max_consecutive(wm) <= 6:
                            rest_days[name].remove(last_rest)
                            rest_days[name].add(w5d)
                            break

        _tick("Phase 2 完成, Phase 3")
        # 诊断：Phase 2 后 day 12 状态
        dw_diag2 = self._daily_workers(rest_days)
        print(f"  [diag] Phase2后 day12 到岗={dw_diag2[11]}, 目标={effective_target[11]}, "
              f"休息={sum(1 for rd in rest_days.values() if 12 in rd)}", flush=True)
        # Phase 3: Target-matching swaps
        for _ in range(800):
            dw = self._daily_workers(rest_days)
            gaps = [(d, dw[d - 1] - effective_target[d - 1]) for d in range(1, self.num_days + 1)]
            overs = [(d, g) for d, g in gaps if g >= 2]
            unders = [(d, g) for d, g in gaps if g <= -2]
            if not overs or not unders:
                break
            overs.sort(key=lambda x: -x[1])
            unders.sort(key=lambda x: x[1])
            improved = False
            for over_d, _ in overs:
                for under_d, _ in unders:
                    if improved:
                        break
                    for p in self.online_staff:
                        name = p['name']
                        if under_d in rest_days[name] and over_d not in rest_days[name]:
                            if under_d in forced_rest_days.get(name, set()):
                                continue
                            test_rest = (rest_days[name] - {under_d}) | {over_d}
                            wm = [d not in test_rest for d in range(1, self.num_days + 1)]
                            rm = [d in test_rest for d in range(1, self.num_days + 1)]
                            mw, mr = self._max_consecutive(wm), self._max_consecutive(rm)
                            max_mw = 6
                            if mw <= max_mw and mr <= 2:
                                rest_days[name].remove(under_d)
                                rest_days[name].add(over_d)
                                improved = True
                                break
            if not improved:
                # 两人回退：一人减超员日到岗，另一人补缺员日到岗
                for over_d, _ in overs:
                    if improved:
                        break
                    for under_d, _ in unders:
                        if improved:
                            break
                        over_candidates = sorted(
                            [p for p in self.online_staff if over_d not in rest_days[p['name']]],
                            key=lambda p: -self._max_consecutive(
                                [d not in rest_days[p['name']] for d in range(1, self.num_days + 1)]))
                        under_candidates = sorted(
                            [p for p in self.online_staff if under_d in rest_days[p['name']]],
                            key=lambda p: self._max_consecutive(
                                [d not in rest_days[p['name']] for d in range(1, self.num_days + 1)]))
                        for p_over in over_candidates:
                            if improved:
                                break
                            test_over = rest_days[p_over['name']] | {over_d}
                            rm_over = [d in test_over for d in range(1, self.num_days + 1)]
                            if self._max_consecutive(rm_over) > 2:
                                continue
                            for p_under in under_candidates:
                                if p_under['name'] == p_over['name']:
                                    continue
                                test_under = rest_days[p_under['name']] - {under_d}
                                wm_under = [d not in test_under for d in range(1, self.num_days + 1)]
                                max_mw_u = 7
                                if self._max_consecutive(wm_under) <= max_mw_u:
                                    rest_days[p_over['name']].add(over_d)
                                    rest_days[p_under['name']].remove(under_d)
                                    improved = True
                                    break
                if not improved:
                    break

        _tick("Phase 3 完成, Phase 4 (SA)")
        # Phase 4: Simulated annealing
        def quick_energy(rd):
            e = 0
            dw = self._daily_workers(rd)
            for d in range(1, self.num_days + 1):
                diff = dw[d - 1] - effective_target[d - 1]
                e += abs(diff) * 500
            for p in self.online_staff:
                name = p['name']
                wm = [d not in rd[name] for d in range(1, self.num_days + 1)]
                mw = self._max_consecutive(wm)
                # 跨月合并连上：上月结转 + 本月月初连续工作
                pm_ws = pm.get('work_streak', {}).get(name, 0)
                cross_mw = mw
                if pm_ws > 0:
                    cons_day1 = 0
                    for w in wm:
                        if w:
                            cons_day1 += 1
                        else:
                            break
                    cross_mw = max(mw, pm_ws + cons_day1)
                # 用跨月合并值做惩罚，促使连上超标在所有人之间均匀分摊
                if cross_mw > 5:
                    excess = cross_mw - 5
                    e += excess * excess * 3000
                if len(rd[name]) != rest_target:
                    e += abs(len(rd[name]) - rest_target) * 10000
            return e

        current_energy = quick_energy(rest_days)
        best_energy = current_energy
        best_rest = {name: set(s) for name, s in rest_days.items()}

        T = 2000.0
        people_names = [p['name'] for p in self.online_staff]

        for it in range(40000):
            p1 = an_rng.choice(people_names)
            if not rest_days[p1]:
                continue
            d1 = an_rng.choice(sorted(rest_days[p1]))
            p2 = an_rng.choice(people_names)
            if p2 == p1:
                continue
            work_days = [d for d in range(1, self.num_days + 1) if d not in rest_days[p2]]
            if not work_days:
                continue
            if d1 in forced_rest_days.get(p1, set()):
                continue  # 不移动跨月强制休息日
            d2 = an_rng.choice(work_days)
            if d1 == d2 or d2 in rest_days[p1] or d1 in rest_days[p2]:
                continue

            new_p1 = (rest_days[p1] - {d1}) | {d2}
            new_p2 = rest_days[p2] | {d1}

            max_mw_1 = 6
            max_mw_2 = 6
            if (self._max_consecutive([d not in new_p1 for d in range(1, self.num_days + 1)]) > max_mw_1 or
                    self._max_consecutive([d not in new_p2 for d in range(1, self.num_days + 1)]) > max_mw_2):
                continue
            if len(new_p2) > rest_target:
                continue

            new_rd = {name: set(s) for name, s in rest_days.items()}
            new_rd[p1] = new_p1
            new_rd[p2] = new_p2
            new_energy = quick_energy(new_rd)

            delta = new_energy - current_energy
            if delta < 0 or (T > 0.01 and an_rng.random() < 2.71828 ** (-delta / T)):
                rest_days[p1] = new_p1
                rest_days[p2] = new_p2
                current_energy = new_energy
                if current_energy < best_energy:
                    best_energy = current_energy
                    best_rest = {name: set(s) for name, s in rest_days.items()}

            T *= 0.9995
            if T < 1.0:
                T = 2000.0 * (0.5 + 0.5 * an_rng.random())
                if best_energy == 0:
                    break

        rest_days = best_rest

        _tick("Phase 4 (SA) 完成, Phase 4.5")
        # Phase 4.5: Direct maxWork repair
        for _ in range(500):
            worst_name = None
            worst_excess = 0
            for p in self.online_staff:
                name = p['name']
                wm = [d not in rest_days[name] for d in range(1, self.num_days + 1)]
                mw = self._max_consecutive(wm)
                # 跨月连上：上月结转 + 月初连续工作
                pm_ws = pm.get('work_streak', {}).get(name, 0)
                if pm_ws > 0:
                    cons_day1 = 0
                    for w in wm:
                        if w:
                            cons_day1 += 1
                        else:
                            break
                    mw = max(mw, pm_ws + cons_day1)
                cur = 0
                longest_end = 0
                for i, w in enumerate(wm):
                    if w:
                        cur += 1
                        if cur == mw:
                            longest_end = i + 1
                    else:
                        cur = 0
                threshold = 7 if longest_end >= self.num_days - 4 else 5
                excess = mw - threshold
                if excess > worst_excess:
                    worst_excess = excess
                    worst_name = name

            if worst_name is None or worst_excess <= 0:
                break

            wm = [d not in rest_days[worst_name] for d in range(1, self.num_days + 1)]
            mw = self._max_consecutive(wm)
            pm_ws_worst = pm.get('work_streak', {}).get(worst_name, 0)
            cur = 0
            longest_start = 1
            longest_end = 1
            if pm_ws_worst > 0 and wm and wm[0]:
                cons_day1 = 0
                for w in wm:
                    if w:
                        cons_day1 += 1
                    else:
                        break
                cross_mw = pm_ws_worst + cons_day1
                if cross_mw > mw:
                    mw = cross_mw
                    longest_start = 1
                    cur = pm_ws_worst
                    for i in range(cons_day1):
                        cur += 1
                        if cur == mw:
                            longest_end = i + 1
                            break
            if cur == 0:
                for i, w in enumerate(wm):
                    if w:
                        if cur == 0:
                            ss = i + 1
                        cur += 1
                        if cur == mw:
                            longest_start = ss
                            longest_end = i + 1
                    else:
                        cur = 0

            best_swap = None
            for offset_frac in [0.3, 0.5, 0.7]:
                target = int(longest_start + (longest_end - longest_start) * offset_frac)
                for pp in self.online_staff:
                    if pp['name'] == worst_name:
                        continue
                    if target not in rest_days[pp['name']]:
                        continue
                    if target in forced_rest_days.get(pp['name'], set()):
                        continue
                    for swap_out in sorted(rest_days[worst_name]):
                        if swap_out in forced_rest_days.get(worst_name, set()):
                            continue
                        if swap_out >= longest_start:
                            break
                        if swap_out in rest_days[pp['name']]:
                            continue
                        new_worst = (rest_days[worst_name] - {swap_out}) | {target}
                        new_pp = (rest_days[pp['name']] - {target}) | {swap_out}
                        ok = True
                        for nm, nr in [(worst_name, new_worst), (pp['name'], new_pp)]:
                            nwm = [d not in nr for d in range(1, self.num_days + 1)]
                            nrm = [d in nr for d in range(1, self.num_days + 1)]
                            nmw = self._max_consecutive(nwm)
                            nmr = self._max_consecutive(nrm)
                            cur2 = 0
                            ne = 0
                            for i, w in enumerate(nwm):
                                if w:
                                    cur2 += 1
                                    if cur2 == nmw:
                                        ne = i + 1
                                else:
                                    cur2 = 0
                            threshold2 = 7 if ne >= self.num_days - 4 else 5
                            if nmw > threshold2 or len(nr) != rest_target:
                                ok = False
                                break
                        if ok:
                            best_swap = (pp['name'], swap_out, target)
                            break
                    if best_swap:
                        break
                if best_swap:
                    break

            if best_swap:
                pp_name, swap_out, target = best_swap
                rest_days[worst_name].remove(swap_out)
                rest_days[worst_name].add(target)
                rest_days[pp_name].remove(target)
                rest_days[pp_name].add(swap_out)
            else:
                break

        _tick("Phase 4.5 完成, Phase 4.6")
        # Phase 4.6: W5 streak breaking via targeted swaps
        for _ in range(500):
            victims = []
            for p in self.online_staff:
                name = p['name']
                wm = [d not in rest_days[name] for d in range(1, self.num_days + 1)]
                mw = self._max_consecutive(wm)
                if mw <= 5:
                    continue
                cur = 0
                end = 0
                start = 1
                for i, w in enumerate(wm):
                    if w:
                        cur += 1
                    else:
                        cur = 0
                    if cur == mw:
                        end = i + 1
                if end >= self.num_days - 4:
                    cur2 = 0
                    for i in range(end - 1, -1, -1):
                        if wm[i]:
                            cur2 += 1
                            if cur2 == mw:
                                start = i + 1
                        else:
                            break
                    victims.append((mw, start, end, name))
            if not victims:
                break
            victims.sort(reverse=True)
            _, v_start, v_end, worst = victims[0]
            old_mw = victims[0][0]

            best_swap = None
            for swap_out in sorted(rest_days[worst], reverse=True):
                if swap_out in forced_rest_days.get(worst, set()):
                    continue
                if swap_out >= v_start:
                    continue
                if best_swap:
                    break
                target = v_start + (v_end - v_start) // 2
                for offset in [0, -1, 1, -2, 2]:
                    t = target + offset
                    if t < v_start or t > v_end:
                        continue
                    if t in rest_days[worst]:
                        continue
                    if best_swap:
                        break
                    for pp in self.online_staff:
                        if pp['name'] == worst:
                            continue
                        if t not in rest_days[pp['name']]:
                            continue
                        if t in forced_rest_days.get(pp['name'], set()):
                            continue
                        if swap_out in rest_days[pp['name']]:
                            continue
                        new_worst = (rest_days[worst] - {swap_out}) | {t}
                        new_donor = (rest_days[pp['name']] - {t}) | {swap_out}
                        ok = True
                        for nm, nr in [(worst, new_worst), (pp['name'], new_donor)]:
                            nwm = [d not in nr for d in range(1, self.num_days + 1)]
                            nrm = [d in nr for d in range(1, self.num_days + 1)]
                            nmw = self._max_consecutive(nwm)
                            nmr = self._max_consecutive(nrm)
                            if nm == worst:
                                if nmw >= old_mw:
                                    ok = False
                                    break
                            else:
                                if nmw > 5:
                                    ok = False
                                    break
                            if len(nr) != rest_target:
                                ok = False
                                break
                        if ok:
                            best_swap = (pp['name'], swap_out, t)
                            break

            if best_swap:
                pp_name, swap_out, target = best_swap
                rest_days[worst].remove(swap_out)
                rest_days[worst].add(target)
                rest_days[pp_name].remove(target)
                rest_days[pp_name].add(swap_out)

        _tick("Phase 4.6 完成, Phase 4.6b")
        # Phase 4.6b: Break remaining long work streaks (≥6 days)
        for _ in range(50):
            any_fixed = False
            for p in self.online_staff:
                name = p['name']
                wm = [d not in rest_days[name] for d in range(1, self.num_days + 1)]
                mw = self._max_consecutive(wm)
                pm_ws_b = pm.get('work_streak', {}).get(name, 0)
                if pm_ws_b > 0:
                    cons_day1 = 0
                    for w in wm:
                        if w:
                            cons_day1 += 1
                        else:
                            break
                    mw = max(mw, pm_ws_b + cons_day1)
                if name in self.rest_requests:
                    if mw < 6:
                        continue
                elif mw < 6:
                    continue
                cur = 0
                best_start, best_end, best_len = 0, 0, 0
                for i, w in enumerate(wm):
                    if w:
                        cur += 1
                    else:
                        cur = 0
                    if cur > best_len:
                        best_len = cur
                        best_end = i + 1
                        best_start = best_end - cur + 1
                streak_days = list(range(best_start, best_end + 1))
                found = False
                w5_boundary = self.num_days - 4
                for work_day in streak_days:
                    if found:
                        break
                    early_rest = sorted([d for d in rest_days[name]
                                        if d <= w5_boundary + 1
                                        and d not in forced_rest_days.get(name, set())], reverse=True)
                    for rest_day in early_rest:
                        if found:
                            break
                        if rest_day in streak_days:
                            continue
                        test_rest = (rest_days[name] - {rest_day}) | {work_day}
                        rm = [d in test_rest for d in range(1, self.num_days + 1)]
                        wm2 = [d not in test_rest for d in range(1, self.num_days + 1)]
                        new_mw = self._max_consecutive(wm2)
                        if new_mw >= 7:
                            continue
                        dw = self._daily_workers(rest_days)
                        new_wd_workers = dw[work_day - 1] - 1
                        new_rd_workers = dw[rest_day - 1] + 1
                        if abs(new_wd_workers - effective_target[work_day - 1]) > 3:
                            continue
                        if abs(new_rd_workers - effective_target[rest_day - 1]) > 3:
                            continue
                        rest_days[name].remove(rest_day)
                        rest_days[name].add(work_day)
                        found = True
                        any_fixed = True
            if not any_fixed:
                break

        _tick("Phase 4.6 完成, Phase 5")
        # 诊断：Phase 5 前的缺口状态
        dw_pre5 = self._daily_workers(rest_days)
        gaps_pre5 = [dw_pre5[d-1] - effective_target[d-1] for d in range(1, self.num_days+1)]
        over_pre5 = [(d+1, g) for d, g in enumerate(gaps_pre5) if g >= 2]
        under_pre5 = [(d+1, g) for d, g in enumerate(gaps_pre5) if g <= -2]
        if over_pre5 or under_pre5:
            print(f"  [diag] Phase 5 前有缺口: 超员{over_pre5[:5]} 缺员{under_pre5[:5]}", flush=True)

        # Phase 5: Final rest count fix
        for p in self.online_staff:
            name = p['name']
            while len(rest_days[name]) > rest_target:
                blocks = self._find_rest_blocks(rest_days[name])
                candidates = []
                for blk in blocks:
                    for d in blk:
                        if d in forced_rest_days.get(name, set()):
                            continue
                        test_rest = rest_days[name] - {d}
                        dw = sum(1 for pp in self.online_staff if d not in rest_days[pp['name']])
                        gap = dw - effective_target[d - 1]
                        mw = self._max_consecutive(
                            [day not in test_rest for day in range(1, self.num_days + 1)])
                        # 优先移除缺员日的休息（删除休息=增加上班人数，应该补到缺员日）
                        candidates.append((gap, mw, d))
                if candidates:
                    candidates.sort()
                    rest_days[name].remove(candidates[0][2])

        # Final target-matching (threshold >= 2)
        for _ in range(400):
            dw = self._daily_workers(rest_days)
            gaps = [(d, dw[d - 1] - effective_target[d - 1]) for d in range(1, self.num_days + 1)]
            overs = [(d, g) for d, g in gaps if g >= 2]
            unders = [(d, g) for d, g in gaps if g <= -2]
            if not overs or not unders:
                break
            overs.sort(key=lambda x: -x[1])
            unders.sort(key=lambda x: x[1])
            improved = False
            for over_d, _ in overs:
                for under_d, _ in unders:
                    if improved:
                        break
                    for p in self.online_staff:
                        name = p['name']
                        if under_d in rest_days[name] and over_d not in rest_days[name]:
                            if under_d in forced_rest_days.get(name, set()):
                                continue
                            test_rest = (rest_days[name] - {under_d}) | {over_d}
                            wm = [d not in test_rest for d in range(1, self.num_days + 1)]
                            rm = [d in test_rest for d in range(1, self.num_days + 1)]
                            mw, mr = self._max_consecutive(wm), self._max_consecutive(rm)
                            max_mw = 6
                            if mw <= max_mw and mr <= 2:
                                rest_days[name].remove(under_d)
                                rest_days[name].add(over_d)
                                improved = True
                                break
            if not improved:
                # 两人回退
                for over_d, _ in overs:
                    if improved:
                        break
                    for under_d, _ in unders:
                        if improved:
                            break
                        over_candidates = sorted(
                            [p for p in self.online_staff if over_d not in rest_days[p['name']]],
                            key=lambda p: -self._max_consecutive(
                                [d not in rest_days[p['name']] for d in range(1, self.num_days + 1)]))
                        under_candidates = sorted(
                            [p for p in self.online_staff if under_d in rest_days[p['name']]],
                            key=lambda p: self._max_consecutive(
                                [d not in rest_days[p['name']] for d in range(1, self.num_days + 1)]))
                        for p_over in over_candidates:
                            if improved:
                                break
                            test_over = rest_days[p_over['name']] | {over_d}
                            rm_over = [d in test_over for d in range(1, self.num_days + 1)]
                            if self._max_consecutive(rm_over) > 2:
                                continue
                            for p_under in under_candidates:
                                if p_under['name'] == p_over['name']:
                                    continue
                                test_under = rest_days[p_under['name']] - {under_d}
                                wm_under = [d not in test_under for d in range(1, self.num_days + 1)]
                                max_mw_u = 7
                                if self._max_consecutive(wm_under) <= max_mw_u:
                                    rest_days[p_over['name']].add(over_d)
                                    rest_days[p_under['name']].remove(under_d)
                                    improved = True
                                    break
                if not improved:
                    break

        # Final target-matching with threshold=1
        for _ in range(400):
            dw = self._daily_workers(rest_days)
            gaps = [(d, dw[d - 1] - effective_target[d - 1]) for d in range(1, self.num_days + 1)]
            overs = [(d, g) for d, g in gaps if g >= 1]
            unders = [(d, g) for d, g in gaps if g <= -1]
            if not overs or not unders:
                break
            overs.sort(key=lambda x: -x[1])
            unders.sort(key=lambda x: x[1])
            improved = False
            for over_d, _ in overs:
                for under_d, _ in unders:
                    if improved:
                        break
                    for p in self.online_staff:
                        name = p['name']
                        if under_d in rest_days[name] and over_d not in rest_days[name]:
                            if under_d in forced_rest_days.get(name, set()):
                                continue
                            test_rest = (rest_days[name] - {under_d}) | {over_d}
                            wm = [d not in test_rest for d in range(1, self.num_days + 1)]
                            rm = [d in test_rest for d in range(1, self.num_days + 1)]
                            mw, mr = self._max_consecutive(wm), self._max_consecutive(rm)
                            max_mw = 6
                            if mw <= max_mw and mr <= 2:
                                rest_days[name].remove(under_d)
                                rest_days[name].add(over_d)
                                improved = True
                                break
            if not improved:
                # 两人回退
                for over_d, _ in overs:
                    if improved:
                        break
                    for under_d, _ in unders:
                        if improved:
                            break
                        over_candidates = sorted(
                            [p for p in self.online_staff if over_d not in rest_days[p['name']]],
                            key=lambda p: -self._max_consecutive(
                                [d not in rest_days[p['name']] for d in range(1, self.num_days + 1)]))
                        under_candidates = sorted(
                            [p for p in self.online_staff if under_d in rest_days[p['name']]],
                            key=lambda p: self._max_consecutive(
                                [d not in rest_days[p['name']] for d in range(1, self.num_days + 1)]))
                        for p_over in over_candidates:
                            if improved:
                                break
                            test_over = rest_days[p_over['name']] | {over_d}
                            rm_over = [d in test_over for d in range(1, self.num_days + 1)]
                            if self._max_consecutive(rm_over) > 2:
                                continue
                            for p_under in under_candidates:
                                if p_under['name'] == p_over['name']:
                                    continue
                                test_under = rest_days[p_under['name']] - {under_d}
                                wm_under = [d not in test_under for d in range(1, self.num_days + 1)]
                                max_mw_u = 7
                                if self._max_consecutive(wm_under) <= max_mw_u:
                                    rest_days[p_over['name']].add(over_d)
                                    rest_days[p_under['name']].remove(under_d)
                                    improved = True
                                    break
                if not improved:
                    break

        _tick("Phase 5 完成")
        # Phase 5.5: 跨月连上最终强制修复（在所有调整阶段之后，班次分配之前）
        for p in self.online_staff:
            name = p['name']
            pm_ws = pm.get('work_streak', {}).get(name, 0)
            if pm_ws == 0:
                continue
            wm = [d not in rest_days[name] for d in range(1, self.num_days + 1)]
            cur = pm_ws
            for i, w in enumerate(wm):
                if w:
                    cur += 1
                    day = i + 1
                    max_allowed = 5
                    if cur > max_allowed:
                        if day not in rest_days[name]:
                            rest_days[name].add(day)
                            # 优先从超员日移除补偿（保持偏差均衡），找不到补偿就回滚
                            if not self._remove_rest_from_overstaffed(name, rest_days, forced_rest_days, effective_target):
                                rest_days[name].discard(day)  # 保偏差=0，放弃此修复
                            else:
                                forced_rest_days[name].add(day)
                        cur = 0
                else:
                    cur = 0

        # Phase 5.6: 月内长连上修复（≥6天）— 使用1:1休息日交换，保偏差=0
        _p56_swapped = set()  # 防乒乓交换
        for _ in range(50):
            # 收集所有连上≥6的人，按连上长度降序
            streak_info = []
            for p in self.online_staff:
                name = p['name']
                wm = [d not in rest_days[name] for d in range(1, self.num_days + 1)]
                mw = self._max_consecutive(wm)
                if mw < 6:
                    continue
                cur = 0
                best_start, best_end, best_len = 0, 0, 0
                for i, w in enumerate(wm):
                    if w:
                        cur += 1
                        if cur > best_len:
                            best_len = cur
                            best_end = i + 1
                            best_start = best_end - cur + 1
                    else:
                        cur = 0
                streak_info.append((best_len, name, best_start, best_end, wm))
            if not streak_info:
                break
            streak_info.sort(key=lambda x: -x[0])
            found_any = False
            for _, name, best_start, best_end, wm in streak_info:
                if found_any:
                    break
                # 优先尝试在连上段内插入休息日：尝试 best_start+2 到 best_end-1
                target_days = list(range(best_start + 2, best_end))
                if best_len == 6:
                    # 对于恰好6天，优先尝试 best_start+3（break into 3+3）
                    target_days.sort(key=lambda d: abs(d - (best_start + 3)))
                # A 的非强制休息日（不在连上段内的优先）
                a_rests = sorted(
                    [d for d in rest_days[name] if d not in forced_rest_days.get(name, set())],
                    key=lambda d: 0 if best_start <= d <= best_end else 1
                )
                for target in target_days:
                    if found_any:
                        break
                    if target in rest_days[name]:
                        continue
                    for give_day in a_rests:
                        if found_any:
                            break
                        if give_day == target:
                            continue
                        # 找交换伙伴 B：在 target 休息（非强制），在 give_day 工作
                        for p2 in self.online_staff:
                            if found_any:
                                break
                            b_name = p2['name']
                            if b_name == name:
                                continue
                            pair_key = frozenset({name, b_name})
                            if pair_key in _p56_swapped:
                                continue
                            if target in forced_rest_days.get(b_name, set()):
                                continue
                            b_wm = [d not in rest_days[b_name] for d in range(1, self.num_days + 1)]
                            if b_wm[target - 1] or not b_wm[give_day - 1]:
                                continue
                            # 尝试交换
                            a_new_rests = (rest_days[name] | {target}) - {give_day}
                            b_new_rests = (rest_days[b_name] | {give_day}) - {target}
                            a_new_wm = [d not in a_new_rests for d in range(1, self.num_days + 1)]
                            b_new_wm = [d not in b_new_rests for d in range(1, self.num_days + 1)]
                            a_new_rm = [d in a_new_rests for d in range(1, self.num_days + 1)]
                            b_new_rm = [d in b_new_rests for d in range(1, self.num_days + 1)]
                            if (self._max_consecutive(a_new_wm) <= 5 and self._max_consecutive(a_new_rm) <= 2 and
                                self._max_consecutive(b_new_wm) <= 5 and self._max_consecutive(b_new_rm) <= 2):
                                rest_days[name] = a_new_rests
                                rest_days[b_name] = b_new_rests
                                forced_rest_days[name].add(target)
                                _p56_swapped.add(pair_key)
                                found_any = True
            if not found_any:
                break

        # Phase 5.7: 均衡偏差（使用原始target与verification一致）
        buru_work = [1 if self.is_workday(d) else 0 for d in range(1, self.num_days + 1)]
        online_target = [max(0, self.daily_targets[d] - buru_work[d]) for d in range(self.num_days)]
        for _ in range(200):
            dw = self._daily_workers(rest_days)
            gaps = [(d + 1, dw[d] - online_target[d]) for d in range(self.num_days)]
            # Step A: 修复 diff <= -3（从 diff >= -1 的日期借休息日）
            deep_unders = [(d, g) for d, g in gaps if g <= -3]
            can_lends = [(d, g) for d, g in gaps if g >= -1]
            improved = False
            if deep_unders and can_lends:
                for under_d, _ in deep_unders:
                    if improved:
                        break
                    for over_d, _ in can_lends:
                        if improved:
                            break
                        for p in self.online_staff:
                            name = p['name']
                            if under_d in rest_days[name] and over_d not in rest_days[name]:
                                if under_d in forced_rest_days.get(name, set()):
                                    continue
                                if over_d in forced_rest_days.get(name, set()):
                                    continue
                                test_rest = (rest_days[name] - {under_d}) | {over_d}
                                wm = [d not in test_rest for d in range(1, self.num_days + 1)]
                                rm = [d in test_rest for d in range(1, self.num_days + 1)]
                                mw, mr = self._max_consecutive(wm), self._max_consecutive(rm)
                                # 跨月检查：上月结转+本月月初连续工作不能超标
                                pm_ws_chk = pm.get('work_streak', {}).get(name, 0)
                                cross_ok = True
                                if pm_ws_chk > 0:
                                    cons = 0
                                    for w in wm:
                                        if w:
                                            cons += 1
                                        else:
                                            break
                                    cross_ok = (pm_ws_chk + cons <= 5)
                                if mw <= 5 and mr <= 2 and cross_ok:
                                    rest_days[name].remove(under_d)
                                    rest_days[name].add(over_d)
                                    improved = True
                                    break
                if improved:
                    continue  # 重新计算 gap 后继续
            # Step B: 修复 diff >= 0 + diff <= -2 配对
            zeros = [(d, g) for d, g in gaps if g >= 0]
            neg_twos = [(d, g) for d, g in gaps if g <= -2]
            if not zeros or not neg_twos:
                break
            improved = False
            for over_d, _ in zeros:
                if improved:
                    break
                for under_d, _ in neg_twos:
                    if improved:
                        break
                    for p in self.online_staff:
                        name = p['name']
                        if under_d in rest_days[name] and over_d not in rest_days[name]:
                            if under_d in forced_rest_days.get(name, set()):
                                continue
                            if over_d in forced_rest_days.get(name, set()):
                                continue
                            test_rest = (rest_days[name] - {under_d}) | {over_d}
                            wm = [d not in test_rest for d in range(1, self.num_days + 1)]
                            rm = [d in test_rest for d in range(1, self.num_days + 1)]
                            mw, mr = self._max_consecutive(wm), self._max_consecutive(rm)
                            # 跨月检查：上月结转+本月月初连续工作不能超标
                            pm_ws_chk = pm.get('work_streak', {}).get(name, 0)
                            cross_ok = True
                            if pm_ws_chk > 0:
                                cons = 0
                                for w in wm:
                                    if w:
                                        cons += 1
                                    else:
                                        break
                                cross_ok = (pm_ws_chk + cons <= 5)
                            if mw <= 5 and mr <= 2 and cross_ok:
                                rest_days[name].remove(under_d)
                                rest_days[name].add(over_d)
                                improved = True
                                break
            if not improved:
                break

        # Phase 5.8: 跨月连上均匀分摊 — 高连上人用月末休息换低连上人月初休息，偏差不变、休息数不变
        _p58_swapped_pairs = set()  # 防止同一对来回乒乓交换
        for _iter58 in range(800):
            cross_scores = []
            for p in self.online_staff:
                name = p['name']
                pm_ws = pm.get('work_streak', {}).get(name, 0)
                wm = [d not in rest_days[name] for d in range(1, self.num_days + 1)]
                mw = self._max_consecutive(wm)
                if pm_ws > 0:
                    cons = 0
                    for w in wm:
                        if w: cons += 1
                        else: break
                    mw = max(mw, pm_ws + cons)
                cross_scores.append((mw, name))
            cross_scores.sort(key=lambda x: -x[0])
            swapped_any = False
            for hi_idx in range(len(cross_scores)):
                if swapped_any: break
                max_cs, max_name = cross_scores[hi_idx]
                if max_cs <= 5: continue
                for lo_idx in range(len(cross_scores) - 1, -1, -1):
                    if swapped_any: break
                    if lo_idx <= hi_idx: break
                    min_cs, min_name = cross_scores[lo_idx]
                    if max_cs - min_cs <= 1: continue
                    pair_key = frozenset((max_name, min_name))
                    if pair_key in _p58_swapped_pairs: continue
                    high_wm = [d not in rest_days[max_name] for d in range(1, self.num_days + 1)]
                    low_wm = [d not in rest_days[min_name] for d in range(1, self.num_days + 1)]
                    found_pair = False
                    for day in range(1, 8):
                        if found_pair: break
                        if day in forced_rest_days.get(min_name, set()): continue
                        if not low_wm[day - 1] and high_wm[day - 1]:
                            early_rest = day
                            for lday in range(self.num_days, 14, -1):
                                if found_pair: break
                                if lday in forced_rest_days.get(max_name, set()): continue
                                if not high_wm[lday - 1] and low_wm[lday - 1]:
                                    late_rest = lday
                                    test_high = (rest_days[max_name] | {early_rest}) - {late_rest}
                                    test_low = (rest_days[min_name] | {late_rest}) - {early_rest}
                                    test_high_wm = [d not in test_high for d in range(1, self.num_days + 1)]
                                    test_low_wm = [d not in test_low for d in range(1, self.num_days + 1)]
                                    test_high_rm = [d in test_high for d in range(1, self.num_days + 1)]
                                    test_low_rm = [d in test_low for d in range(1, self.num_days + 1)]
                                    h_mw = self._max_consecutive(test_high_wm)
                                    h_mr = self._max_consecutive(test_high_rm)
                                    l_mw = self._max_consecutive(test_low_wm)
                                    l_mr = self._max_consecutive(test_low_rm)
                                    h_pm = pm.get('work_streak', {}).get(max_name, 0)
                                    l_pm = pm.get('work_streak', {}).get(min_name, 0)
                                    h_cons = 0
                                    for w in test_high_wm:
                                        if w: h_cons += 1
                                        else: break
                                    l_cons = 0
                                    for w in test_low_wm:
                                        if w: l_cons += 1
                                        else: break
                                    new_h_cross = max(h_mw, h_pm + h_cons) if h_pm > 0 else h_mw
                                    new_l_cross = max(l_mw, l_pm + l_cons) if l_pm > 0 else l_mw
                                    if (h_mw <= 6 and h_mr <= 2 and l_mw <= 6 and l_mr <= 2 and
                                        new_h_cross < max_cs and new_l_cross > min_cs and
                                        new_l_cross <= 6):
                                        rest_days[max_name] = test_high
                                        rest_days[min_name] = test_low
                                        _p58_swapped_pairs.add(pair_key)
                                        forced_rest_days.setdefault(max_name, set()).add(early_rest)
                                        swapped_any = True
                                        found_pair = True
            if not swapped_any:
                break

        # Phase 5.9: 跨月连上均匀分摊 — 高跨月者用靠后休息换低跨月者靠前休息
        _p59_swapped = set()
        for _iter in range(200):
            cross_map = {}
            for p in self.online_staff:
                name = p['name']
                pm_ws = pm.get('work_streak', {}).get(name, 0)
                wm = [d not in rest_days[name] for d in range(1, self.num_days + 1)]
                if pm_ws > 0:
                    cons = 0
                    for w in wm:
                        if w: cons += 1
                        else: break
                    cross_map[name] = pm_ws + cons
                else:
                    cross_map[name] = 0
            highs = [(n, cross_map[n]) for n in cross_map if cross_map[n] >= 6]
            if not highs:
                break
            highs.sort(key=lambda x: -x[1])
            swapped_any = False
            for sname, s_cross in highs:
                s_wm = [d not in rest_days[sname] for d in range(1, self.num_days + 1)]
                s_pm = pm.get('work_streak', {}).get(sname, 0)
                first_rest = None
                for d in range(1, self.num_days + 1):
                    if not s_wm[d - 1]:
                        first_rest = d
                        break
                if first_rest is None or first_rest == 1:
                    continue
                s_non_forced = sorted([d for d in rest_days[sname]
                                       if d not in forced_rest_days.get(sname, set())],
                                      reverse=True)
                low_candidates = sorted(cross_map.keys(), key=lambda n: cross_map[n])
                for fname in low_candidates:
                    f_cross = cross_map[fname]
                    if sname == fname: continue
                    if s_cross - f_cross <= 1: continue
                    pair_key = frozenset({sname, fname})
                    if pair_key in _p59_swapped: continue
                    f_wm = [d not in rest_days[fname] for d in range(1, self.num_days + 1)]
                    f_pm = pm.get('work_streak', {}).get(fname, 0)
                    found_pair = False
                    for day in range(1, first_rest):
                        if found_pair: break
                        if day in forced_rest_days.get(fname, set()): continue
                        if not f_wm[day - 1] and s_wm[day - 1]:
                            # 预检：fname 交出 day 休息日后，月内连上是否会超标？
                            tf_pre = rest_days[fname] - {day}
                            tf_pre_wm = [d not in tf_pre for d in range(1, self.num_days + 1)]
                            if self._max_consecutive(tf_pre_wm) > 5:
                                continue
                            for give_day in s_non_forced:
                                if found_pair: break
                                if give_day <= day: continue
                                if give_day in rest_days[fname]: continue
                                if not f_wm[give_day - 1]: continue
                                ts = (rest_days[sname] | {day}) - {give_day}
                                tf = (rest_days[fname] | {give_day}) - {day}
                                ts_wm = [d not in ts for d in range(1, self.num_days + 1)]
                                tf_wm = [d not in tf for d in range(1, self.num_days + 1)]
                                ts_rm = [d in ts for d in range(1, self.num_days + 1)]
                                tf_rm = [d in tf for d in range(1, self.num_days + 1)]
                                s_cons = 0
                                for w in ts_wm:
                                    if w: s_cons += 1
                                    else: break
                                f_cons = 0
                                for w in tf_wm:
                                    if w: f_cons += 1
                                    else: break
                                new_s = max(self._max_consecutive(ts_wm), s_pm + s_cons) if s_pm > 0 else self._max_consecutive(ts_wm)
                                new_f = max(self._max_consecutive(tf_wm), f_pm + f_cons) if f_pm > 0 else self._max_consecutive(tf_wm)
                                s_mw_val = self._max_consecutive(ts_wm)
                                s_mr_val = self._max_consecutive(ts_rm)
                                f_mw_val = self._max_consecutive(tf_wm)
                                f_mr_val = self._max_consecutive(tf_rm)
                                new_f_ok = new_f <= 5 if f_pm > 0 else True
                                ok = (s_mw_val <= 5 and s_mr_val <= 2 and
                                      f_mw_val <= 5 and f_mr_val <= 2 and
                                      new_s < s_cross and new_f_ok)
                                if ok:
                                    rest_days[sname] = ts
                                    rest_days[fname] = tf
                                    forced_rest_days.setdefault(sname, set()).add(day)
                                    _p59_swapped.add(pair_key)
                                    found_pair = True
                                    swapped_any = True
                                    cross_map[sname] = new_s
                                    cross_map[fname] = new_f
                    if found_pair:
                        s_cross = cross_map[sname]
                        if s_cross <= 5:
                            break
            if not swapped_any:
                break

        # Phase 5.9b: 修复 Phase 5.9 交换可能引入的月内连上≥6
        for _ in range(50):
            streak_info = []
            for p in self.online_staff:
                name = p['name']
                wm = [d not in rest_days[name] for d in range(1, self.num_days + 1)]
                mw = self._max_consecutive(wm)
                if mw < 6:
                    continue
                cur = 0
                best_start, best_end, best_len = 0, 0, 0
                for i, w in enumerate(wm):
                    if w:
                        cur += 1
                        if cur > best_len:
                            best_len = cur
                            best_end = i + 1
                            best_start = best_end - cur + 1
                    else:
                        cur = 0
                streak_info.append((best_len, name, best_start, best_end, wm))
            if not streak_info:
                break
            streak_info.sort(key=lambda x: -x[0])
            found_any = False
            for _, name, best_start, best_end, wm in streak_info:
                if found_any:
                    break
                target_days = list(range(best_start + 2, best_end))
                if best_len == 6:
                    target_days.sort(key=lambda d: abs(d - (best_start + 3)))
                a_rests = sorted(
                    [d for d in rest_days[name] if d not in forced_rest_days.get(name, set())],
                    key=lambda d: 0 if best_start <= d <= best_end else 1
                )
                for target in target_days:
                    if found_any:
                        break
                    if target in rest_days[name]:
                        continue
                    for give_day in a_rests:
                        if found_any:
                            break
                        if give_day == target:
                            continue
                        for p2 in self.online_staff:
                            if found_any:
                                break
                            b_name = p2['name']
                            if b_name == name:
                                continue
                            if target in forced_rest_days.get(b_name, set()):
                                continue
                            b_wm = [d not in rest_days[b_name] for d in range(1, self.num_days + 1)]
                            if b_wm[target - 1] or not b_wm[give_day - 1]:
                                continue
                            a_new_rests = (rest_days[name] | {target}) - {give_day}
                            b_new_rests = (rest_days[b_name] | {give_day}) - {target}
                            a_new_wm = [d not in a_new_rests for d in range(1, self.num_days + 1)]
                            b_new_wm = [d not in b_new_rests for d in range(1, self.num_days + 1)]
                            a_new_rm = [d in a_new_rests for d in range(1, self.num_days + 1)]
                            b_new_rm = [d in b_new_rests for d in range(1, self.num_days + 1)]
                            if (self._max_consecutive(a_new_wm) <= 5 and self._max_consecutive(a_new_rm) <= 2 and
                                self._max_consecutive(b_new_wm) <= 5 and self._max_consecutive(b_new_rm) <= 2):
                                rest_days[name] = a_new_rests
                                rest_days[b_name] = b_new_rests
                                forced_rest_days[name].add(target)
                                found_any = True
            if not found_any:
                break

        # Phase 5.10: 休息日数量平衡（纠正前序阶段可能引入的偏差）
        overs = [(p['name'], len(rest_days[p['name']])) for p in self.online_staff
                 if len(rest_days[p['name']]) > rest_target]
        unders = [(p['name'], len(rest_days[p['name']])) for p in self.online_staff
                  if len(rest_days[p['name']]) < rest_target]
        _p510_swapped = set()
        for oname, ocount in overs:
            for uname, ucount in unders:
                if ocount <= rest_target or ucount >= rest_target:
                    continue
                pair_key = frozenset({oname, uname})
                if pair_key in _p510_swapped:
                    continue
                # 从多余者找一个非强制休息日给不足者
                for gday in sorted(rest_days[oname]):
                    if gday in forced_rest_days.get(oname, set()):
                        continue
                    if gday in rest_days[uname]:
                        continue  # 不足者已在这天休息
                    # 检查不足者在 gday 是否工作
                    if gday not in rest_days[uname]:
                        o_new = rest_days[oname] - {gday}
                        u_new = rest_days[uname] | {gday}
                        o_wm = [d not in o_new for d in range(1, self.num_days + 1)]
                        u_wm = [d not in u_new for d in range(1, self.num_days + 1)]
                        o_rm = [d in o_new for d in range(1, self.num_days + 1)]
                        u_rm = [d in u_new for d in range(1, self.num_days + 1)]
                        if (self._max_consecutive(o_wm) <= 5 and self._max_consecutive(o_rm) <= 2 and
                            self._max_consecutive(u_wm) <= 5 and self._max_consecutive(u_rm) <= 2):
                            rest_days[oname] = o_new
                            rest_days[uname] = u_new
                            ocount -= 1
                            ucount += 1
                            _p510_swapped.add(pair_key)
                            break
                if ocount <= rest_target:
                    break

        # Phase 5.11: 休息需求交换优化
        _p511_reqs = getattr(self, 'rest_requests', {}) or {}
        if _p511_reqs:
            for _round in range(20):
                any_swap = False
                for name in _p511_reqs:
                    requested = _p511_reqs[name]
                    best_swap = None
                    best_score = -1
                    for md in requested:
                        if md in rest_days[name]:
                            continue
                        s_pm = pm.get('work_streak', {}).get(name, 0)
                        for giver in self.online_staff:
                            gname = giver['name']
                            if gname == name:
                                continue
                            if md not in rest_days[gname]:
                                continue
                            if md in _p511_reqs.get(gname, []):
                                continue
                            if md in forced_rest_days.get(gname, set()):
                                continue
                            g_pm = pm.get('work_streak', {}).get(gname, 0)
                            for give_day in sorted(rest_days[name]):
                                if give_day in rest_days[gname]:
                                    continue
                                # 仅当有跨月风险(prev_ws>0)时才保护强制休息日
                                # 无跨月风险的人被误标forced也不应阻止交换
                                if give_day in forced_rest_days.get(name, set()) and s_pm > 0:
                                    continue
                                # 需求人员只能用非需求日交换，不能给出自己申请的需求日
                                if give_day in _p511_reqs.get(name, []):
                                    continue
                                ts = (rest_days[name] | {md}) - {give_day}
                                tf = (rest_days[gname] | {give_day}) - {md}
                                ts_wm = [d not in ts for d in range(1, self.num_days + 1)]
                                tf_wm = [d not in tf for d in range(1, self.num_days + 1)]
                                if self._max_consecutive(ts_wm) > 5:
                                    continue
                                if self._max_consecutive(tf_wm) > 5:
                                    continue
                                ts_rm = [d in ts for d in range(1, self.num_days + 1)]
                                tf_rm = [d in tf for d in range(1, self.num_days + 1)]
                                if self._max_consecutive(ts_rm) > 2:
                                    continue
                                if self._max_consecutive(tf_rm) > 2:
                                    continue
                                s_cons = 0
                                for w in ts_wm:
                                    if w: s_cons += 1
                                    else: break
                                f_cons = 0
                                for w in tf_wm:
                                    if w: f_cons += 1
                                    else: break
                                if s_pm + s_cons > 5:
                                    continue
                                if g_pm > 0 and g_pm + f_cons > 5:
                                    continue
                                score = 5 - g_pm
                                if score > best_score:
                                    best_score = score
                                    best_swap = (md, gname, give_day)
                    if best_swap is not None:
                        md, gname, give_day = best_swap
                        rest_days[name] = (rest_days[name] | {md}) - {give_day}
                        rest_days[gname] = (rest_days[gname] | {give_day}) - {md}
                        forced_rest_days.setdefault(name, set()).add(md)
                        any_swap = True
                if not any_swap:
                    break

        _tick("Phase 6")
        # Phase 6: Assign shifts
        schedules = {p['name']: [XI] * self.num_days for p in self.online_staff}
        prev_shift = {}
        for p in self.online_staff:
            name = p['name']
            prev_shift[name] = pm.get('last_shift', {}).get(name)
        balance = {p['name']: {'early': 0, 'late': 0} for p in self.online_staff}
        jd_count = {p['name']: 0 for p in self.online_staff}
        wan2_count = {p['name']: 0 for p in self.online_staff}
        zaozao2_count = {p['name']: 0 for p in self.online_staff}
        shift_type_counts = {p['name']: {} for p in self.online_staff}

        for day in range(1, self.num_days + 1):
            working = [p for p in self.online_staff if day not in rest_days[p['name']]]
            if day == 1:
                _tick(f"Phase 6 day 1 start, n_workers={len(working)}")
            self._assign_shifts_for_day(working, day, prev_shift, rest_days,
                                        fenxiao_rest_days, schedules, balance, jd_count, wan2_count,
                                        zaozao2_count, shift_type_counts)
            for p in working:
                prev_shift[p['name']] = schedules[p['name']][day - 1]
            for p in self.online_staff:
                if day in rest_days[p['name']]:
                    prev_shift[p['name']] = None

        # Post-fix late→early
        for day in range(1, self.num_days):
            for p in self.online_staff:
                name = p['name']
                if schedules[name][day - 1] in LATE_SHIFTS and schedules[name][day] in EARLY_SHIFTS:
                    next_working = [pp for pp in self.online_staff
                                    if day + 1 not in rest_days[pp['name']] and pp['name'] != name]
                    for pp in next_working:
                        if schedules[pp['name']][day] in LATE_SHIFTS:
                            if day == 1 or schedules[pp['name']][day - 1] not in LATE_SHIFTS:
                                schedules[name][day], schedules[pp['name']][day] = \
                                    schedules[pp['name']][day], schedules[name][day]
                                break

        # Post-fix: 需求人员未满足的请求日避免关键班次（早早2、晚二），优先换行政班
        _rest_reqs = getattr(self, 'rest_requests', {}) or {}
        if _rest_reqs:
            _critical = {ZAO_ZAO_2, WAN_ER}
            _admin_shifts = {XING_ZHENG}
            for name, req_days in _rest_reqs.items():
                if name not in schedules:
                    continue
                for day in req_days:
                    if day in rest_days.get(name, set()):
                        continue  # 需求已满足
                    shift = schedules[name][day - 1]
                    if shift not in _critical:
                        continue  # 不是关键班次，无需处理
                    # 找同天可换的同事：优先行政班，其次任意非关键班次
                    coworkers = []
                    for pp in self.online_staff:
                        pname = pp['name']
                        if pname == name:
                            continue
                        if day in rest_days.get(pname, set()):
                            continue  # 同事休息
                        coworker_shift = schedules[pname][day - 1]
                        if coworker_shift in _critical:
                            continue  # 同事也是关键班次，换了没意义
                        if coworker_shift == XI:
                            continue
                        # 同事的这一天也不是他自己的未满足需求日
                        if day in _rest_reqs.get(pname, []):
                            continue
                        # 检查交换后双方都不会产生晚接早
                        # 需求人员拿到 coworker_shift
                        ok = True
                        if coworker_shift in EARLY_SHIFTS and day > 1:
                            if schedules[name][day - 2] in LATE_SHIFTS:
                                ok = False
                        if coworker_shift in LATE_SHIFTS and day < self.num_days:
                            if schedules[name][day] in EARLY_SHIFTS:
                                ok = False
                        # 同事拿到 critical shift
                        if shift in EARLY_SHIFTS and day > 1:
                            if schedules[pname][day - 2] in LATE_SHIFTS:
                                ok = False
                        if shift in LATE_SHIFTS and day < self.num_days:
                            if schedules[pname][day] in EARLY_SHIFTS:
                                ok = False
                        if not ok:
                            continue
                        is_admin = coworker_shift in _admin_shifts
                        coworkers.append((is_admin, pname, coworker_shift))
                    if not coworkers:
                        continue
                    # 优先行政班，其次任意非关键班次
                    coworkers.sort(key=lambda x: (0 if x[0] else 1, x[2]))
                    swap_name = coworkers[0][1]
                    schedules[name][day - 1], schedules[swap_name][day - 1] = \
                        schedules[swap_name][day - 1], schedules[name][day - 1]

        # 最终诊断：输出所有缺口
        dw_final = self._daily_workers(rest_days)
        gaps_final = [dw_final[d-1] - effective_target[d-1] for d in range(1, self.num_days+1)]
        over_final = [(d+1, g) for d, g in enumerate(gaps_final) if g >= 2]
        under_final = [(d+1, g) for d, g in enumerate(gaps_final) if g <= -2]
        if over_final or under_final:
            print(f"  [diag] 最终剩余缺口: 超员{over_final} 缺员{under_final}", flush=True)
        else:
            print(f"  [diag] 最终无≥2缺口, 最大偏差={max(abs(g) for g in gaps_final)}", flush=True)

        _tick("Phase 6 完成, 即将返回")
        if zhou_sched:
            schedules[zhou_name] = zhou_sched
            self.online_staff = original_online_staff
        return schedules

    # ── 生成入口 ──

    def generate(self):
        """执行全部排班，返回 {group_name: {name: [shifts]}} 和 fenxiao_rest"""
        banzhang_sched = self._generate_banzhang()
        fenxiao_sched, fenxiao_rest = self._generate_fenxiao()
        night_sched = self._generate_night_shift()
        buru_sched, buru_support_sched = self._generate_buru_ban()
        changbingjia_sched = self._generate_changbingjia()
        online_sched = self._generate_online(fenxiao_rest)

        return {
            'banzhang': banzhang_sched,
            'fenxiao': fenxiao_sched,
            'night_shift': night_sched,
            'buru_ban': buru_sched,
            'buru_support': buru_support_sched,
            'changbingjia': changbingjia_sched,
            'online': online_sched,
        }, fenxiao_rest

    # ── 验证 ──

    def verify(self, schedules, fenxiao_rest):
        """返回验证结果字典"""
        all_people = (
            [(p, schedules['banzhang'][p['name']], '班长') for p in self.banzhang] +
            [(p, schedules['fenxiao'][p['name']], '分销') for p in self.fenxiao] +
            [(p, schedules['night_shift'][p['name']], '夜班') for p in self.night_shift] +
            [(p, schedules['buru_ban'][p['name']], '行政班(在线)') for p in self.buru_ban] +
            [(p, schedules['buru_support'][p['name']], '行政班(支援)') for p in self.buru_support] +
            [(p, schedules['changbingjia'][p['name']], '长病假') for p in self.changbingjia] +
            [(p, schedules['online'][p['name']], '在线') for p in self.online_staff]
        )

        online_verify = []
        fenxiao_verify = []
        for person, sched, cat in all_people:
            rest_count = sum(1 for s in sched if s == XI)
            work_count = self.num_days - rest_count
            mw, mr = self._streak_stats(sched)
            flag = ''
            if cat in ('在线', '分销') and mw > 5:
                if person['name'] in self.rest_requests:
                    flag = f'连上={mw}(因请休需求)'
                else:
                    flag = f'连上={mw}'
            if cat in ('在线', '分销') and mr > 2:
                flag = f'连休={mr}'
            late_early = 0
            for d in range(1, self.num_days):
                if sched[d - 1] in LATE_SHIFTS and sched[d] in EARLY_SHIFTS:
                    late_early += 1
            if late_early > 0:
                flag += f' 晚接早x{late_early}'

            # 跨月最大连上：仅计算上月结转部分，不混入当月内部连上
            pm_ws = self.prev_month_data.get('work_streak', {}).get(person['name'], 0)
            start_work = 0
            for s in sched:
                if s != XI:
                    start_work += 1
                else:
                    break
            cross_mw = pm_ws + start_work if pm_ws > 0 else 0

            entry = {
                'name': person['name'], 'id': person.get('id', ''),
                'cat': cat, 'rest': rest_count, 'work': work_count,
                'max_work': mw, 'max_rest': mr, 'cross_max_work': cross_mw, 'flag': flag,
            }
            if cat == '分销':
                zao = sum(1 for s in sched if s == ZAO_ZAO_IM)
                wan = sum(1 for s in sched if s == WAN_YI_IM)
                jd = sum(1 for s in sched if s == XING_ZHENG_JD)
                entry['zao_im'] = zao
                entry['wan_im'] = wan
                entry['jidong'] = jd
                fenxiao_verify.append(entry)
            online_verify.append(entry)

        daily_headcount = []
        for d in range(1, self.num_days + 1):
            online_n = sum(1 for p in self.online_staff if schedules['online'][p['name']][d - 1] != XI)
            buru_n = sum(1 for p in self.buru_ban if schedules['buru_ban'][p['name']][d - 1] != XI)
            total = online_n + buru_n
            target = self.daily_targets[d - 1]
            daily_headcount.append({
                'day': d, 'weekday': (self.start_weekday + d - 1) % 7,
                'online': online_n, 'buru': buru_n, 'total': total,
                'target': target, 'diff': total - target,
            })

        # 跨月衔接检查
        pm = self.prev_month_data
        cross_month = {'has_prev_data': bool(pm)}
        if pm:
            last_shift_map = pm.get('last_shift', {})
            daye_names = [n for n, s in last_shift_map.items() if s == DA_YE]
            cross_month['daye_count'] = len(daye_names)
            daye_issues = []
            prev_roles = pm.get('roles', {})
            for name in daye_names:
                for cat_sched in [schedules['night_shift'], schedules['online'], schedules['fenxiao'],
                                   schedules['buru_ban'], schedules['buru_support']]:
                    if name in cat_sched:
                        # 上月也是专职大夜 → 延顺，不检查大夜后休息
                        if prev_roles.get(name) == '专职大夜' and name in schedules.get('night_shift', {}):
                            break
                        s = cat_sched[name]
                        if s[0] != XI:
                            daye_issues.append(f'{name}上月大夜→Day1未休(排了{s[0]})')
                        break
            cross_month['daye_rest_ok'] = len(daye_issues) == 0
            cross_month['daye_issues'] = daye_issues

            work_streak_map = pm.get('work_streak', {})
            cross_issues = []
            for name, ws_val in work_streak_map.items():
                if ws_val >= 5:
                    found = False
                    for cat_key, cat_sched in [('online', schedules['online']), ('fenxiao', schedules['fenxiao']),
                                               ('night_shift', schedules['night_shift']),
                                               ('buru_ban', schedules['buru_ban']),
                                               ('buru_support', schedules['buru_support'])]:
                        if name in cat_sched:
                            found = True
                            if cat_sched[name][0] != XI:
                                reason = ''
                                if cat_key == 'fenxiao':
                                    reason = '(分销需保障2人在岗配IM)'
                                cross_issues.append(f'{name}跨月连上{ws_val}天→Day1未休{reason}')
                            break
                    if not found:
                        cross_issues.append(f'{name}上月连上{ws_val}天→本月不在排班中')
            # 也检查跨月合并后连上>5但上月连上<5的人（如7月末连上3+8月初连上4=7）
            for name, ws_val in work_streak_map.items():
                if ws_val == 0 or ws_val >= 5:
                    continue
                for cat_key, cat_sched in [('online', schedules['online']), ('fenxiao', schedules['fenxiao']),
                                           ('night_shift', schedules['night_shift']),
                                           ('buru_ban', schedules['buru_ban']),
                                           ('buru_support', schedules['buru_support'])]:
                    if name in cat_sched:
                        start_work = 0
                        for s in cat_sched[name]:
                            if s != XI:
                                start_work += 1
                            else:
                                break
                        combined = ws_val + start_work
                        if combined > 5:
                            cross_issues.append(f'{name}跨月连上{combined}天(上月{ws_val}+本月{start_work})')
                        break
            cross_month['work_streak_ok'] = len(cross_issues) == 0
            cross_month['cross_issues'] = cross_issues

            # 晚不能接早跨月检查
            late_early_issues = []
            for name, last in last_shift_map.items():
                if last in LATE_SHIFTS:
                    for cat_sched in [schedules['online'], schedules['fenxiao'],
                                      schedules['buru_ban'], schedules['buru_support']]:
                        if name in cat_sched:
                            if cat_sched[name][0] in EARLY_SHIFTS:
                                late_early_issues.append(f'{name}上月{last}→Day1早班({cat_sched[name][0]})')
                            break
            cross_month['late_early_ok'] = len(late_early_issues) == 0
            cross_month['late_early_issues'] = late_early_issues

        # 班次每日统计（每个班次每天几人 + 合计，大夜单独）
        online_shift_types = [ZAO_BAN, ZAO_SAN, ZAO_ZAO_2, XING_ZHENG, BAI_BAN, XING_ZHENG_JD,
                              ZHONG_SAN, ZHONG_SI, WAN_YI, WAN_ER, ZAO_ZAO_IM, WAN_YI_IM]
        shift_daily = []
        for s_type in online_shift_types:
            counts = []
            for d in range(1, self.num_days + 1):
                c = 0
                for cat_sched in [schedules['online'], schedules['fenxiao'], schedules['buru_ban'],
                                  schedules['buru_support'], schedules['banzhang'],
                                  schedules['changbingjia']]:
                    for sched in cat_sched.values():
                        if sched[d - 1] == s_type:
                            c += 1
                counts.append(c)
            total = sum(counts)
            if total > 0:
                shift_daily.append({'name': s_type, 'daily': counts, 'total': total})
        # 大夜每日统计
        night_daily = []
        for d in range(1, self.num_days + 1):
            c = 0
            for sched in schedules.get('night_shift', {}).values():
                if sched[d - 1] == DA_YE:
                    c += 1
            night_daily.append(c)

        return {
            'online': online_verify,
            'fenxiao': fenxiao_verify,
            'daily_headcount': daily_headcount,
            'cross_month': cross_month,
            'shift_daily': shift_daily,
            'night_daily': night_daily,
        }

    # ── 创建预览数据（不含公式的纯数据） ──

    def preview_data(self, schedules):
        """返回前端可用的二维数组预览数据"""
        rows = []

        # 日期表头行
        header_row = ['姓名', '工号', '组别']
        for d in range(1, self.num_days + 1):
            header_row.append(f'{self.month}/{d}')
        rows.append(header_row)

        def add_group(label, people_list, sched_dict):
            for p in people_list:
                name = p['name']
                s = sched_dict[name]
                row = [name, p.get('id', ''), label]
                for shift in s:
                    row.append(shift)
                rows.append(row)

        add_group('班长', self.banzhang, schedules['banzhang'])
        add_group('专职分销', self.fenxiao, schedules['fenxiao'])
        night_sorted = sorted(self.night_shift, key=lambda p: p.get('group', 'A'))
        add_group('专职大夜', night_sorted, schedules['night_shift'])
        add_group('在线', self.buru_ban, schedules['buru_ban'])
        if self.buru_support:
            add_group('支援', self.buru_support, schedules['buru_support'])
        add_group('在线', self.online_staff, schedules['online'])

        if self.changbingjia:
            add_group('长病假', self.changbingjia, schedules['changbingjia'])

        return rows

    # ── 创建 Excel 文件 ──

    def create_excel(self, schedules):
        """生成 Excel 并返回 BytesIO"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'{self.month}月班'

        header_font = Font(name='微软雅黑', size=8, bold=True, color='FF000000')
        data_font = Font(name='微软雅黑', size=8, color='FF000000')
        bold_font = Font(name='微软雅黑', size=8, bold=True, color='FF000000')
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        shift_colors = {
            XI: 'FFC000', QING_JIA: 'FFC000',
            XING_ZHENG: 'FFF3E0', ZAO_BAN: 'E3F2FD', ZAO_SAN: 'F3E5F5',
            BAI_BAN: 'E8F5E9', ZHONG_SAN: 'E0F7FA', ZHONG_SI: 'FFF8E1',
            WAN_YI: 'FCE4EC', WAN_ER: 'EDE7F6', DA_YE: 'FFFFFF',
            ZAO_ZAO_IM: 'FFFF00', WAN_YI_IM: 'FFFF00', XING_ZHENG_JD: 'EE822F',
            ZAO_ZAO_2: 'D1C4E9',
            CHANG_BAI: 'FFFFFF',
            CHANG_BAI_ZAO: 'FFF9C4',
        }

        row = 1

        # Row 1: Date headers
        ws.cell(row=row, column=1, value='姓名')
        ws.cell(row=row, column=2, value='工号')
        ws.cell(row=row, column=3, value='')
        for day in range(1, self.num_days + 1):
            col = 3 + day
            dt = datetime(self.year, self.month, day)
            cell = ws.cell(row=row, column=col, value=dt)
            cell.number_format = 'm/d;@'
            is_weekend_day = (self.start_weekday + day - 1) % 7 >= 5
            if is_weekend_day:
                cell.font = Font(name='微软雅黑', size=8, bold=True, color='FFFF0000')
            else:
                cell.font = Font(name='微软雅黑', size=8, bold=True, color='FF000000')
            cell.alignment = center_align

        # Person-level summary headers (AI-AU)
        summary_start = 3 + self.num_days + 1
        summary_headers = [
            XI, XING_ZHENG, XING_ZHENG_JD, WAN_ER,
            ZAO_ZAO_IM, ZAO_BAN, ZAO_SAN, BAI_BAN,
            ZHONG_SAN, ZHONG_SI, WAN_YI, WAN_ER, DA_YE,
            ZAO_ZAO_2,
        ]
        for si, sh in enumerate(summary_headers):
            cell = ws.cell(row=row, column=summary_start + si, value=sh)
            cell.font = header_font
            cell.alignment = center_align

        for c in [1, 2]:
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.alignment = center_align

        ws.row_dimensions[1].height = 27

        # Column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 6
        ws.column_dimensions['C'].width = 6
        for day in range(1, self.num_days + 1):
            ws.column_dimensions[get_column_letter(3 + day)].width = 6
        for si in range(len(summary_headers)):
            ws.column_dimensions[get_column_letter(summary_start + si)].width = 6

        row += 1

        def write_person_row(name, emp_id, shifts_list, role_label=''):
            nonlocal row
            ws.row_dimensions[row].height = 33
            ws.cell(row=row, column=1, value=f' {name}')
            ws.cell(row=row, column=2, value=emp_id)
            ws.cell(row=row, column=3, value=role_label)
            for day in range(1, self.num_days + 1):
                col = 3 + day
                shift = shifts_list[day - 1]
                cell = ws.cell(row=row, column=col, value=shift)
                cell.alignment = center_align
                cell.font = data_font
                if shift in shift_colors:
                    cell.fill = PatternFill(start_color=shift_colors[shift],
                                            end_color=shift_colors[shift],
                                            fill_type='solid')
            # COUNTIF formulas
            erow = row
            d_col = get_column_letter(4)
            e_col = get_column_letter(5)
            ah_col = get_column_letter(3 + self.num_days)
            d_range = f'{d_col}{erow}:{ah_col}{erow}'
            e_range = f'{e_col}{erow}:{ah_col}{erow}'
            formulas = [
                (summary_start, f'=COUNTIF({d_range},"{XI}")+COUNTIF({d_range},"{QING_JIA}")'),
                (summary_start + 1, f'=COUNTIF({e_range},"{XING_ZHENG}")'),
                (summary_start + 2, f'=COUNTIF({e_range},"{XING_ZHENG_JD}")'),
                (summary_start + 3, f'=COUNTIF({e_range},"{WAN_ER}")'),
                (summary_start + 4, f'=COUNTIF({d_range},"{ZAO_ZAO_IM}")'),
                (summary_start + 5, f'=COUNTIF({d_range},"{ZAO_BAN}")'),
                (summary_start + 6, f'=COUNTIF({d_range},"{ZAO_SAN}")'),
                (summary_start + 7, f'=COUNTIF({d_range},"{BAI_BAN}")'),
                (summary_start + 8, f'=COUNTIF({d_range},"{ZHONG_SAN}")'),
                (summary_start + 9, f'=COUNTIF({d_range},"{ZHONG_SI}")'),
                (summary_start + 10, f'=COUNTIF({d_range},"{WAN_YI}")+COUNTIF({d_range},"{WAN_YI_IM}")'),
                (summary_start + 11, f'=COUNTIF({d_range},"{WAN_ER}")'),
                (summary_start + 12, f'=COUNTIF({d_range},"{DA_YE}")'),
                (summary_start + 13, f'=COUNTIF({d_range},"{ZAO_ZAO_2}")'),
            ]
            for col_idx, formula in formulas:
                cell = ws.cell(row=row, column=col_idx, value=formula)
                cell.alignment = center_align
                cell.font = data_font
            for c in [1, 2, 3]:
                ws.cell(row=row, column=c).font = data_font
                ws.cell(row=row, column=c).alignment = center_align
            row += 1

        # Write rows
        for p in self.banzhang:
            write_person_row(p['name'], p['id'], schedules['banzhang'][p['name']], '班长')
        row += 1

        fx_start_row = row
        for p in self.fenxiao:
            write_person_row(p['name'], p['id'], schedules['fenxiao'][p['name']], '专职分销')
        fx_end_row = row - 1
        row += 1

        night_start_row = row
        for p in sorted(self.night_shift, key=lambda p: p.get('group', 'A')):
            write_person_row(p['name'], p['id'], schedules['night_shift'][p['name']], '专职大夜')
        night_end_row = row - 1
        row += 1

        buru_start_row = row
        for p in self.buru_ban:
            write_person_row(p['name'], p['id'], schedules['buru_ban'][p['name']], '在线')
        buru_end_row = row - 1
        row += 1

        if self.buru_support:
            for p in self.buru_support:
                write_person_row(p['name'], p['id'], schedules['buru_support'][p['name']], '支援')
            row += 1

        online_start_row = row
        for p in self.online_staff:
            write_person_row(p['name'], p['id'], schedules['online'][p['name']], '在线')
        online_end_row = row - 1

        row += 1
        row += 1

        # Summary Group 1
        all_staff_range = f"{{col}}{fx_start_row}:{{col}}{online_end_row}"
        night_range = f"{{col}}{night_start_row}:{{col}}{online_end_row}"

        group1_shifts = [ZAO_BAN, ZAO_SAN, ZAO_ZAO_2, XING_ZHENG, BAI_BAN,
                         ZHONG_SAN, ZHONG_SI, WAN_YI, WAN_ER]
        group1_start_row = row
        for shift_type in group1_shifts:
            ws.row_dimensions[row].height = 33
            ws.cell(row=row, column=3, value=shift_type).font = data_font
            ws.cell(row=row, column=3).alignment = center_align
            for day in range(1, self.num_days + 1):
                col = 3 + day
                col_letter = get_column_letter(col)
                formula = f'=COUNTIF({all_staff_range.format(col=col_letter)},"{shift_type}")'
                cell = ws.cell(row=row, column=col, value=formula)
                cell.alignment = center_align
                cell.font = data_font
            row += 1

        # 大夜 row
        ws.row_dimensions[row].height = 33
        ws.cell(row=row, column=3, value=DA_YE).font = data_font
        ws.cell(row=row, column=3).alignment = center_align
        for day in range(1, self.num_days + 1):
            col = 3 + day
            col_letter = get_column_letter(col)
            formula = f'=COUNTIF({night_range.format(col=col_letter)},"{DA_YE}")'
            cell = ws.cell(row=row, column=col, value=formula)
            cell.alignment = center_align
            cell.font = data_font
        row += 1

        # 合计 row
        ws.row_dimensions[row].height = 33
        ws.cell(row=row, column=3, value='合计').font = bold_font
        ws.cell(row=row, column=3).alignment = center_align
        for day in range(1, self.num_days + 1):
            col = 3 + day
            col_letter = get_column_letter(col)
            formula = f'=SUM({col_letter}{group1_start_row}:{col_letter}{row - 2})'
            cell = ws.cell(row=row, column=col, value=formula)
            cell.alignment = center_align
            cell.font = bold_font
        row += 1

        row += 1

        # Summary Group 2: IM / 机动
        fx_range = f"{{col}}{fx_start_row}:{{col}}{online_end_row}"
        group2_shifts = [ZAO_ZAO_IM, WAN_YI_IM, XING_ZHENG_JD]
        group2_start_row = row
        for shift_type in group2_shifts:
            ws.row_dimensions[row].height = 33
            ws.cell(row=row, column=3, value=shift_type).font = data_font
            ws.cell(row=row, column=3).alignment = center_align
            for day in range(1, self.num_days + 1):
                col = 3 + day
                col_letter = get_column_letter(col)
                formula = f'=COUNTIF({fx_range.format(col=col_letter)},"{shift_type}")'
                cell = ws.cell(row=row, column=col, value=formula)
                cell.alignment = center_align
                cell.font = data_font
            row += 1

        # 合计 row
        ws.row_dimensions[row].height = 33
        ws.cell(row=row, column=3, value='合计').font = bold_font
        ws.cell(row=row, column=3).alignment = center_align
        for day in range(1, self.num_days + 1):
            col = 3 + day
            col_letter = get_column_letter(col)
            formula = f'=SUM({col_letter}{group2_start_row}:{col_letter}{row - 1})'
            cell = ws.cell(row=row, column=col, value=formula)
            cell.alignment = center_align
            cell.font = bold_font

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


# ── 默认人员配置 ──

DEFAULT_CONFIG = {
    'staff_names': list(STAFF_DB.keys()),
    'total_staff_count': len(STAFF_DB),
    'night_shift_groups': NIGHT_SHIFT_GROUPS,
}
