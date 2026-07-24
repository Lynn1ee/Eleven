import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random

# ============================================================
# CONFIGURATION
# ============================================================
JULY_YEAR, JULY_MONTH = 2026, 7
NUM_DAYS = 31
JULY_1_WEEKDAY = 2  # 0=Mon... 6=Sun, July 1 2026 = Wednesday

# Shift names
XI = '休'
CHANG_BAI = '长白班'
XING_ZHENG = '行政班'
ZAO_ZAO_IM = '早早班\n(IM)'
WAN_YI_IM = '晚一\n(IM)'
BAI_BAN_JD = '白班\n(机动)'
ZAO_BAN = '早班'
ZAO_SAN = '早三'
BAI_BAN = '白班'
ZHONG_SAN = '中三'
ZHONG_SI = '中四'
WAN_YI = '晚一'
WAN_ER = '晚二'
DA_YE = '大夜'
ZAO_ZAO_2 = '早早2'
TIAN_DI = '天地班'
QING_JIA = '请假'

# Late-shift classes (off duty >= 22:00) → next day can't do early shifts
LATE_SHIFTS = {ZHONG_SAN, ZHONG_SI, WAN_YI, WAN_ER, WAN_YI_IM}
# Early-shift classes (banned after a late shift)
EARLY_SHIFTS = {ZAO_BAN, ZAO_SAN, XING_ZHENG, BAI_BAN, ZAO_ZAO_IM, BAI_BAN_JD}
# 早早班 is also early (7:00 start) but the shift name constant isn't in the list above
# ZAO_ZAO_IM is already covered; let me add 早早2 if needed
EARLY_SHIFTS.add(ZAO_ZAO_2)

# ============================================================
# PERSONNEL
# ============================================================

# 班长 (3 people)
banzhang = [
    {'name': '曾瑾', 'id': '75103', 'primary': '长白'},
    {'name': '李玲', 'id': '62416', 'primary': '长白'},
    {'name': '何丹妮', 'id': '63032', 'primary': '行政'},
]

# 专职分销 (3 people)
fenxiao = [
    {'name': '周妙笛', 'id': '32054'},
    {'name': '李珂', 'id': '63171'},
    {'name': '陈佳慧', 'id': '32014'},
]

# 专职夜班 (4 people, 2 groups, B组先上)
night_shift = [
    {'name': '王颖', 'id': '75112', 'group': 'B'},
    {'name': '王欣怡', 'id': '63898', 'group': 'B'},
    {'name': '张玉婕', 'id': '63885', 'group': 'A'},
    {'name': '魏溪', 'id': '75105', 'group': 'A'},
]

# 行政班固定 (1人, 工作日行政班 周末休, 算在线人力)
buru_ban = [
    {'name': '舒丹', 'id': '63661'},
]

# 长病假 (empty)
changbingjia = [
]

# 在线正常轮转 (25人 = 原23 + 龚格 + 李蒙)
online_staff = [
    {'name': '盛坤坤', 'id': '63715'},
    {'name': '褚婉婷', 'id': '63810'},
    {'name': '余晓芬', 'id': '63574'},
    {'name': '胡文思', 'id': '63802'},
    {'name': '余飞滔', 'id': '63610'},
    {'name': '周雨襄', 'id': '63704'},
    {'name': '杜梦薇', 'id': '63891'},
    {'name': '罗小婷', 'id': '63359'},
    {'name': '李艾洛', 'id': '63882'},
    {'name': '李蒙', 'id': '63367'},
    {'name': '贺亚兰', 'id': '63809'},
    {'name': '邵今令', 'id': '63849'},
    {'name': '马宇杰', 'id': '63805'},
    {'name': '林雪薇', 'id': '63388'},
    {'name': '李娜', 'id': '63655'},
    {'name': '陈文萍', 'id': '63874'},
    {'name': '张怡顺', 'id': '63837'},
    {'name': '张慧琳', 'id': '63680'},
    {'name': '葛宇', 'id': '63872'},
    {'name': '龚格', 'id': '63480'},
    {'name': '富惟佳', 'id': '63845'},
    {'name': '陈瑞楠', 'id': '63819'},
    {'name': '范敏', 'id': '63914'},
    {'name': '徐学宇', 'id': '63894'},
    {'name': '邹秋霞', 'id': '32007'},
]

# 每日在岗目标 (在线轮转25人, 不含行政班1人固定班)  sum=575
DAILY_TARGETS = [
    18, 17, 17, 16, 16,  # W1: 7/1 Wed-7/4 Sat-7/5 Sun
    17, 16, 17, 17, 17, 13, 15,  # W2: 7/6 Mon-7/10 Fri, 7/11 Sat, 7/12 Sun
    18, 19, 19, 21, 20, 16, 18,  # W3: 7/13 Mon-7/17 Fri, 7/18 Sat, 7/19 Sun
    19, 20, 20, 22, 21, 17, 19,  # W4: 7/20 Mon-7/24 Fri, 7/25 Sat, 7/26 Sun
    22, 22, 22, 22, 22,  # W5: 7/27 Mon-7/31 Fri
]

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def is_weekend(day):
    weekday = (JULY_1_WEEKDAY + day - 1) % 7
    return weekday >= 5

def is_saturday(day):
    return (JULY_1_WEEKDAY + day - 1) % 7 == 5

def is_sunday(day):
    return (JULY_1_WEEKDAY + day - 1) % 7 == 6

def _find_rest_blocks(rest_days_set):
    rest_list = sorted(rest_days_set)
    if not rest_list:
        return []
    blocks = []
    block_start = rest_list[0]
    prev = rest_list[0]
    for d in rest_list[1:]:
        if d == prev + 1:
            prev = d
        else:
            blocks.append(list(range(block_start, prev + 1)))
            block_start = d
            prev = d
    blocks.append(list(range(block_start, prev + 1)))
    return blocks

def _max_consecutive(work_mask):
    mx = cur = 0
    for v in work_mask:
        if v:
            cur += 1
            mx = max(mx, cur)
        else:
            cur = 0
    return mx

def _streak_stats(sched, name_key=None):
    """Return (max_work, max_rest) for a schedule list."""
    max_w = max_r = cur_w = cur_r = 0
    for s in sched:
        if s != XI:
            cur_w += 1; max_w = max(max_w, cur_w); cur_r = 0
        else:
            cur_r += 1; max_r = max(max_r, cur_r); cur_w = 0
    return max_w, max_r

# ============================================================
# 1. BANZHANG SCHEDULING
# ============================================================

def generate_banzhang():
    def zeng_changbai(day):
        return ((day - 1) % 4) in [0, 1]
    def li_changbai(day):
        return ((day - 3) % 4) in [0, 1] or day == 1

    schedules = {p['name']: [None] * NUM_DAYS for p in banzhang}
    for day in range(1, NUM_DAYS + 1):
        schedules['何丹妮'][day-1] = XING_ZHENG if not is_weekend(day) else XI
        schedules['曾瑾'][day-1] = CHANG_BAI if zeng_changbai(day) else XI
        schedules['李玲'][day-1] = CHANG_BAI if li_changbai(day) else XI
    return schedules

# ============================================================
# 2. FENXIAO SCHEDULING
# ============================================================

def generate_fenxiao():
    """
    3人轮换, 上4休2, 每人8休23上班
    每日1早早IM + 1晚一IM, 第3人休或白班机动
    加晚接早约束：晚一IM次日不接早早IM
    """
    names = ['周妙笛', '李珂', '陈佳慧']
    offsets = [0, 2, 4]
    rest_days = {}
    for i, name in enumerate(names):
        rest_days[name] = set()
        offset = offsets[i]
        for day in range(1, NUM_DAYS + 1):
            pos = (day - 1 + offset) % 6
            if pos >= 4:
                rest_days[name].add(day)

    # Remove extra rest days (10→8), smart removal
    for name in names:
        to_remove = len(rest_days[name]) - 8
        for _ in range(to_remove):
            blocks = _find_rest_blocks(rest_days[name])
            best_day, best_score = None, -1
            for blk in blocks:
                for d in blk:
                    test_rest = rest_days[name] - {d}
                    wm = [day not in test_rest for day in range(1, NUM_DAYS + 1)]
                    mw = _max_consecutive(wm)
                    score = 100 - mw * 10 + (5 if len(blk) == 1 else 0)
                    if mw <= 5:
                        score += 50
                    if score > best_score:
                        best_score, best_day = score, d
            if best_day:
                rest_days[name].remove(best_day)

    # Assign roles with counter-based balancing and late→early constraint
    schedules = {n: [XI] * NUM_DAYS for n in names}
    zao_im_cnt = {n: 0 for n in names}
    wan_im_cnt = {n: 0 for n in names}
    jidong_cnt = {n: 0 for n in names}

    for day in range(1, NUM_DAYS + 1):
        resting = [n for n in names if day in rest_days[n]]
        working = [n for n in names if day not in rest_days[n]]
        prev_shifts = {n: schedules[n][day-2] if day > 1 else None for n in names}

        if len(resting) == 1:
            # 2 work: 早早IM + 晚一IM
            r = resting[0]
            schedules[r][day-1] = XI

            if zao_im_cnt[working[0]] <= zao_im_cnt[working[1]]:
                zao_cand, wan_cand = working[0], working[1]
            else:
                zao_cand, wan_cand = working[1], working[0]

            if day > 1 and prev_shifts[zao_cand] == WAN_YI_IM:
                zao_cand, wan_cand = wan_cand, zao_cand

            schedules[zao_cand][day-1] = ZAO_ZAO_IM
            schedules[wan_cand][day-1] = WAN_YI_IM
            zao_im_cnt[zao_cand] += 1
            wan_im_cnt[wan_cand] += 1
        else:
            # 3 work: 早早IM + 晚一IM + 白班机动
            # Rotate tie-breaking by day so same person doesn't always get 机动
            offset = day % 3
            rotated = working[offset:] + working[:offset]
            rotated.sort(key=lambda n: jidong_cnt[n])
            jidong_name = rotated[0]
            others = [n for n in working if n != jidong_name]

            if day > 1 and prev_shifts[jidong_name] == WAN_YI_IM:
                for alt in others:
                    if prev_shifts[alt] != WAN_YI_IM:
                        jidong_name = alt
                        others = [n for n in working if n != jidong_name]
                        break

            if zao_im_cnt[others[0]] <= zao_im_cnt[others[1]]:
                zao_cand, wan_cand = others[0], others[1]
            else:
                zao_cand, wan_cand = others[1], others[0]

            if day > 1 and prev_shifts[zao_cand] == WAN_YI_IM:
                zao_cand, wan_cand = wan_cand, zao_cand

            schedules[jidong_name][day-1] = BAI_BAN_JD
            schedules[zao_cand][day-1] = ZAO_ZAO_IM
            schedules[wan_cand][day-1] = WAN_YI_IM
            jidong_cnt[jidong_name] += 1
            zao_im_cnt[zao_cand] += 1
            wan_im_cnt[wan_cand] += 1

    return schedules, rest_days

# ============================================================
# 3. NIGHT SHIFT SCHEDULING
# ============================================================

def generate_night_shift():
    schedules = {p['name']: [] for p in night_shift}
    for day in range(1, NUM_DAYS + 1):
        block = (day - 1) // 2
        for p in night_shift:
            if p['group'] == 'B':
                schedules[p['name']].append(DA_YE if block % 2 == 0 else XI)
            else:
                schedules[p['name']].append(DA_YE if block % 2 == 1 else XI)
    return schedules

# ============================================================
# 4. BURU BAN (哺乳班, 2 people, 工作日行政班)
# ============================================================

def generate_buru_ban():
    schedules = {}
    for p in buru_ban:
        s = []
        for day in range(1, NUM_DAYS + 1):
            s.append(XING_ZHENG if not is_weekend(day) else XI)
        schedules[p['name']] = s
    return schedules

# ============================================================
# 5. CHANGBINGJIA (长病假, 1 person, 全休)
# ============================================================

def generate_changbingjia():
    schedules = {}
    for p in changbingjia:
        schedules[p['name']] = [XI] * NUM_DAYS
    return schedules

# ============================================================
# ============================================================
# ============================================================
# ============================================================
# ============================================================
# ============================================================
# ============================================================
# ============================================================
# ============================================================
# ============================================================
# ============================================================
# ============================================================
# 6. ONLINE STAFF SCHEDULING (23 people, daily targets, 晚接早约束)
# ============================================================

def _assign_shifts_for_day(working, day, prev_shift, rest_days, fenxiao_rest_days, online_staff_list, schedules, balance=None, jd_count=None, wan2_count=None):
    """Assign shifts for one day with late→early constraint and 白班机动.

    balance: dict {name: {'early': int, 'late': int}} for cumulative balancing.
    """
    rng_day = random.Random(day * 137 + 42)
    n_workers = len(working)

    base_shifts_wd = {ZAO_BAN: 3, ZAO_SAN: 2, XING_ZHENG: 2, BAI_BAN: 2,
                      ZHONG_SAN: 1, ZHONG_SI: 1, WAN_YI: 4, WAN_ER: 2}
    base_shifts_we = {ZAO_BAN: 3, ZAO_SAN: 2, XING_ZHENG: 2, BAI_BAN: 1,
                      ZHONG_SAN: 1, ZHONG_SI: 1, WAN_YI: 4, WAN_ER: 1}

    base = base_shifts_wd if not is_weekend(day) else base_shifts_we
    base_total = sum(base.values())
    day_shifts = []
    for shift, count in base.items():
        scaled = max(1, round(count * n_workers / base_total))
        day_shifts.extend([shift] * scaled)

    while len(day_shifts) > n_workers:
        for s in [XING_ZHENG, BAI_BAN, WAN_YI, ZAO_BAN]:
            if s in day_shifts:
                day_shifts.remove(s)
                break
    while len(day_shifts) < n_workers:
        day_shifts.append(XING_ZHENG)

    late_pool = [s for s in day_shifts if s in LATE_SHIFTS]
    early_pool = [s for s in day_shifts if s not in LATE_SHIFTS]

    restricted = [p for p in working if prev_shift.get(p['name']) in LATE_SHIFTS]
    free = [p for p in working if prev_shift.get(p['name']) not in LATE_SHIFTS]

    rng_day.shuffle(restricted)
    rng_day.shuffle(late_pool)
    rng_day.shuffle(early_pool)

    # Sort free by late ratio: lower ratio → more likely to get late (balances month)
    if balance:
        def _late_ratio(name):
            b = balance.get(name, {'early': 0, 'late': 0})
            return b['late'] / max(1, b['early'] + b['late'])
        free.sort(key=lambda p: _late_ratio(p['name']))
    else:
        rng_day.shuffle(free)

    assignments = {}
    for p in restricted:
        if late_pool:
            assignments[p['name']] = late_pool.pop()
        else:
            free.append(p)

    n_late = min(len(late_pool), len(free))
    for p in free[:n_late]:
        if late_pool:
            assignments[p['name']] = late_pool.pop()

    for p in free[n_late:]:
        if early_pool:
            assignments[p['name']] = early_pool.pop()
        elif late_pool:
            assignments[p['name']] = late_pool.pop()

    # Balance WAN_ER: among late-assigned people, swap WAN_ER to those with fewer counts
    if wan2_count is not None:
        late_assigned = [(p, assignments[p['name']]) for p in working
                        if p['name'] in assignments and assignments[p['name']] in LATE_SHIFTS]
        wan_er_names = set(p['name'] for p, s in late_assigned if s == WAN_ER)
        non_wan_er = [(p, s) for p, s in late_assigned if s != WAN_ER]
        for wp_name in list(wan_er_names):
            wp = next(p for p in working if p['name'] == wp_name)
            better = [(p, s) for p, s in non_wan_er
                     if wan2_count[p['name']] < wan2_count[wp_name]]
            if better:
                swap_p, swap_s = min(better, key=lambda x: wan2_count[x[0]['name']])
                assignments[wp_name] = swap_s
                assignments[swap_p['name']] = WAN_ER
                non_wan_er = [(p, s) for p, s in non_wan_er if p['name'] != swap_p['name']]
                non_wan_er.append((wp, swap_s))
                wan_er_names.discard(wp_name)
                wan_er_names.add(swap_p['name'])
        for p in working:
            if assignments.get(p['name']) == WAN_ER:
                wan2_count[p['name']] += 1

    unassigned = [p for p in working if p['name'] not in assignments]
    leftover = early_pool + late_pool
    rng_day.shuffle(leftover)
    for p in unassigned:
        if leftover:
            assignments[p['name']] = leftover.pop()
        else:
            assignments[p['name']] = XING_ZHENG

    fenxiao_all_work = False
    if fenxiao_rest_days:
        wc = sum(1 for n in ['周妙笛', '李珂', '陈佳慧']
                if day not in fenxiao_rest_days.get(n, set()))
        fenxiao_all_work = (wc == 3)

    if not fenxiao_all_work:
        candidates = [p for p in working
                     if p['name'] in assignments
                     and prev_shift.get(p['name']) not in LATE_SHIFTS
                     and assignments[p['name']] in EARLY_SHIFTS]
        if candidates:
            # Pick candidate with fewest BAI_BAN_JD shifts so far
            if jd_count is not None:
                min_jd = min(jd_count[p['name']] for p in candidates)
                candidates = [p for p in candidates if jd_count[p['name']] == min_jd]
                chosen = rng_day.choice(candidates)
                jd_count[chosen['name']] += 1
            else:
                chosen = rng_day.choice(candidates)
            assignments[chosen['name']] = BAI_BAN_JD

    for p in working:
        if p['name'] in assignments:
            schedules[p['name']][day-1] = assignments[p['name']]
        else:
            schedules[p['name']][day-1] = XING_ZHENG

    # Update cumulative balance
    if balance:
        for p in working:
            name = p['name']
            shift = assignments.get(name, XING_ZHENG)
            if shift in LATE_SHIFTS:
                balance[name]['late'] += 1
            else:
                balance[name]['early'] += 1


def generate_online(fenxiao_rest_days=None):
    """
    25人在线轮转
    - Phase 1: Forward greedy with progressive rest cap
    - Phase 2: Fix rest counts
    - Phase 3: Target-matching swaps
    - Phase 4: SA fine-tuning
    - Phase 4.5: Direct maxWork repair
    - Phase 5: Final constraint fix
    - Phase 6: Shift assignment
    """
    rng = random.Random(42)
    an_rng = random.Random(67890)
    rest_target = 8
    n = len(online_staff)

    # Pre-compute expected rest
    cum_rest = 0
    expected_rest = {}
    for d in range(1, NUM_DAYS + 1):
        cum_rest += n - DAILY_TARGETS[d-1]
        expected_rest[d] = cum_rest / n

    # ---- Phase 1: Greedy forward with progressive cap ----
    rest_days = {p['name']: set() for p in online_staff}
    work_streak = {p['name']: 0 for p in online_staff}
    rest_streak = {p['name']: 0 for p in online_staff}

    for day in range(1, NUM_DAYS + 1):
        target_rest = n - DAILY_TARGETS[day-1]

        max_rest_by_now = min(rest_target, int(expected_rest[day]) + 2)

        scored = []
        for p in online_staff:
            name = p['name']
            if len(rest_days[name]) >= rest_target:
                continue
            if rest_streak[name] >= 2:
                continue
            if len(rest_days[name]) >= max_rest_by_now:
                continue

            wstreak = work_streak[name]
            must_rest = wstreak >= 5
            deficit = expected_rest[day] - len(rest_days[name])
            block_bonus = 200 if (day > 1 and (day - 1) in rest_days[name]) else 0
            ahead = len(rest_days[name]) - expected_rest[day]
            ahead_penalty = int(max(0, ahead - 0.0) * 500)

            score = (100000 if must_rest else 0)
            score += int(deficit * 400)
            # Stronger push for 4+ consecutive work days
            if wstreak >= 4:
                score += (2 ** wstreak) * 10  # 4→160, 5→320, 6→640
            else:
                score += wstreak * 40
            score += block_bonus
            score -= ahead_penalty
            score += rng.randint(0, 20)
            scored.append((score, name))

        scored.sort(key=lambda x: -x[0])

        day_rest = set()
        for _, name in scored:
            if len(day_rest) >= target_rest:
                break
            day_rest.add(name)

        # Fill shortfall
        if len(day_rest) < target_rest:
            remaining = sorted(
                [p for p in online_staff
                 if p['name'] not in day_rest
                 and len(rest_days[p['name']]) < rest_target
                 and rest_streak[p['name']] < 2
                 and len(rest_days[p['name']]) < max_rest_by_now],
                key=lambda p: work_streak[p['name']],
                reverse=True
            )
            for p in remaining:
                if len(day_rest) >= target_rest:
                    break
                day_rest.add(p['name'])

        for p in online_staff:
            name = p['name']
            if name in day_rest:
                rest_days[name].add(day)
                work_streak[name] = 0
                rest_streak[name] += 1
            else:
                work_streak[name] += 1
                rest_streak[name] = 0

    # ---- Phase 1.5: Break excessive work streaks ----
    # W1-W4 (d<=26): maxWork > 5 is bad. W5 (d>=27): maxWork > 7 is bad.
    for p in online_staff:
        name = p['name']
        wm = [d not in rest_days[name] for d in range(1, NUM_DAYS + 1)]
        # Find the worst streak that violates constraints
        cur = 0
        worst_start, worst_end, worst_len = 0, 0, 0
        worst_severity = 0
        for i, w in enumerate(wm):
            if w:
                cur += 1
            else:
                cur = 0
            day = i + 1
            in_w5 = day >= 27
            max_allowed = 6 if in_w5 else 5
            if cur > max_allowed:
                severity = cur - max_allowed
                if severity > worst_severity:
                    worst_severity = severity
                    worst_start = day - cur + 1
                    worst_end = day
                    worst_len = cur
        if worst_severity == 0:
            continue
        # Remove the last rest day before this streak
        rest_before = [d for d in sorted(rest_days[name]) if d < worst_start]
        if rest_before:
            rest_days[name].remove(rest_before[-1])

    # ---- Phase 2: Fix rest counts ----
    # Bias: remove early rest, add late rest → reduces W5 streaks
    for p in online_staff:
        name = p['name']
        while len(rest_days[name]) > rest_target:
            blocks = _find_rest_blocks(rest_days[name])
            candidates = []
            for blk in blocks:
                for d in blk:
                    test_rest = rest_days[name] - {d}
                    wm = [day not in test_rest for day in range(1, NUM_DAYS + 1)]
                    mw = _max_consecutive(wm)
                    dw = sum(1 for pp in online_staff if d not in rest_days[pp['name']])
                    candidates.append((mw, DAILY_TARGETS[d-1] - dw, d))
            candidates.sort(key=lambda x: (x[0], -x[1], x[2]))
            rest_days[name].remove(candidates[0][2])

        while len(rest_days[name]) < rest_target:
            for d in range(NUM_DAYS, 0, -1):
                if d in rest_days[name]:
                    continue
                test_rest = rest_days[name] | {d}
                rm = [day in test_rest for day in range(1, NUM_DAYS + 1)]
                wm = [day not in test_rest for day in range(1, NUM_DAYS + 1)]
                if _max_consecutive(rm) <= 2 and _max_consecutive(wm) <= 5:
                    current_rest = sum(1 for pp in online_staff if d in rest_days[pp['name']])
                    target_rest_day = n - DAILY_TARGETS[d-1]
                    max_extra = 1 if day >= 27 else 2
                    if current_rest < target_rest_day + max_extra:
                        rest_days[name].add(d)
                        break
            else:
                best_d, best_mw = None, 999
                for d in range(NUM_DAYS, 0, -1):
                    if d in rest_days[name]:
                        continue
                    test = rest_days[name] | {d}
                    wm = [day not in test for day in range(1, NUM_DAYS + 1)]
                    rm = [day in test for day in range(1, NUM_DAYS + 1)]
                    mw = _max_consecutive(wm)
                    current_rest = sum(1 for pp in online_staff if d in rest_days[pp['name']])
                    target_rest_day = n - DAILY_TARGETS[d-1]
                    max_extra = 1 if day >= 27 else 2
                    if mw <= 7 and _max_consecutive(rm) <= 2 and current_rest < target_rest_day + max_extra:
                        rest_days[name].add(d)
                        break
                    if mw < best_mw:
                        best_mw, best_d = mw, d
                else:
                    rest_days[name].add(best_d if best_d else
                        next(d for d in range(NUM_DAYS, 0, -1) if d not in rest_days[name]))

    # ---- Phase 2.5: Push rest into W5 for people with 0 W5 rest ----
    w5_days = list(range(27, 32))
    for p in online_staff:
        name = p['name']
        w5_rest = len(rest_days[name] & set(w5_days))
        if w5_rest > 0:
            continue
        # Find last rest before W5
        early_rest = sorted(rest_days[name] - set(w5_days))
        if not early_rest:
            continue
        last_rest = early_rest[-1]
        # Swap to W5 if last rest is D25 or earlier
        if last_rest <= 25:
            # Try to find a W5 day that can take another rest
            for w5d in [28, 29, 27, 30, 31]:
                if w5d in rest_days[name]:
                    continue
                current_rest_w5d = sum(1 for pp in online_staff if w5d in rest_days[pp['name']])
                target_rest_w5d = n - DAILY_TARGETS[w5d-1]
                if current_rest_w5d < target_rest_w5d + 3:
                    # Check constraints
                    test_rest = (rest_days[name] - {last_rest}) | {w5d}
                    rm = [d in test_rest for d in range(1, NUM_DAYS + 1)]
                    wm = [d not in test_rest for d in range(1, NUM_DAYS + 1)]
                    if _max_consecutive(rm) <= 2 and _max_consecutive(wm) <= 6:
                        rest_days[name].remove(last_rest)
                        rest_days[name].add(w5d)
                        break

    # ---- Phase 3: Target-matching swaps ----
    def daily_workers(rd):
        return [sum(1 for p in online_staff if d not in rd[p['name']])
                for d in range(1, NUM_DAYS + 1)]

    for _ in range(800):
        dw = daily_workers(rest_days)
        gaps = [(d, dw[d-1] - DAILY_TARGETS[d-1]) for d in range(1, NUM_DAYS + 1)]
        overs = [(d, g) for d, g in gaps if g >= 2]
        unders = [(d, g) for d, g in gaps if g <= -2]
        if not overs or not unders:
            break
        overs.sort(key=lambda x: -x[1])
        unders.sort(key=lambda x: x[1])
        improved = False
        for over_d, _ in overs:
            for under_d, _ in unders:
                if improved:
                    break
                for p in online_staff:
                    name = p['name']
                    if under_d in rest_days[name] and over_d not in rest_days[name]:
                        test_rest = (rest_days[name] - {under_d}) | {over_d}
                        wm = [d not in test_rest for d in range(1, NUM_DAYS + 1)]
                        rm = [d in test_rest for d in range(1, NUM_DAYS + 1)]
                        mw, mr = _max_consecutive(wm), _max_consecutive(rm)
                        max_mw = 7 if over_d >= 27 else 5
                        if mw <= max_mw and mr <= 2:
                            rest_days[name].remove(under_d)
                            rest_days[name].add(over_d)
                            improved = True
                            break
        if not improved:
            break

    # ---- Phase 4: Simulated annealing ----
    def quick_energy(rd):
        e = 0
        dw = daily_workers(rd)
        for d in range(1, NUM_DAYS + 1):
            diff = dw[d-1] - DAILY_TARGETS[d-1]
            e += abs(diff) * 500

        for p in online_staff:
            name = p['name']
            wm = [d not in rd[name] for d in range(1, NUM_DAYS + 1)]
            mw = _max_consecutive(wm)
            if mw > 5:
                excess = mw - 5
                cur = 0
                streak_end = 0
                for i, w in enumerate(wm):
                    if w:
                        if cur == 0:
                            cur += 1
                        else:
                            cur += 1
                        if cur == mw:
                            streak_end = i + 1
                    else:
                        cur = 0
                if streak_end >= 27:
                    if mw > 7:
                        e += (mw - 7) * (mw - 7) * 500 + excess * excess * 2000
                    else:
                        e += excess * excess * 2000
                else:
                    e += excess * excess * 3000
            rm = [d in rd[name] for d in range(1, NUM_DAYS + 1)]
            mr = _max_consecutive(rm)
            if mr > 2:
                e += 100000
            if mr < 2 and len(rd[name]) >= 4:
                e += 2000
            if len(rd[name]) != rest_target:
                e += abs(len(rd[name]) - rest_target) * 10000
        return e

    current_energy = quick_energy(rest_days)
    best_energy = current_energy
    best_rest = {name: set(s) for name, s in rest_days.items()}

    T = 2000.0
    people_names = [p['name'] for p in online_staff]

    for it in range(40000):
        p1 = an_rng.choice(people_names)
        if not rest_days[p1]:
            continue
        d1 = an_rng.choice(sorted(rest_days[p1]))

        p2 = an_rng.choice(people_names)
        if p2 == p1:
            continue
        work_days = [d for d in range(1, NUM_DAYS + 1) if d not in rest_days[p2]]
        if not work_days:
            continue
        d2 = an_rng.choice(work_days)

        if d1 == d2 or d2 in rest_days[p1] or d1 in rest_days[p2]:
            continue

        new_p1 = (rest_days[p1] - {d1}) | {d2}
        new_p2 = rest_days[p2] | {d1}

        if (_max_consecutive([d in new_p1 for d in range(1, NUM_DAYS + 1)]) > 2 or
            _max_consecutive([d in new_p2 for d in range(1, NUM_DAYS + 1)]) > 2):
            continue
        if (_max_consecutive([d not in new_p1 for d in range(1, NUM_DAYS + 1)]) > 7 or
            _max_consecutive([d not in new_p2 for d in range(1, NUM_DAYS + 1)]) > 7):
            continue
        if len(new_p2) > rest_target:
            continue

        new_rd = {name: set(s) for name, s in rest_days.items()}
        new_rd[p1] = new_p1
        new_rd[p2] = new_p2
        new_energy = quick_energy(new_rd)

        delta = new_energy - current_energy
        if delta < 0 or (T > 0.01 and an_rng.random() < 2.71828 ** (-delta / T)):
            rest_days[p1] = new_p1
            rest_days[p2] = new_p2
            current_energy = new_energy

            if current_energy < best_energy:
                best_energy = current_energy
                best_rest = {name: set(s) for name, s in rest_days.items()}

        T *= 0.9995
        if T < 1.0:
            T = 2000.0 * (0.5 + 0.5 * an_rng.random())
            if best_energy == 0:
                break

    rest_days = best_rest

    # ---- Phase 4.5: Direct maxWork repair ----
    # For people with maxWork > 7 (W5) or > 5 (W1-W4), try to break longest streak
    for _ in range(500):
        # Find person with worst maxWork violation
        worst_name = None
        worst_excess = 0
        for p in online_staff:
            name = p['name']
            wm = [d not in rest_days[name] for d in range(1, NUM_DAYS + 1)]
            mw = _max_consecutive(wm)
            # Find longest streak end
            cur = 0
            longest_end = 0
            for i, w in enumerate(wm):
                if w:
                    cur += 1
                    if cur == mw:
                        longest_end = i + 1
                else:
                    cur = 0
            threshold = 7 if longest_end >= 27 else 5
            excess = mw - threshold
            if excess > worst_excess:
                worst_excess = excess
                worst_name = name

        if worst_name is None or worst_excess <= 0:
            break

        # Get streak info for worst person
        wm = [d not in rest_days[worst_name] for d in range(1, NUM_DAYS + 1)]
        mw = _max_consecutive(wm)
        cur = 0
        longest_start = 1
        longest_end = 1
        for i, w in enumerate(wm):
            if w:
                if cur == 0:
                    ss = i + 1
                cur += 1
                if cur == mw:
                    longest_start = ss
                    longest_end = i + 1
            else:
                cur = 0

        # Try inserting rest at different positions within the streak
        best_swap = None
        for offset_frac in [0.3, 0.5, 0.7]:
            target = int(longest_start + (longest_end - longest_start) * offset_frac)
            # Find someone resting on 'target' who can swap
            for pp in online_staff:
                if pp['name'] == worst_name:
                    continue
                if target not in rest_days[pp['name']]:
                    continue
                # Find a rest day of worst_name before the streak to give to pp
                for swap_out in sorted(rest_days[worst_name]):
                    if swap_out >= longest_start:
                        break
                    if swap_out in rest_days[pp['name']]:
                        continue
                    # Try swap
                    new_worst = (rest_days[worst_name] - {swap_out}) | {target}
                    new_pp = (rest_days[pp['name']] - {target}) | {swap_out}
                    ok = True
                    for nm, nr in [(worst_name, new_worst), (pp['name'], new_pp)]:
                        nwm = [d not in nr for d in range(1, NUM_DAYS + 1)]
                        nrm = [d in nr for d in range(1, NUM_DAYS + 1)]
                        nmw = _max_consecutive(nwm)
                        nmr = _max_consecutive(nrm)
                        # Check: find if new longest streak ends in W5
                        cur2 = 0
                        ne = 0
                        for i, w in enumerate(nwm):
                            if w:
                                cur2 += 1
                                if cur2 == nmw:
                                    ne = i + 1
                            else:
                                cur2 = 0
                        threshold2 = 7 if ne >= 27 else 5
                        if nmw > threshold2 or nmr > 2 or len(nr) != rest_target:
                            ok = False
                            break
                    if ok:
                        best_swap = (pp['name'], swap_out, target)
                        break
                if best_swap:
                    break
            if best_swap:
                break

        if best_swap:
            pp_name, swap_out, target = best_swap
            rest_days[worst_name].remove(swap_out)
            rest_days[worst_name].add(target)
            rest_days[pp_name].remove(target)
            rest_days[pp_name].add(swap_out)
        else:
            break  # can't fix anyone, stop trying

    # ---- Phase 4.6: W5 streak breaking via targeted swaps ----
    # For people with maxWork>5 ending in W5, try to swap a rest into W5
    for _ in range(500):
        victims = []
        for p in online_staff:
            name = p['name']
            wm = [d not in rest_days[name] for d in range(1, NUM_DAYS + 1)]
            mw = _max_consecutive(wm)
            if mw <= 5:
                continue
            cur = 0
            end = 0
            start = 1
            for i, w in enumerate(wm):
                if w: cur += 1
                else: cur = 0
                if cur == mw: end = i + 1
            if end >= 27:
                # Find start of this streak
                cur2 = 0
                for i in range(end - 1, -1, -1):
                    if wm[i]:
                        cur2 += 1
                        if cur2 == mw:
                            start = i + 1
                    else:
                        break
                victims.append((mw, start, end, name))
        if not victims:
            break
        victims.sort(reverse=True)
        _, v_start, v_end, worst = victims[0]
        old_mw = victims[0][0]

        # Try all rest days before the streak for swapping
        best_swap = None
        for swap_out in sorted(rest_days[worst], reverse=True):
            if swap_out >= v_start:
                continue
            if best_swap: break
            # Target: insert rest near the middle of the streak
            target = v_start + (v_end - v_start) // 2
            for offset in [0, -1, 1, -2, 2]:
                t = target + offset
                if t < v_start or t > v_end: continue
                if t in rest_days[worst]: continue
                if best_swap: break
                # Find donor resting on day t
                for pp in online_staff:
                    if pp['name'] == worst: continue
                    if t not in rest_days[pp['name']]: continue
                    if swap_out in rest_days[pp['name']]: continue
                    new_worst = (rest_days[worst] - {swap_out}) | {t}
                    new_donor = (rest_days[pp['name']] - {t}) | {swap_out}
                    ok = True
                    for nm, nr in [(worst, new_worst), (pp['name'], new_donor)]:
                        nwm = [d not in nr for d in range(1, NUM_DAYS + 1)]
                        nrm = [d in nr for d in range(1, NUM_DAYS + 1)]
                        nmw = _max_consecutive(nwm)
                        nmr = _max_consecutive(nrm)
                        if nm == worst:
                            if nmw >= old_mw:  # Must reduce maxWork
                                ok = False; break
                        else:
                            if nmw > 7:  # Donor must stay ≤ 7
                                ok = False; break
                        if nmr > 2 or len(nr) != rest_target:
                            ok = False; break
                    if ok:
                        best_swap = (pp['name'], swap_out, t)
                        break

        if best_swap:
            pp_name, swap_out, target = best_swap
            rest_days[worst].remove(swap_out)
            rest_days[worst].add(target)
            rest_days[pp_name].remove(target)
            rest_days[pp_name].add(swap_out)

    # ---- Phase 4.6: Break remaining long work streaks (≥6 days) ----
    # Iterate until no more fixes
    for _ in range(50):
        any_fixed = False
        for p in online_staff:
            name = p['name']
            wm = [d not in rest_days[name] for d in range(1, NUM_DAYS + 1)]
            mw = _max_consecutive(wm)
            if mw < 6:
                continue
            # Find the longest streak
            cur = 0
            best_start, best_end, best_len = 0, 0, 0
            for i, w in enumerate(wm):
                if w: cur += 1
                else: cur = 0
                if cur > best_len:
                    best_len = cur
                    best_end = i + 1
                    best_start = best_end - cur + 1
            streak_days = list(range(best_start, best_end + 1))
            # Try each work day in the streak as candidate to swap out
            found = False
            for work_day in streak_days:
                if found: break
                # Find rest days in W1-W4 (D1-D26) to swap in
                early_rest = sorted([d for d in rest_days[name] if d <= 26], reverse=True)
                for rest_day in early_rest:
                    if found: break
                    if rest_day in streak_days:
                        continue
                    test_rest = (rest_days[name] - {rest_day}) | {work_day}
                    rm = [d in test_rest for d in range(1, NUM_DAYS + 1)]
                    wm2 = [d not in test_rest for d in range(1, NUM_DAYS + 1)]
                    if _max_consecutive(rm) > 3:
                        continue
                    new_mw = _max_consecutive(wm2)
                    if new_mw >= 6:
                        continue
                    # Check daily target impact
                    dw = daily_workers(rest_days)
                    new_wd_workers = dw[work_day - 1] - 1
                    new_rd_workers = dw[rest_day - 1] + 1
                    if abs(new_wd_workers - DAILY_TARGETS[work_day - 1]) > 3:
                        continue
                    if abs(new_rd_workers - DAILY_TARGETS[rest_day - 1]) > 3:
                        continue
                    rest_days[name].remove(rest_day)
                    rest_days[name].add(work_day)
                    found = True
                    any_fixed = True
        if not any_fixed:
            break

    # ---- Phase 5: Final rest count fix ----
    for p in online_staff:
        name = p['name']
        while len(rest_days[name]) > rest_target:
            blocks = _find_rest_blocks(rest_days[name])
            candidates = []
            for blk in blocks:
                for d in blk:
                    test_rest = rest_days[name] - {d}
                    candidates.append((_max_consecutive([day not in test_rest for day in range(1, NUM_DAYS + 1)]), d))
            candidates.sort()
            rest_days[name].remove(candidates[0][1])

    # Final target-matching
    for _ in range(400):
        dw = daily_workers(rest_days)
        gaps = [(d, dw[d-1] - DAILY_TARGETS[d-1]) for d in range(1, NUM_DAYS + 1)]
        overs = [(d, g) for d, g in gaps if g >= 2]
        unders = [(d, g) for d, g in gaps if g <= -2]
        if not overs or not unders:
            break
        overs.sort(key=lambda x: -x[1])
        unders.sort(key=lambda x: x[1])
        improved = False
        for over_d, _ in overs:
            for under_d, _ in unders:
                if improved:
                    break
                for p in online_staff:
                    name = p['name']
                    if under_d in rest_days[name] and over_d not in rest_days[name]:
                        test_rest = (rest_days[name] - {under_d}) | {over_d}
                        wm = [d not in test_rest for d in range(1, NUM_DAYS + 1)]
                        rm = [d in test_rest for d in range(1, NUM_DAYS + 1)]
                        mw, mr = _max_consecutive(wm), _max_consecutive(rm)
                        max_mw = 7 if over_d >= 27 else 5
                        if mw <= max_mw and mr <= 2:
                            rest_days[name].remove(under_d)
                            rest_days[name].add(over_d)
                            improved = True
                            break
        if not improved:
            break

    # Final target-matching with threshold=1 to fix small gaps
    for _ in range(400):
        dw = daily_workers(rest_days)
        gaps = [(d, dw[d-1] - DAILY_TARGETS[d-1]) for d in range(1, NUM_DAYS + 1)]
        overs = [(d, g) for d, g in gaps if g >= 1]
        unders = [(d, g) for d, g in gaps if g <= -1]
        if not overs or not unders:
            break
        overs.sort(key=lambda x: -x[1])
        unders.sort(key=lambda x: x[1])
        improved = False
        for over_d, _ in overs:
            for under_d, _ in unders:
                if improved:
                    break
                for p in online_staff:
                    name = p['name']
                    if under_d in rest_days[name] and over_d not in rest_days[name]:
                        test_rest = (rest_days[name] - {under_d}) | {over_d}
                        wm = [d not in test_rest for d in range(1, NUM_DAYS + 1)]
                        rm = [d in test_rest for d in range(1, NUM_DAYS + 1)]
                        mw, mr = _max_consecutive(wm), _max_consecutive(rm)
                        max_mw = 7 if over_d >= 27 else 5
                        if mw <= max_mw and mr <= 2:
                            rest_days[name].remove(under_d)
                            rest_days[name].add(over_d)
                            improved = True
                            break
        if not improved:
            break

    # ---- Phase 6: Assign shifts ----
    schedules = {p['name']: [XI] * NUM_DAYS for p in online_staff}
    prev_shift = {p['name']: None for p in online_staff}
    balance = {p['name']: {'early': 0, 'late': 0} for p in online_staff}
    jd_count = {p['name']: 0 for p in online_staff}
    wan2_count = {p['name']: 0 for p in online_staff}

    for day in range(1, NUM_DAYS + 1):
        working = [p for p in online_staff if day not in rest_days[p['name']]]
        _assign_shifts_for_day(working, day, prev_shift, rest_days,
                              fenxiao_rest_days, online_staff, schedules, balance, jd_count, wan2_count)
        for p in working:
            prev_shift[p['name']] = schedules[p['name']][day-1]
        for p in online_staff:
            if day in rest_days[p['name']]:
                prev_shift[p['name']] = None

    # Post-fix late→early
    for day in range(1, NUM_DAYS):
        for p in online_staff:
            name = p['name']
            if schedules[name][day-1] in LATE_SHIFTS and schedules[name][day] in EARLY_SHIFTS:
                next_working = [pp for pp in online_staff
                               if day+1 not in rest_days[pp['name']] and pp['name'] != name]
                for pp in next_working:
                    if schedules[pp['name']][day] in LATE_SHIFTS:
                        if day == 1 or schedules[pp['name']][day-1] not in LATE_SHIFTS:
                            schedules[name][day], schedules[pp['name']][day] = \
                                schedules[pp['name']][day], schedules[name][day]
                            break

    return schedules













# ============================================================
# 7. CREATE EXCEL
# ============================================================

def create_excel():
    print("Generating schedules...")

    banzhang_sched = generate_banzhang()
    fenxiao_sched, fenxiao_rest = generate_fenxiao()
    night_sched = generate_night_shift()
    buru_sched = generate_buru_ban()
    changbingjia_sched = generate_changbingjia()
    online_sched = generate_online(fenxiao_rest)

    print("Creating Excel workbook...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '7月班'

    # Styles
    header_font = Font(name='微软雅黑', size=8, bold=True, color='FF000000')
    data_font = Font(name='微软雅黑', size=8, color='FF000000')
    bold_font = Font(name='微软雅黑', size=8, bold=True, color='FF000000')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    center_align_nowrap = Alignment(horizontal='center', vertical='center')

    shift_colors = {
        XI: 'FFC000', QING_JIA: 'FFC000',
        XING_ZHENG: 'FFF3E0', ZAO_BAN: 'E3F2FD', ZAO_SAN: 'F3E5F5',
        BAI_BAN: 'E8F5E9', ZHONG_SAN: 'E0F7FA', ZHONG_SI: 'FFF8E1',
        WAN_YI: 'FCE4EC', WAN_ER: 'EDE7F6', DA_YE: 'FFFFFF',
        ZAO_ZAO_IM: 'FFFF00', WAN_YI_IM: 'FFFF00', BAI_BAN_JD: 'EE822F',
        CHANG_BAI: 'FFFFFF',
    }

    row = 1

    # ---- Row 1: Date headers ----
    ws.cell(row=row, column=1, value='姓名')
    ws.cell(row=row, column=2, value='工号')
    ws.cell(row=row, column=3, value='')  # C column is blank in user's version
    for day in range(1, NUM_DAYS + 1):
        col = 3 + day
        dt = datetime(JULY_YEAR, JULY_MONTH, day)
        cell = ws.cell(row=row, column=col, value=dt)
        cell.number_format = 'm/d;@'
        # Weekend dates in RED (Sat=5, Sun=6 in weekday())
        is_weekend_day = (JULY_1_WEEKDAY + day - 1) % 7 >= 5
        if is_weekend_day:
            cell.font = Font(name='微软雅黑', size=8, bold=True, color='FFFF0000')
        else:
            cell.font = Font(name='微软雅黑', size=8, bold=True, color='FF000000')
        cell.alignment = center_align

    # Person-level summary headers (columns AI-AU)
    summary_start = 3 + NUM_DAYS + 1  # = column 35 = AI
    summary_headers = [
        XI, XING_ZHENG, BAI_BAN_JD, WAN_ER,
        ZAO_ZAO_IM, ZAO_BAN, ZAO_SAN, BAI_BAN,
        ZHONG_SAN, ZHONG_SI, WAN_YI, WAN_ER, DA_YE,
    ]
    # Note: AL1=晚二(重复), AT1=晚二(重复) - matching user's file exactly
    for si, sh in enumerate(summary_headers):
        cell = ws.cell(row=row, column=summary_start + si, value=sh)
        cell.font = header_font
        cell.alignment = center_align

    # Apply header font/alignment to A1, B1
    for c in [1, 2]:
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.alignment = center_align

    ws.row_dimensions[1].height = 27

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 6
    for day in range(1, NUM_DAYS + 1):
        ws.column_dimensions[get_column_letter(3 + day)].width = 6
    for si in range(len(summary_headers)):
        ws.column_dimensions[get_column_letter(summary_start + si)].width = 6

    row += 1

    # ---- Helper: write a person row ----
    def write_person_row(name, emp_id, shifts_list, role_label=''):
        nonlocal row
        ws.row_dimensions[row].height = 33
        ws.cell(row=row, column=1, value=f' {name}')
        ws.cell(row=row, column=2, value=emp_id)
        ws.cell(row=row, column=3, value=role_label)
        for day in range(1, NUM_DAYS + 1):
            col = 3 + day
            shift = shifts_list[day - 1]
            cell = ws.cell(row=row, column=col, value=shift)
            cell.alignment = center_align
            cell.font = data_font
            if shift in shift_colors:
                cell.fill = PatternFill(start_color=shift_colors[shift],
                                       end_color=shift_colors[shift],
                                       fill_type='solid')
        # COUNTIF formulas in summary columns (AI-AU)
        erow = row  # current row
        d_col = get_column_letter(4)   # D
        e_col = get_column_letter(5)   # E
        ah_col = get_column_letter(3 + NUM_DAYS)  # AH
        d_range = f'{d_col}{erow}:{ah_col}{erow}'
        e_range = f'{e_col}{erow}:{ah_col}{erow}'
        formulas = [
            (summary_start, f'=COUNTIF({d_range},"{XI}")+COUNTIF({d_range},"{QING_JIA}")'),
            (summary_start + 1, f'=COUNTIF({e_range},"{XING_ZHENG}")'),
            (summary_start + 2, f'=COUNTIF({e_range},"{BAI_BAN_JD}")'),
            (summary_start + 3, f'=COUNTIF({e_range},"{WAN_ER}")'),
            (summary_start + 4, f'=COUNTIF({d_range},"{ZAO_ZAO_IM}")'),
            (summary_start + 5, f'=COUNTIF({d_range},"{ZAO_BAN}")'),
            (summary_start + 6, f'=COUNTIF({d_range},"{ZAO_SAN}")'),
            (summary_start + 7, f'=COUNTIF({d_range},"{BAI_BAN}")'),
            (summary_start + 8, f'=COUNTIF({d_range},"{ZHONG_SAN}")'),
            (summary_start + 9, f'=COUNTIF({d_range},"{ZHONG_SI}")'),
            (summary_start + 10, f'=COUNTIF({d_range},"{WAN_YI}")+COUNTIF({d_range},"{WAN_YI_IM}")'),
            (summary_start + 11, f'=COUNTIF({d_range},"{WAN_ER}")'),
            (summary_start + 12, f'=COUNTIF({d_range},"{DA_YE}")'),
        ]
        for col_idx, formula in formulas:
            cell = ws.cell(row=row, column=col_idx, value=formula)
            cell.alignment = center_align
            cell.font = data_font
        for c in [1, 2, 3]:
            ws.cell(row=row, column=c).font = data_font
            ws.cell(row=row, column=c).alignment = center_align
        row += 1

    def write_empty_row():
        nonlocal row
        row += 1

    # ---- Write rows in order ----
    # 班长
    for name in ['曾瑾', '李玲', '何丹妮']:
        p = next(p for p in banzhang if p['name'] == name)
        write_person_row(p['name'], p['id'], banzhang_sched[name], '班长')
    write_empty_row()

    # 专职分销
    fx_start_row = row
    for name in ['周妙笛', '李珂', '陈佳慧']:
        p = next(p for p in fenxiao if p['name'] == name)
        write_person_row(p['name'], p['id'], fenxiao_sched[name], '专职分销')
    fx_end_row = row - 1
    write_empty_row()

    # 专职夜班
    night_start_row = row
    for name in ['王颖', '王欣怡', '张玉婕', '魏溪']:
        p = next(p for p in night_shift if p['name'] == name)
        write_person_row(p['name'], p['id'], night_sched[name], '专职大夜')
    night_end_row = row - 1
    write_empty_row()

    # 行政班 (舒丹, 算在线人力)
    buru_start_row = row
    for p in buru_ban:
        write_person_row(p['name'], p['id'], buru_sched[p['name']], '在线')
    buru_end_row = row - 1
    write_empty_row()

    # 在线 (邹秋霞在内)
    online_start_row = row
    for p in online_staff:
        write_person_row(p['name'], p['id'], online_sched[p['name']], '在线')
    online_end_row = row - 1

    write_empty_row()
    write_empty_row()

    # ---- Summary Group 1: online regular shifts (COUNTIF formulas) ----
    # Range: all staff from fenxiao through online (D6:AH{online_end_row})
    all_staff_range = f"{{col}}{fx_start_row}:{{col}}{online_end_row}"
    # 大夜 range: from night shift start
    night_range = f"{{col}}{night_start_row}:{{col}}{online_end_row}"

    group1_shifts = [ZAO_BAN, ZAO_SAN, XING_ZHENG, BAI_BAN,
                     ZHONG_SAN, ZHONG_SI, WAN_YI, WAN_ER]
    group1_start_row = row
    for shift_type in group1_shifts:
        ws.row_dimensions[row].height = 33
        ws.cell(row=row, column=3, value=shift_type).font = data_font
        ws.cell(row=row, column=3).alignment = center_align
        for day in range(1, NUM_DAYS + 1):
            col = 3 + day
            col_letter = get_column_letter(col)
            formula = f'=COUNTIF({all_staff_range.format(col=col_letter)},"{shift_type}")'
            cell = ws.cell(row=row, column=col, value=formula)
            cell.alignment = center_align
            cell.font = data_font
        row += 1

    # 大夜 row
    ws.row_dimensions[row].height = 33
    ws.cell(row=row, column=3, value=DA_YE).font = data_font
    ws.cell(row=row, column=3).alignment = center_align
    for day in range(1, NUM_DAYS + 1):
        col = 3 + day
        col_letter = get_column_letter(col)
        formula = f'=COUNTIF({night_range.format(col=col_letter)},"{DA_YE}")'
        cell = ws.cell(row=row, column=col, value=formula)
        cell.alignment = center_align
        cell.font = data_font
    row += 1

    # 合计 row for group 1 (SUM over 早班~晚二, not 大夜)
    ws.row_dimensions[row].height = 33
    ws.cell(row=row, column=3, value='合计').font = bold_font
    ws.cell(row=row, column=3).alignment = center_align
    for day in range(1, NUM_DAYS + 1):
        col = 3 + day
        col_letter = get_column_letter(col)
        formula = f'=SUM({col_letter}{group1_start_row}:{col_letter}{row - 2})'
        cell = ws.cell(row=row, column=col, value=formula)
        cell.alignment = center_align
        cell.font = bold_font
    row += 1

    write_empty_row()

    # ---- Summary Group 2: IM / 机动 shifts (COUNTIF formulas) ----
    # Range: from fx start to online end
    fx_range = f"{{col}}{fx_start_row}:{{col}}{online_end_row}"

    group2_shifts = [ZAO_ZAO_IM, WAN_YI_IM, BAI_BAN_JD]
    group2_start_row = row
    for shift_type in group2_shifts:
        ws.row_dimensions[row].height = 33
        ws.cell(row=row, column=3, value=shift_type).font = data_font
        ws.cell(row=row, column=3).alignment = center_align
        for day in range(1, NUM_DAYS + 1):
            col = 3 + day
            col_letter = get_column_letter(col)
            formula = f'=COUNTIF({fx_range.format(col=col_letter)},"{shift_type}")'
            cell = ws.cell(row=row, column=col, value=formula)
            cell.alignment = center_align
            cell.font = data_font
        row += 1

    # 合计 row for group 2
    ws.row_dimensions[row].height = 33
    ws.cell(row=row, column=3, value='合计').font = bold_font
    ws.cell(row=row, column=3).alignment = center_align
    for day in range(1, NUM_DAYS + 1):
        col = 3 + day
        col_letter = get_column_letter(col)
        formula = f'=SUM({col_letter}{group2_start_row}:{col_letter}{row - 1})'
        cell = ws.cell(row=row, column=col, value=formula)
        cell.alignment = center_align
        cell.font = bold_font

    # ---- Verification ----
    print("\n=== Verification ===")
    all_people = (
        [(p, banzhang_sched[p['name']], '班长') for p in banzhang] +
        [(p, fenxiao_sched[p['name']], '分销') for p in fenxiao] +
        [(p, night_sched[p['name']], '夜班') for p in night_shift] +
        [(p, buru_sched[p['name']], '行政班(在线)') for p in buru_ban] +
        [(p, changbingjia_sched[p['name']], '长病假') for p in changbingjia] +
        [(p, online_sched[p['name']], '在线') for p in online_staff]
    )

    for person, sched, cat in all_people:
        rest_count = sum(1 for s in sched if s == XI)
        work_count = NUM_DAYS - rest_count
        mw, mr = _streak_stats(sched)
        flag = ''
        if cat in ('在线', '分销') and mw > 5:
            flag = f' *** MAXWORK={mw}'
        if cat in ('在线', '分销') and mr > 2:
            flag = f' *** MAXREST={mr}'
        # Check late→early
        late_early_violations = 0
        for d in range(1, NUM_DAYS):
            if sched[d-1] in LATE_SHIFTS and sched[d] in EARLY_SHIFTS:
                late_early_violations += 1
        if late_early_violations > 0:
            flag += f' LATE->EARLY x{late_early_violations}'
        shift_info = ''
        if cat == '分销':
            zao = sum(1 for s in sched if s == ZAO_ZAO_IM)
            wan = sum(1 for s in sched if s == WAN_YI_IM)
            jd = sum(1 for s in sched if s == BAI_BAN_JD)
            shift_info = f' 早早IM={zao} 晚一IM={wan} 机动={jd}'
        print(f"  [{cat}] {person['name']:6s}: 休={rest_count} 上班={work_count} maxW={mw} maxR={mr}{flag}{shift_info}")

    # Daily headcount check
    print("\n=== Daily Headcount (在线25人 + 行政班1人) ===")
    for d in range(1, NUM_DAYS + 1):
        online_n = sum(1 for p in online_staff if online_sched[p['name']][d-1] != XI)
        buru_n = sum(1 for p in buru_ban if buru_sched[p['name']][d-1] != XI)
        total = online_n + buru_n
        target = DAILY_TARGETS[d-1]
        diff = total - target
        wd = 'WD' if not is_weekend(d) else 'WE'
        bar = '***' if abs(diff) >= 3 else ''
        print(f"  7/{d:2d} {wd}: 在线={online_n:2d} +行政班={buru_n} ={total:2d} (target={target:2d}) {bar}")

    # ---- Save ----
    import os, time
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, f'{JULY_MONTH}月份班表.xlsx')
    for attempt in range(3):
        try:
            wb.save(output_path)
            print(f"\nSaved to: {output_path}")
            break
        except PermissionError:
            if attempt == 0:
                output_path = os.path.join(script_dir, f'{JULY_MONTH}月份班表_new.xlsx')
            elif attempt == 1:
                output_path = os.path.join(script_dir, f'{JULY_MONTH}月份班表_v2.xlsx')
            else:
                output_path = os.path.join(script_dir, f'{JULY_MONTH}月份班表_{int(time.time())}.xlsx')
                wb.save(output_path)
                print(f"\nSaved to: {output_path}")

    return output_path


if __name__ == '__main__':
    create_excel()
