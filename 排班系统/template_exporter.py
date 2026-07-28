"""排班系统 — 在线客服班表模版格式导出器

严格按照「在线客服班表（2026年7月）.xlsx」参考模版的格式生成。
仅在输出时转换班次名称，不修改引擎内部常量。
"""

import json
from datetime import datetime, time, timedelta
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).parent
_MAPPING_FILE = SCRIPT_DIR / 'moka_mapping.json'

# ── 班次名称转换（引擎 → 模版） ──
SHIFT_NAME_MAP = {
    '早早班\n(IM)': '早早班',
    '晚一\n(IM)': '晚一',
    '行政班\n(机动)': '行政班',
    '大夜': '专职大夜',
    '请假': '休',  # 模版中请假归入休
}

# 统计列全量顺序（用于确定列序）
ALL_STAT_ORDER = [
    '休', '早早班', '长白班', '长白-早', '早早2', '早班', '天地班',
    '早三', '行政班', '白班', '中二', '中三', '中四',
    '晚一', '晚二', '孕妇（哺乳）行政班', '专职大夜',
]

# 班次对应工时（小时）
SHIFT_HOURS = {
    '长白班': 11.5,
    '长白-早': 12,
    '专职大夜': 10.5,
}
DEFAULT_SHIFT_HOURS = 8  # 未特指的日班

# 星期名
WEEKDAY_NAMES = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']

# ── 样式 ──
FONT_H = Font(name='微软雅黑', size=9)
FONT_D8 = Font(name='微软雅黑', size=8)
FONT_D9 = Font(name='微软雅黑', size=9)

ALIGN_C = Alignment(horizontal='center', vertical='center')
ALIGN_W = Alignment(horizontal='center', vertical='center', wrap_text=True)

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _load_mapping():
    if _MAPPING_FILE.exists():
        with open(_MAPPING_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'by_emp_id': {}, 'by_name': {}}


def _convert_shift(engine_shift):
    if engine_shift is None:
        return ''
    s = str(engine_shift).strip()
    return SHIFT_NAME_MAP.get(s, s)


def _get_moka_info(emp_id, name, mapping):
    emp_id_str = str(emp_id).strip() if emp_id else ''
    name_str = str(name).strip() if name else ''
    if emp_id_str and emp_id_str in mapping.get('by_emp_id', {}):
        info = mapping['by_emp_id'][emp_id_str]
        return info['moka_id'], info.get('name', name_str)
    if name_str and name_str in mapping.get('by_name', {}):
        info = mapping['by_name'][name_str]
        return info['moka_id'], name_str
    return '', name_str


def _set_cell(ws, row, col, value, font, alignment, number_format=None):
    """写入单元格并统一设置字体/对齐/边框/数字格式"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = alignment
    cell.border = THIN_BORDER
    if number_format:
        cell.number_format = number_format
    return cell


def create_template_excel(schedules, engine):
    mapping = _load_mapping()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{engine.month}月在线客服班表'

    num_days = engine.num_days
    year = engine.year
    month = engine.month
    start_weekday = engine.start_weekday

    # ── 收集全部人员和排班 ──
    all_people = []
    all_shifts_set = set()

    for group_key in ['banzhang', 'fenxiao', 'night_shift', 'buru_ban',
                       'buru_support', 'online']:
        group_schedule = schedules.get(group_key, {})
        group_people = (engine.online_staff if group_key == 'online'
                        else getattr(engine, group_key, []))
        for p in group_people:
            name = p['name']
            emp_id = str(p['id']).strip()
            if name in group_schedule:
                converted = [_convert_shift(s) for s in group_schedule[name]]
                all_people.append((name, emp_id, converted))
                all_shifts_set.update(converted)

    # ── 动态统计列：按全量顺序，仅保留当月出现的 ──
    active_stats = [s for s in ALL_STAT_ORDER if s in all_shifts_set]
    num_stat_cols = len(active_stats)

    # 统计列 → 工时分类
    # 长白班: 11.5h, 长白-早: 12h, 专职大夜: 10.5h, 休: 不计, 其余: 8h
    stat_eight_hour = []     # 8h 日班索引
    stat_changbai = None     # 长白班索引
    stat_changbai_zao = None  # 长白-早索引
    stat_night = None        # 专职大夜索引
    for i, s in enumerate(active_stats):
        if s == '长白班':
            stat_changbai = i
        elif s == '长白-早':
            stat_changbai_zao = i
        elif s == '专职大夜':
            stat_night = i
        elif s != '休':
            stat_eight_hour.append(i)

    # ── 列位置 ──
    C_DAY = 5
    C_STAT = C_DAY + num_days  # 统计起始列

    def stat_col(offset):
        return get_column_letter(C_STAT + offset)

    day_start_letter = get_column_letter(C_DAY)
    day_end_letter = get_column_letter(C_DAY + num_days - 1)

    # ── 班长名字（三个班长） ──
    leader_names = '\n'.join(p['name'] for p in engine.banzhang)

    # ── 动态标准工时 ──
    workday_count = sum(1 for d in range(1, num_days + 1) if engine.is_workday(d))
    standard_hours = workday_count * 8

    # ═══════════════════════════════════════════
    # Row 1: 表头
    # ═══════════════════════════════════════════
    row = 1
    for i, hdr in enumerate(['组长', 'moka工号', '工号', '姓名'], 1):
        _set_cell(ws, row, i, hdr, FONT_H, ALIGN_C)

    for day in range(1, num_days + 1):
        col = C_DAY + day - 1
        wd_idx = (start_weekday + day - 1) % 7
        _set_cell(ws, row, col, WEEKDAY_NAMES[wd_idx], FONT_H, ALIGN_C)

    # Row 1 统计列区留空（参考模版）

    # ═══════════════════════════════════════════
    # Row 2: 日期子头 / 统计列头
    # ═══════════════════════════════════════════
    row = 2
    ws.row_dimensions[row].height = 65

    # A-D 空但有边框（因为合并了 Row1+Row2，边框在 Row1 已设）
    for c in range(1, 5):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_H
        cell.alignment = ALIGN_C
        cell.border = THIN_BORDER

    for day in range(1, num_days + 1):
        col = C_DAY + day - 1
        dt = datetime(year, month, day)
        _set_cell(ws, row, col, dt, FONT_H, ALIGN_W,
                  number_format='m"月"d"日";@')

    # 统计列头
    stat_labels = active_stats + ['合计', '标准工时', '合计工时']
    for si, label in enumerate(stat_labels):
        align = ALIGN_C if label in ('标准工时', '合计工时') else ALIGN_W
        _set_cell(ws, row, C_STAT + si, label, FONT_H, align)

    # ═══════════════════════════════════════════
    # 列宽（在写入前设置）
    # ═══════════════════════════════════════════
    ws.column_dimensions['A'].width = 6.25
    ws.column_dimensions['B'].width = 12.63
    ws.column_dimensions['C'].width = 8.63
    ws.column_dimensions['D'].width = 7.87
    for day in range(1, num_days + 1):
        ws.column_dimensions[get_column_letter(C_DAY + day - 1)].width = 7.63
    for si, label in enumerate(stat_labels):
        cl = get_column_letter(C_STAT + si)
        ws.column_dimensions[cl].width = 7.87 if label in ('标准工时', '合计工时') else 4.63

    # ═══════════════════════════════════════════
    # 数据行
    # ═══════════════════════════════════════════
    row = 3
    stat_range_end_col = stat_col(num_stat_cols - 1)
    stat_range_start_col = stat_col(0)

    for name, emp_id, shifts_list in all_people:
        ws.row_dimensions[row].height = 14.25

        moka_id, _ = _get_moka_info(emp_id, name, mapping)
        day_range = f'{day_start_letter}{row}:{day_end_letter}{row}'

        # A: 组长 — 写班长名字
        _set_cell(ws, row, 1, leader_names, FONT_D9, ALIGN_W)

        # B: moka工号
        _set_cell(ws, row, 2, moka_id, FONT_D9, ALIGN_W)

        # C: 工号
        _set_cell(ws, row, 3,
                  int(emp_id) if emp_id.isdigit() else emp_id,
                  FONT_D9, ALIGN_C)

        # D: 姓名
        _set_cell(ws, row, 4, name, FONT_D9, ALIGN_C)

        # 每日班次
        for day in range(1, num_days + 1):
            col = C_DAY + day - 1
            _set_cell(ws, row, col,
                      shifts_list[day - 1] if day - 1 < len(shifts_list) else '',
                      FONT_D8, ALIGN_W)

        # ── 动态统计列 COUNTIF ──
        # 休列需要额外统计"请假"（已转换但 COUNTIF 需要原生"休"）
        for si, shift_name in enumerate(active_stats):
            if shift_name == '休':
                formula = f'=COUNTIF({day_range},"休")+COUNTIF({day_range},"请假")'
            else:
                formula = f'=COUNTIF({day_range},"{shift_name}")'
            _set_cell(ws, row, C_STAT + si, formula, FONT_D9, ALIGN_W)

        # 合计 =SUM(统计首列:统计末列)
        sum_formula = f'=SUM({stat_range_start_col}{row}:{stat_range_end_col}{row})'
        _set_cell(ws, row, C_STAT + num_stat_cols, sum_formula, FONT_D9, ALIGN_W)

        # 标准工时
        _set_cell(ws, row, C_STAT + num_stat_cols + 1, standard_hours, FONT_D9, ALIGN_C)

        # 合计工时 — 动态生成
        # 长白班*11.5 + 长白-早*12 + (8h班次之和)*8 + 专职大夜*10.5
        formula_parts = []
        if stat_changbai is not None:
            formula_parts.append(f'{stat_col(stat_changbai)}{row}*11.5')
        if stat_changbai_zao is not None:
            formula_parts.append(f'{stat_col(stat_changbai_zao)}{row}*12')
        if stat_eight_hour:
            eight_sum = '+'.join(f'{stat_col(i)}{row}' for i in stat_eight_hour)
            formula_parts.append(f'({eight_sum})*8')
        if stat_night is not None:
            formula_parts.append(f'{stat_col(stat_night)}{row}*10.5')

        if formula_parts:
            hours_formula = '=' + '+'.join(formula_parts)
        else:
            hours_formula = '0'
        _set_cell(ws, row, C_STAT + num_stat_cols + 2, hours_formula, FONT_D9, ALIGN_C)

        row += 1

    # ═══════════════════════════════════════════
    # 班次时间对照表
    # ═══════════════════════════════════════════
    row += 1

    # 表头
    _set_cell(ws, row, 4, '上班时间', FONT_D9, ALIGN_C)
    _set_cell(ws, row, 6, '下班时间', FONT_D9, ALIGN_C)
    ws.merge_cells(f'D{row}:E{row}')
    ws.merge_cells(f'F{row}:G{row}')
    row += 1

    time_table = [
        ('长白班', time(8, 0), time(20, 30)),
        ('长白晚', time(9, 0), time(22, 0)),
        ('长白-早', time(8, 0), time(21, 0)),
        ('早早班', time(6, 0), time(15, 0)),
        ('早早2', time(7, 0), time(16, 0)),
        ('早班', time(8, 0), time(17, 0)),
        ('天地班', time(8, 0), time(13, 0)),
        ('早三', time(8, 30), time(17, 30)),
        ('行政班', time(9, 0), time(18, 0)),
        ('白班', time(10, 0), time(19, 0)),
        ('中二', time(12, 0), time(21, 0)),
        ('中三', time(13, 0), time(22, 0)),
        ('中四', time(14, 0), time(23, 0)),
        ('晚一', time(15, 0), timedelta(0)),
        ('晚二', time(16, 0), time(1, 0)),
    ]

    for shift_label, start_t, end_t in time_table:
        _set_cell(ws, row, 2, shift_label, FONT_D9, ALIGN_C)
        _set_cell(ws, row, 4, start_t, FONT_D9, ALIGN_C, number_format='h:mm;@')
        _set_cell(ws, row, 6, end_t, FONT_D9, ALIGN_C, number_format='h:mm;@')
        ws.merge_cells(f'D{row}:E{row}')
        ws.merge_cells(f'F{row}:G{row}')
        row += 1

    # ═══════════════════════════════════════════
    # 合并单元格（全部写入完成后执行，避免 MergedCell 冲突）
    # ═══════════════════════════════════════════
    data_end_row = 2 + len(all_people)
    for col_letter in ['A', 'B', 'C', 'D']:
        ws.merge_cells(f'{col_letter}1:{col_letter}2')
    if data_end_row >= 3:
        ws.merge_cells(f'A3:A{data_end_row}')

    # ═══════════════════════════════════════════
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
