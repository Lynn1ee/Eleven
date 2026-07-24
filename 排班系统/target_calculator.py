"""从原始进线数据计算每日目标人数，支持导出系数模板 Excel"""
import io
import math
from datetime import datetime, date, timedelta
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
import urllib.request
import json
import calendar as cal_mod

# 节假日 API（复用 schedule_engine 的逻辑）
HOLIDAY_API = "https://timor.tech/api/holiday/year"
_holiday_cache = {}


def _fetch_holidays(year):
    """获取国务院节假日数据，内存缓存"""
    if year in _holiday_cache:
        return _holiday_cache[year]
    try:
        req = urllib.request.Request(
            f"{HOLIDAY_API}/{year}",
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        )
        resp = urllib.request.urlopen(req, timeout=10)
        data = json.loads(resp.read())
        if data.get("code") == 0 and data.get("holiday"):
            parsed = {}
            for date_str, info in data["holiday"].items():
                parts = date_str.split("-")
                m = int(parts[0])
                d = int(parts[1])
                if m not in parsed:
                    parsed[m] = {}
                parsed[m][d] = {
                    "holiday": info.get("holiday", False),
                    "name": info.get("name", ""),
                    "wage": info.get("wage", 1),
                }
            _holiday_cache[year] = parsed
            return parsed
    except Exception:
        pass
    _holiday_cache[year] = {}
    return _holiday_cache[year]


def _count_workdays(year, month):
    """计算指定月份的工作日数（排除周末和节假日，加回调休补班）"""
    num_days = cal_mod.monthrange(year, month)[1]
    holidays = _fetch_holidays(year).get(month, {})
    workdays = 0
    for d in range(1, num_days + 1):
        wd = date(year, month, d).weekday()
        is_weekend = wd >= 5
        h = holidays.get(d, {})
        if h.get("holiday"):
            continue  # 法定假日 → 休息
        if h.get("holiday") is False and h.get("wage", 1) <= 1:
            workdays += 1  # 调休补班
        elif not is_weekend:
            workdays += 1  # 普通工作日
    return workdays


class TargetCalculator:
    """读取去年同月进线数据，计算周系数/日系数，生成每日目标人数"""

    def __init__(self, file_bytes: bytes):
        self._file_bytes = file_bytes
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

        # 查找"时间段进线情况明细表" sheet
        sheet_name = None
        for sn in wb.sheetnames:
            if '明细表' in sn or '明细' in sn:
                sheet_name = sn
                break
        if not sheet_name:
            raise ValueError(f'未找到"时间段进线情况明细表"，可用 Sheet: {wb.sheetnames}')

        ws = wb[sheet_name]
        self._daily_data = defaultdict(list)
        # 检测列位置（根据表头）
        col_datetime = 0  # 0-indexed
        col_click = 2     # 默认 C 列
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for i, h in enumerate(header):
            if h and '时间' in str(h):
                col_datetime = i
            if h and '点击人工' in str(h):
                col_click = i

        for row in ws.iter_rows(min_row=2, values_only=True):
            dt_val = row[col_datetime] if col_datetime < len(row) else None
            click_val = row[col_click] if col_click < len(row) else 0
            if dt_val is None:
                continue
            if isinstance(dt_val, str):
                try:
                    dt_val = datetime.strptime(dt_val.strip(), '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    continue
            if isinstance(dt_val, datetime):
                d = dt_val.date()
                val = float(click_val or 0)
                self._daily_data[d].append(val)

        wb.close()

        if not self._daily_data:
            raise ValueError('未能解析到任何数据，请检查文件格式')

        self._dates = sorted(self._daily_data.keys())
        self._first_date = self._dates[0]
        self._last_date = self._dates[-1]
        # 验证每天48条
        for d in self._dates:
            count = len(self._daily_data[d])
            if count != 48:
                raise ValueError(f'日期 {d} 有 {count} 条记录，预期 48 条（半小时一条）')

    # ── 基本属性 ──

    @property
    def source_year(self):
        return self._first_date.year

    @property
    def source_month(self):
        return self._dates[len(self._dates) // 2].month  # 用中间日期避免跨月问题

    @property
    def source_num_days(self):
        return len(self._dates)

    @property
    def source_start_weekday(self):
        return self._first_date.weekday()

    # ── 数据获取 ──

    def get_daily_totals(self):
        return {d: sum(vals) for d, vals in self._daily_data.items()}

    def _get_week_label(self, day_num, start_weekday=None):
        """根据 day_num (1-indexed) 和 start_weekday (0=Mon) 确定周标签，按周一~周日分组"""
        if start_weekday is None:
            start_weekday = self.source_start_weekday
        # 第一个周一是第几天
        if start_weekday == 0:
            first_monday = 1
        else:
            first_monday = 8 - start_weekday
        if day_num < first_monday:
            return 'W1'
        return f'W{(day_num - first_monday) // 7 + 2}'

    # ── 系数计算 ──

    def compute_all_coefficients(self):
        daily_totals = self.get_daily_totals()
        num_days = self.source_num_days
        monthly_total = sum(daily_totals.values())
        monthly_avg = monthly_total / num_days

        # 按周分组
        week_groups = defaultdict(list)
        for d in self._dates:
            day_num = (d - self._first_date).days + 1
            label = self._get_week_label(day_num)
            week_groups[label].append(daily_totals[d])

        weekly_coeffs = {}
        for label in sorted(week_groups.keys()):
            totals = week_groups[label]
            weekly_coeffs[label] = round((sum(totals) / len(totals)) / monthly_avg, 6)

        # 按星期几分组
        weekday_groups = defaultdict(list)
        for d in self._dates:
            wd = d.weekday()
            weekday_groups[wd].append(daily_totals[d])

        weekday_coeffs = {}
        for wd in range(7):
            if weekday_groups[wd]:
                avg = sum(weekday_groups[wd]) / len(weekday_groups[wd])
                weekday_coeffs[wd] = round(avg / monthly_avg, 6)
            else:
                weekday_coeffs[wd] = 1.0

        # 每日比例
        daily_ratios = {}
        for d in self._dates:
            day_num = (d - self._first_date).days + 1
            wd = d.weekday()
            wl = self._get_week_label(day_num)
            ratio = weekly_coeffs.get(wl, 1.0) * weekday_coeffs.get(wd, 1.0)
            daily_ratios[day_num] = ratio

        all_ratios_sum = sum(daily_ratios.values())

        return {
            'source_monthly_avg': round(monthly_avg, 2),
            'source_total': round(monthly_total, 2),
            'weekly_coeffs': weekly_coeffs,
            'weekday_coeffs': weekday_coeffs,
            'daily_ratios': daily_ratios,
            'all_ratios_sum': round(all_ratios_sum, 6),
        }

    # ── 目标计算 ──

    def compute_daily_targets(self, target_year, target_month, online_count,
                              workday_count=None, max_daily_cap=None, min_daily_floor=None):
        num_days = cal_mod.monthrange(target_year, target_month)[1]
        target_start_weekday = date(target_year, target_month, 1).weekday()

        if workday_count is None:
            workday_count = _count_workdays(target_year, target_month)

        total_person_trips = online_count * workday_count

        coeffs = self.compute_all_coefficients()
        weekly_c = coeffs['weekly_coeffs']
        weekday_c = coeffs['weekday_coeffs']

        # 为每天计算比例
        all_weekly = list(weekly_c.values())
        last_weekly = all_weekly[-1] if all_weekly else 1.0

        ratios = []
        raw_theoretical = []
        for day in range(1, num_days + 1):
            wd = (target_start_weekday + day - 1) % 7
            wl = self._get_week_label(day, target_start_weekday)
            weekly_coeff = weekly_c.get(wl, last_weekly)
            day_coeff = weekday_c.get(wd, 1.0)
            ratio = weekly_coeff * day_coeff
            ratios.append(ratio)

        sum_ratios = sum(ratios)
        for r in ratios:
            raw_theoretical.append(r / sum_ratios * total_person_trips)

        if max_daily_cap is None:
            max_daily_cap = online_count - 1
        if min_daily_floor is None:
            # 下限取在线人数的 30%，只防极端低值，不干扰正常四舍五入
            min_daily_floor = max(1, int(online_count * 0.3))

        daily_targets = self._allocate_and_round(ratios, total_person_trips, max_daily_cap, min_daily_floor)

        return {
            'daily_targets': daily_targets,
            'total_person_trips': total_person_trips,
            'workday_count': workday_count,
            'max_cap': max_daily_cap,
            'min_floor': min_daily_floor,
            'raw_theoretical': [round(t, 2) for t in raw_theoretical],
        }

    def _allocate_and_round(self, ratios, total, max_cap, min_floor):
        """四舍五入初始化，差额按离进位边界(XX.5)距离微调——离边界越近越优先调整。"""
        sum_r = sum(ratios)
        theoretical = [r / sum_r * total for r in ratios]

        def round_half_up(x):
            return int(math.floor(x + 0.5))

        rounded = [round_half_up(t) for t in theoretical]
        result = [max(min_floor, min(max_cap, v)) for v in rounded]
        adjusted = [False] * len(result)

        diff = total - sum(result)
        if diff != 0:
            for _ in range(abs(diff) * 3):
                if diff == 0:
                    break
                candidates = []
                for i in range(len(result)):
                    boundary = int(theoretical[i]) + 0.5
                    dist = abs(theoretical[i] - boundary)
                    if diff > 0 and result[i] < max_cap:
                        candidates.append((dist, i))
                    elif diff < 0 and result[i] > min_floor:
                        candidates.append((dist, i))
                if not candidates:
                    break
                candidates.sort(key=lambda x: x[0])
                _, idx = candidates[0]
                result[idx] += 1 if diff > 0 else -1
                adjusted[idx] = True
                diff = total - sum(result)

        # 保存调试信息
        self._debug_alloc = []
        for i in range(len(result)):
            boundary = int(theoretical[i]) + 0.5
            self._debug_alloc.append({
                'theoretical': theoretical[i],
                'rounded': rounded[i],
                'dist': abs(theoretical[i] - boundary),
                'adjusted': adjusted[i],
            })

        return result

    # ── 模板 Excel 导出 ──

    def _week_label_cn(self, week_num):
        """W1→第一周, W2→第二周..."""
        labels = ['', '第一周', '第二周', '第三周', '第四周', '第五周', '第六周', '第七周']
        n = int(week_num[1:])
        if n < len(labels):
            return labels[n]
        return f'第{n}周'

    def export_template_excel(self):
        daily_totals = self.get_daily_totals()
        monthly_total = sum(daily_totals.values())
        monthly_avg = monthly_total / self.source_num_days
        wd_names = ['一', '二', '三', '四', '五', '六', '日']
        full_wd_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']

        from datetime import time as dt_time

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'

        # 样式（对齐原模版）
        font_normal = Font(name='微软雅黑', size=10)
        font_bold = Font(name='微软雅黑', size=10, bold=True)
        font_header = Font(name='微软雅黑', size=10, bold=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        weekend_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # 第1行：表头
        ws.cell(1, 1, '星期').font = font_header
        ws.cell(1, 2, '日期').font = font_header
        ws.cell(1, 3, '点击量').font = font_header
        ws.cell(1, 4, '系数').font = font_header
        for i in range(48):
            h = i // 2
            m = (i % 2) * 30
            ws.cell(1, 5 + i, dt_time(h, m)).font = font_header
        for c in range(1, 53):
            ws.cell(1, c).alignment = align_center
            ws.cell(1, c).border = thin_border
            ws.cell(1, c).fill = header_fill

        # 按日期汇总半小时数据
        sorted_dates = sorted(daily_totals.keys())
        # 周分组
        week_boundaries = []
        current_week = None
        for d in sorted_dates:
            day_num = (d - self._first_date).days + 1
            wl = self._get_week_label(day_num)
            if wl != current_week:
                week_boundaries.append(d)
                current_week = wl

        # 数据行
        row_num = 2
        wd_group_vals = defaultdict(list)      # {weekday: [daily_totals]}
        wd_group_hh = {i: defaultdict(list) for i in range(7)}  # {wd: {half_hour_idx: [values]}}
        week_group_vals = defaultdict(list)    # {week_label: [daily_totals]}
        week_group_hh = {}                     # {week_label: {half_hour_idx: [values]}}
        week_idx = 0

        all_week_labels = set()
        for d in sorted_dates:
            vals = self._daily_data.get(d, [0] * 48)
            day_total = sum(vals)
            coeff = day_total / monthly_avg if monthly_avg > 0 else 0
            wd = d.weekday()
            wd_name = wd_names[wd]
            excel_date = (d - datetime(1899, 12, 30).date()).days

            # 周间插入空行
            if week_idx < len(week_boundaries) and d == week_boundaries[week_idx]:
                if row_num > 2:
                    row_num += 1
                week_idx += 1

            ws.cell(row_num, 1, wd_name).font = font_normal
            ws.cell(row_num, 2, excel_date).font = font_normal
            ws.cell(row_num, 2).number_format = 'm/d'
            ws.cell(row_num, 3, day_total).font = font_normal
            ws.cell(row_num, 4, round(coeff, 4)).font = font_normal

            for i, v in enumerate(vals):
                ws.cell(row_num, 5 + i, v).font = font_normal

            for c in range(1, 53):
                ws.cell(row_num, c).alignment = align_center
                ws.cell(row_num, c).border = thin_border
                if wd >= 5:
                    ws.cell(row_num, c).fill = weekend_fill

            # 累积汇总数据
            wd_group_vals[wd].append(day_total)
            for i, v in enumerate(vals):
                wd_group_hh[wd][i].append(v)

            day_num = (d - self._first_date).days + 1
            wl = self._get_week_label(day_num)
            all_week_labels.add(wl)
            week_group_vals[wl].append(day_total)
            if wl not in week_group_hh:
                week_group_hh[wl] = {i: [] for i in range(48)}
            for i, v in enumerate(vals):
                week_group_hh[wl][i].append(v)

            row_num += 1

        # 空行分隔
        row_num += 1

        # 星期平均（周一~日）
        wd_coeffs_export = {}
        for wd in range(7):
            vals = wd_group_vals.get(wd, [])
            if not vals:
                continue
            avg = sum(vals) / len(vals)
            c = round(avg / monthly_avg, 4) if monthly_avg > 0 else 0
            wd_coeffs_export[wd] = c
            ws.cell(row_num, 1, full_wd_names[wd]).font = font_bold
            ws.cell(row_num, 2, '周平均值').font = font_bold
            ws.cell(row_num, 3, round(avg, 2)).font = font_bold
            ws.cell(row_num, 4, c).font = font_bold
            for i in range(48):
                hh_vals = wd_group_hh[wd].get(i, [0])
                if hh_vals:
                    ws.cell(row_num, 5 + i, round(sum(hh_vals) / len(hh_vals), 2)).font = font_bold
            for c2 in range(1, 53):
                ws.cell(row_num, c2).alignment = align_center
                ws.cell(row_num, c2).border = thin_border
            row_num += 1

        row_num += 1

        # 周平均（第一周~第N周）
        week_coeffs_export = {}
        sorted_weeks = sorted(all_week_labels, key=lambda x: int(x[1:]))
        for wl in sorted_weeks:
            vals = week_group_vals[wl]
            if not vals:
                continue
            avg = sum(vals) / len(vals)
            c = round(avg / monthly_avg, 4) if monthly_avg > 0 else 0
            week_coeffs_export[wl] = c
            ws.cell(row_num, 1, self._week_label_cn(wl)).font = font_bold
            ws.cell(row_num, 2, '周平均值').font = font_bold
            ws.cell(row_num, 3, round(avg, 2)).font = font_bold
            ws.cell(row_num, 4, c).font = font_bold
            for i in range(48):
                hh_vals = week_group_hh[wl][i]
                if hh_vals:
                    ws.cell(row_num, 5 + i, round(sum(hh_vals) / len(hh_vals), 2)).font = font_bold
            for c2 in range(1, 53):
                ws.cell(row_num, c2).alignment = align_center
                ws.cell(row_num, c2).border = thin_border
            row_num += 1

        row_num += 1

        # 月平均（对齐原模版格式）
        last_date_serial = (sorted_dates[-1] - datetime(1899, 12, 30).date()).days
        ws.cell(row_num, 1, last_date_serial).font = font_bold
        ws.cell(row_num, 2, '月平均值').font = font_bold
        ws.cell(row_num, 3, round(monthly_avg, 2)).font = font_bold
        # D列留空（和原模版一致）
        for i in range(48):
            all_hh = []
            for d2 in sorted_dates:
                hh_vals = self._daily_data.get(d2, [0] * 48)
                if i < len(hh_vals):
                    all_hh.append(hh_vals[i])
            if all_hh:
                ws.cell(row_num, 5 + i, round(sum(all_hh) / len(all_hh), 2)).font = font_bold
        for c2 in range(1, 53):
            ws.cell(row_num, c2).alignment = align_center
            ws.cell(row_num, c2).border = thin_border
        # D列留空
        ws.cell(row_num, 4).value = None

        # 列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        for i in range(48):
            ws.column_dimensions[get_column_letter(5 + i)].width = 5

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
