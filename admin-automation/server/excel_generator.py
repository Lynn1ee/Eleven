"""Excel 生成：月度任务量统计表、月考成绩表"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 字体
FONT_DEFAULT = Font(name="微软雅黑", size=10)
FONT_BOLD = Font(name="微软雅黑", size=10, bold=True)
# 对齐
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
# 边框
THIN_SIDE = Side(style="thin", color="000000")
BORDER_THIN = Border(top=THIN_SIDE, bottom=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE)
# 标题行填充色
TITLE_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
# 数字格式（两位小数）
NUM_FORMAT_2D = '0.00'

COL_WIDTHS = [9.0, 10.625, 10.875, 14.375, 17.125, 17.25, 13.0, 13.0, 13.0]

HEADERS_FENXIAO = ["工号", "姓名", "在线会话量", "机票分销处理量", "火车票分销处理量",
                    "IM会话量\n(拼团+千牛+微信)", "外部支援", "合计", "当月分值\n(四舍五入保留2位小数)"]
HEADERS_DAYE    = ["工号", "姓名", "在线会话量", "机票分销处理量", "火车票分销处理量",
                    "IM会话量\n(拼团+千牛+微信)", "外部支援", "合计"]
HEADERS_KEFU    = ["工号", "姓名", "在线会话量", "机票分销处理量", "火车票分销处理量",
                    "IM会话量\n(拼团+千牛+微信)", "带教/支援会话量", "合计", "当月分值\n(四舍五入保留2位小数)"]


def generate_excel(groups, month_str):
    """按照模板格式生成 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "在线合计任务量"

    # 列宽
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── 专职分销 ──
    row = write_group(ws, row, "专职分销", HEADERS_FENXIAO, groups.get("fenxiao", []), calc_score=True)
    row = write_separator(ws, row)

    # ── 专职大夜 ──
    row = write_group(ws, row, "专职大夜", HEADERS_DAYE, groups.get("daye", []), calc_score=False)
    row = write_separator(ws, row)

    # ── 在线客服 ──
    row = write_group(ws, row, "在线客服", HEADERS_KEFU, groups.get("kefu", []), calc_score=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def write_group(ws, start_row, title, headers, rows, calc_score):
    r = start_row

    # 标题行（合并 A-I）
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    cell = ws.cell(row=r, column=1, value=title)
    cell.font = FONT_BOLD
    cell.fill = TITLE_FILL
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER_THIN
    for c in range(2, 10):
        ws.cell(row=r, column=c).border = BORDER_THIN
    ws.row_dimensions[r].height = 24
    r += 1

    # 表头行
    ws.row_dimensions[r].height = 33
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.font = FONT_DEFAULT
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
    # 如果实际表头少于 9 列（专职大夜），补齐空列边框
    for c in range(len(headers) + 1, 10):
        ws.cell(row=r, column=c).border = BORDER_THIN
    r += 1

    # 找最大合计
    max_total = 0
    if calc_score and rows:
        for row_data in rows:
            total = sum(row_data.get(k, 0) for k in ["online", "flight", "train", "im", "support"])
            if total > max_total:
                max_total = total

    for idx, row_data in enumerate(rows):
        is_top = (idx == 0 and calc_score)

        ws.cell(row=r, column=1, value=row_data.get("id", "")).font = FONT_DEFAULT
        ws.cell(row=r, column=2, value=row_data.get("name", "")).font = FONT_DEFAULT
        ws.cell(row=r, column=3, value=row_data.get("online", 0)).font = FONT_DEFAULT
        ws.cell(row=r, column=4, value=row_data.get("flight", 0)).font = FONT_DEFAULT
        ws.cell(row=r, column=5, value=row_data.get("train", 0)).font = FONT_DEFAULT
        ws.cell(row=r, column=6, value=row_data.get("im", 0)).font = FONT_DEFAULT
        ws.cell(row=r, column=7, value=row_data.get("support", 0)).font = FONT_DEFAULT
        # 合计 = 直接写入计算值而非公式（确保邮件预览可见）
        total_val = sum(row_data.get(k, 0) for k in ["online", "flight", "train", "im", "support"])
        ws.cell(row=r, column=8, value=total_val).font = FONT_DEFAULT
        # 当月分值
        if calc_score:
            if is_top:
                c = ws.cell(row=r, column=9, value=20)
                c.font = FONT_DEFAULT
                c.number_format = NUM_FORMAT_2D
            else:
                score = round(total_val / max_total * 20, 2) if max_total > 0 else 0
                c = ws.cell(row=r, column=9, value=score)
                c.font = FONT_DEFAULT
                c.number_format = NUM_FORMAT_2D

        for c in range(1, 10):
            ws.cell(row=r, column=c).alignment = ALIGN_CENTER
            ws.cell(row=r, column=c).border = BORDER_THIN

        r += 1

    return r


def write_separator(ws, row):
    """写入空行作为分组间隔"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    for c in range(1, 10):
        ws.cell(row=row, column=c).border = BORDER_THIN
    ws.row_dimensions[row].height = 33
    return row + 1


def generate_exam_excel(records, month_str):
    """生成月考成绩 Excel，格式：月份 | 姓名 | 工号 | 成绩"""
    wb = Workbook()
    ws = wb.active
    ws.title = "月考成绩"

    # 列宽
    ws.column_dimensions['A'].width = 9.0
    ws.column_dimensions['B'].width = 12.375
    ws.column_dimensions['C'].width = 11.875
    ws.column_dimensions['D'].width = 9.0

    # 标题行
    headers = ["月份", "姓名", "工号", "成绩"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = FONT_DEFAULT
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN

    # 数据行
    for r, rec in enumerate(records, 2):
        ws.cell(row=r, column=1, value=month_str).font = FONT_DEFAULT
        ws.cell(row=r, column=2, value=rec.get("name", "")).font = FONT_DEFAULT
        ws.cell(row=r, column=3, value=rec.get("id", "")).font = FONT_DEFAULT
        ws.cell(row=r, column=4, value=rec.get("score", 0)).font = FONT_DEFAULT
        for c in range(1, 5):
            ws.cell(row=r, column=c).alignment = ALIGN_CENTER
            ws.cell(row=r, column=c).border = BORDER_THIN

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
