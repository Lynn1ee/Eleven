"""客服升降岗预警 — API 处理函数"""
import io
import json
import math
import openpyxl
from ..db import get_db


# ── 常量 ──
FIXED_EXCLUDED = {"何丹妮", "曾瑾", "李玲"}  # 固定排除名单

# 历史 Excel 单元格颜色 → 档位映射
# 显式 RGB 色值集合
_VIP_RGB = {"FFFF0000"}
_SENIOR_RGB = {"FFFFFF00"}  # 黄色 = 资深客服
_SUPPORT_RGB = {"FFFFC000", "FFED7D31", "FFF4B084", "FFF4A460"}  # 橙色/金色 = 支援
_ADVANCED_RGB = {"FF92D050", "FF00B050", "FF00B04F", "FF548235",
                 "FFA9D18E", "FF70AD47", "FF375623", "FF9DC35A",
                 "FFB6D7A8", "FFC5E0B4", "FFE2EFDA", "FFA8D08D"}
_NIGHT_RGB = {"FF00B0F0", "FF0070C0", "FF4472C4", "FF8DB4E2",
              "FF4F81BD", "FF2E75B6", "FF5B9BD5", "FF9DC3E6",
              "FFBDD7EE", "FFD6E4F0", "FF0066CC", "FF3399FF"}


def _detect_cell_tier(cell):
    """根据单元格背景色返回历史档位，无颜色返回 None"""
    try:
        fill = cell.fill
        if not fill or not fill.patternType:
            return None
        fg = fill.fgColor
        if not fg:
            return None

        rgb = None
        theme = None
        try:
            if fg.rgb and isinstance(fg.rgb, str):
                rgb = str(fg.rgb)
        except Exception:
            pass
        try:
            if fg.theme is not None:
                theme = int(fg.theme)
        except (ValueError, TypeError):
            pass

        if rgb:
            if rgb in _VIP_RGB:
                return "VIP客服"
            if rgb in _SENIOR_RGB:
                return "资深客服"
            if rgb in _ADVANCED_RGB:
                return "高级客服"
            if rgb in _NIGHT_RGB:
                return "大夜"
            if rgb in _SUPPORT_RGB:
                return "支援"
        if theme is not None:
            if theme == 4:  # Accent 1 = 蓝色 = 大夜
                return "大夜"
            if theme == 6:  # Accent 6 = 绿色 = 高级客服
                return "高级客服"
    except Exception:
        pass
    return None


# ── 导入 ──

def handle_import(handler, data):
    """POST /api/performance/import
    body: {year: int, month: int, entries: [{name, score, is_night_shift}]}
    """
    year = data.get("year")
    month = data.get("month")
    entries = data.get("entries", [])
    if not year or not month or not entries:
        return 400, {"success": False, "error": "缺少 year/month/entries"}

    db = get_db()
    for e in entries:
        name = e.get("name", "").strip()
        score = float(e.get("score", 0))
        is_night = 1 if e.get("is_night_shift") else 0
        if not name:
            continue
        db.execute("""
            INSERT INTO performance_scores (name, year, month, score, is_night_shift)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(name, year, month) DO UPDATE SET
                score = excluded.score,
                is_night_shift = excluded.is_night_shift
        """, (name, year, month, score, is_night))

    db.commit()
    return 200, {"success": True, "message": f"已导入 {len(entries)} 条数据"}


# ── 文件导入（含橘色大夜检测）──

def handle_import_file(handler, data):
    """POST /api/performance/import-file (multipart)
    用 openpyxl 解析 Excel，检测橘色背景行作为大夜标记
    """
    year = int(data.get("year", 0))
    month = int(data.get("month", 0))
    file_field = data.get("file")

    if not year or not month or not file_field:
        return 400, {"success": False, "error": "缺少 year/month/file"}

    file_data = file_field.value if hasattr(file_field, 'value') else file_field
    wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
    ws = wb[wb.sheetnames[0]]

    entries = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=False), start=3):
        name_cell = row[1] if len(row) > 1 else None  # B列
        score_cell = row[42] if len(row) > 42 else None  # AQ列

        if not name_cell or not name_cell.value:
            continue
        name = str(name_cell.value).strip()
        if not name:
            continue

        score = 0
        if score_cell and score_cell.value is not None:
            try:
                score = float(score_cell.value)
            except (ValueError, TypeError):
                score = 0

        # 检测橘色背景（大夜标记）
        is_night = False
        try:
            fill = name_cell.fill
            if fill and fill.fgColor and fill.fgColor.rgb:
                rgb = str(fill.fgColor.rgb)
                # 橘色相关色值
                if rgb in ('FFFFC000', 'FFF4B084', 'FFED7D31', 'FFF4A460'):
                    is_night = True
        except Exception:
            pass

        entries.append({
            "name": name,
            "score": score,
            "is_night_shift": is_night,
        })

    wb.close()

    # 写入数据库
    db = get_db()
    for e in entries:
        db.execute("""
            INSERT INTO performance_scores (name, year, month, score, is_night_shift)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(name, year, month) DO UPDATE SET
                score = excluded.score,
                is_night_shift = excluded.is_night_shift
        """, (e["name"], year, month, e["score"], 1 if e["is_night_shift"] else 0))

    db.commit()
    night_count = sum(1 for e in entries if e["is_night_shift"])
    return 200, {
        "success": True,
        "message": f"已导入 {len(entries)} 条数据（其中大夜 {night_count} 人）",
        "count": len(entries),
        "night_count": night_count,
        "entries": entries,
    }


# ── 历史数据导入（年度排名 Excel）──

def handle_import_history(handler, data):
    """POST /api/performance/import-history (multipart)
    解析历史排名 Excel（多 sheet：2024年/2025年/2026年）
    列：工号 | 人员 | 客服等级 | 1月~12月
    """
    file_field = data.get("file")
    if not file_field:
        return 400, {"success": False, "error": "缺少 file"}

    file_data = file_field.value if hasattr(file_field, 'value') else file_field
    wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)

    total_scores = 0
    total_notes = 0
    total_ranks = 0
    db = get_db()

    for sheet_name in wb.sheetnames:
        year_str = sheet_name.replace("年", "").strip()
        try:
            year = int(year_str)
        except ValueError:
            continue

        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=2, values_only=False):
            if len(row) < 3:
                continue
            name_cell = row[1]
            rank_cell = row[2]

            if not name_cell or not name_cell.value:
                continue
            name = str(name_cell.value).strip()
            if not name:
                continue

            rank_level = str(rank_cell.value).strip() if rank_cell and rank_cell.value else "客服专员"

            db.execute("""
                INSERT INTO staff_ranks (name, rank_level) VALUES (?, ?)
                ON CONFLICT(name) DO UPDATE SET
                    rank_level = excluded.rank_level,
                    updated_at = CURRENT_TIMESTAMP
            """, (name, rank_level))
            total_ranks += 1

            for m in range(12):
                col_idx = 3 + m
                month = m + 1
                if col_idx >= len(row):
                    continue

                cell = row[col_idx]
                if not cell or cell.value is None:
                    continue

                val = cell.value
                score = None
                is_night = 0
                hist_tier = ""

                if isinstance(val, (int, float)):
                    score = float(val)
                    tier = _detect_cell_tier(cell)
                    if tier:
                        hist_tier = tier
                        is_night = 1 if tier == "大夜" else 0
                    else:
                        hist_tier = "客服专员"  # 无颜色 = 白色 = 客服专员
                elif isinstance(val, str):
                    val_stripped = val.strip()
                    if val_stripped == "大夜":
                        is_night = 1
                        score = 0
                    else:
                        # 其他文本作为状态标注（病假、产假、离职等），分数记为0
                        score = 0
                        note_text = val_stripped
                        db.execute("""
                            INSERT INTO score_status_notes (name, year, month, note) VALUES (?, ?, ?, ?)
                            ON CONFLICT(name, year, month) DO UPDATE SET note = excluded.note
                        """, (name, year, month, note_text))
                        total_notes += 1
                else:
                    continue

                if score is None:
                    continue

                db.execute("""
                    INSERT INTO performance_scores (name, year, month, score, is_night_shift, historical_tier)
                    VALUES (?, ?, ?, ?, ?, ?)
                    ON CONFLICT(name, year, month) DO UPDATE SET
                        score = excluded.score,
                        is_night_shift = MAX(is_night_shift, excluded.is_night_shift),
                        historical_tier = CASE WHEN excluded.historical_tier != '' THEN excluded.historical_tier ELSE historical_tier END
                """, (name, year, month, score, is_night, hist_tier))
                total_scores += 1

    wb.close()
    db.commit()
    return 200, {
        "success": True,
        "message": f"已导入历史数据：{total_ranks} 人次岗级信息，{total_scores} 条月度分数，{total_notes} 条状态标注",
    }


# ── 历史月份 ──

def handle_history(handler, data):
    """GET /api/performance/history"""
    db = get_db()
    rows = db.execute("""
        SELECT DISTINCT year, month FROM performance_scores
        ORDER BY year DESC, month DESC
    """).fetchall()
    result = [{"year": r["year"], "month": r["month"]} for r in rows]
    return 200, {"success": True, "history": result}


# ── 排名计算 ──

def handle_ranking(handler, data):
    """GET /api/performance/ranking?year=&month="""
    year = int(data.get("year", 0))
    month = int(data.get("month", 0))
    if not year or not month:
        return 400, {"success": False, "error": "缺少 year/month 参数"}

    db = get_db()

    # 查询当月数据
    rows = db.execute("""
        SELECT ps.name, ps.score, ps.is_night_shift, COALESCE(sr.rank_level, '客服专员') as rank_level
        FROM performance_scores ps
        LEFT JOIN staff_ranks sr ON sr.name = ps.name
        WHERE ps.year = ? AND ps.month = ?
    """, (year, month)).fetchall()

    if not rows:
        return 200, {"success": True, "ranking": [], "summary": None}

    # 区分参与排名和非参与排名的人员
    all_entries = []
    for r in rows:
        all_entries.append({
            "name": r["name"],
            "score": r["score"],
            "is_night_shift": bool(r["is_night_shift"]),
            "rank_level": r["rank_level"],
        })

    # 查询当月手动排除名单（含原因）
    excl_rows = db.execute(
        "SELECT name, COALESCE(reason, '') as reason FROM ranking_exclusions WHERE year = ? AND month = ?",
        (year, month)
    ).fetchall()
    excluded_reasons = {r["name"]: r["reason"] for r in excl_rows}
    excluded_names = set(excluded_reasons.keys())

    # 有效人数计算
    active_entries = [
        e for e in all_entries
        if e["name"] not in FIXED_EXCLUDED and e["score"] > 0
    ]
    excluded_entries = [e for e in active_entries if e["name"] in excluded_names]
    ranking_entries = [e for e in active_entries if e["name"] not in excluded_names]

    # 按分数降序排列
    ranking_entries.sort(key=lambda x: x["score"], reverse=True)
    effective_count = len(active_entries)

    # 名额计算
    vip_quota = math.ceil(effective_count * 0.05)
    senior_quota = math.ceil(effective_count * 0.30) - vip_quota
    advanced_quota = math.ceil(effective_count * 0.60) - vip_quota - senior_quota
    # junior = effective_count - vip_quota - senior_quota - advanced_quota

    def get_tier(idx):
        if idx < vip_quota:
            return "VIP客服"
        elif idx < vip_quota + senior_quota:
            return "资深客服"
        elif idx < vip_quota + senior_quota + advanced_quota:
            return "高级客服"
        else:
            return "客服专员"

    # 分配档位
    for i, e in enumerate(ranking_entries):
        e["rank"] = i + 1
        e["month_tier"] = get_tier(i)

    # 排除人员排在最后，档位显示原因
    for e in excluded_entries:
        e["rank"] = None
        e["month_tier"] = excluded_reasons.get(e["name"]) or "不参与排名"

    # 历史排名数据（用于升降岗判断）
    history = _load_ranking_history(db, year, month)

    # 计算升降岗
    for e in ranking_entries:
        change = _check_promotion_demotion(db, e["name"], e["rank_level"], e["month_tier"], year, month, history)
        e["change_type"] = change["type"]  # "up" | "down" | "keep"
        e["change_target"] = change.get("target", "")
        e["change_note"] = change.get("note", "")
        e["qualified"] = change.get("qualified", False)
        e["critical"] = change.get("critical", False)

    for e in excluded_entries:
        change = _check_promotion_demotion(db, e["name"], e["rank_level"], None, year, month, history)
        e["change_type"] = change["type"]
        e["change_target"] = change.get("target", "")
        e["change_note"] = change.get("note", "")
        e["qualified"] = change.get("qualified", False)
        e["critical"] = change.get("critical", False)

    # 组装返回
    ranked = ranking_entries + excluded_entries

    summary = {
        "effective_count": effective_count,
        "excluded_count": len(excluded_entries),
        "zero_excluded": sum(1 for e in all_entries if e["score"] == 0 and e["name"] not in FIXED_EXCLUDED),
        "vip_quota": vip_quota,
        "senior_quota": senior_quota,
        "advanced_quota": advanced_quota,
        "junior_quota": effective_count - vip_quota - senior_quota - advanced_quota,
    }
    return 200, {"success": True, "ranking": ranked, "summary": summary}


def _load_ranking_history(db, current_year, current_month):
    """加载所有历史月份的排名数据，返回 {name: {ym: tier}}"""
    rows = db.execute("""
        SELECT ps.name, ps.year, ps.month, ps.is_night_shift, ps.score
        FROM performance_scores ps
        ORDER BY ps.year, ps.month
    """).fetchall()

    # 按年月分组
    by_month = {}
    for r in rows:
        ym = f"{r['year']}-{r['month']:02d}"
        if ym not in by_month:
            by_month[ym] = []
        by_month[ym].append({
            "name": r["name"],
            "score": r["score"],
            "is_night_shift": bool(r["is_night_shift"]),
        })

    # 为每个月计算排名
    history = {}
    for ym, entries in by_month.items():
        active = [e for e in entries if e["name"] not in FIXED_EXCLUDED and e["score"] > 0]
        ranking = [e for e in active if not e["is_night_shift"]]
        ranking.sort(key=lambda x: x["score"], reverse=True)
        eff = len(active)
        vip_q = math.ceil(eff * 0.05)
        senior_q = math.ceil(eff * 0.30) - vip_q
        adv_q = math.ceil(eff * 0.60) - vip_q - senior_q

        for e in entries:
            if e["name"] not in history:
                history[e["name"]] = {}
            if e["name"] in FIXED_EXCLUDED or e["score"] == 0:
                history[e["name"]][ym] = None  # 不参与
            elif e["is_night_shift"]:
                history[e["name"]][ym] = "大夜"
            else:
                # 找排名
                idx = next((i for i, re in enumerate(ranking) if re["name"] == e["name"]), -1)
                if idx < vip_q:
                    history[e["name"]][ym] = "VIP客服"
                elif idx < vip_q + senior_q:
                    history[e["name"]][ym] = "资深客服"
                elif idx < vip_q + senior_q + adv_q:
                    history[e["name"]][ym] = "高级客服"
                else:
                    history[e["name"]][ym] = "客服专员"

    return history


def _fmt_month_range(ym_list):
    """将月份列表格式化为可读字符串，有间隔时逐月罗列"""
    if not ym_list:
        return ""
    parts = [(int(ym.split("-")[0]), int(ym.split("-")[1])) for ym in ym_list]
    if len(parts) == 1:
        return f"{parts[0][0]}年{parts[0][1]}月"
    # 检测是否有间隔（如 5月→7月 跳过了6月）
    continuous = True
    for i in range(1, len(parts)):
        y1, m1 = parts[i - 1]
        y2, m2 = parts[i]
        expected_m = m1 + 1 if m1 < 12 else 1
        expected_y = y1 if m1 < 12 else y1 + 1
        if not (y2 == expected_y and m2 == expected_m):
            continuous = False
            break
    if continuous:
        y1, m1 = parts[0]
        y2, m2 = parts[-1]
        if y1 == y2:
            return f"{y1}年{m1}-{m2}月"
        return f"{y1}年{m1}月-{y2}年{m2}月"
    # 有间隔，逐月罗列
    return "、".join(f"{y}年{m}月" for y, m in parts)


def _check_promotion_demotion(db, name, current_rank, current_tier, year, month, history):
    """判断升降岗，返回 {type, target, note}"""
    rank_order = {"客服专员": 0, "高级客服": 1, "资深客服": 2, "VIP客服": 3}
    rank_names = {0: "客服专员", 1: "高级客服", 2: "资深客服", 3: "VIP客服"}
    current_level = rank_order.get(current_rank, 0)

    # 大夜不参与升降岗判定
    if current_tier == "大夜":
        return {"type": "keep", "target": "", "note": ""}

    # 获取该员工的历史月度档位（按时间排序，跳过大夜和无效月份）
    name_history = history.get(name, {})
    sorted_months = sorted(name_history.keys())
    tier_seq = []  # [(ym, tier)]
    for ym in sorted_months:
        tier = name_history[ym]
        if tier == "大夜" or tier is None:
            continue
        tier_seq.append((ym, tier))

    if not tier_seq:
        return {"type": "keep", "target": "", "note": ""}

    # ── 升岗判断 ──
    promo_rules = [
        # (from_level, required_tier, required_months)
        (0, "高级客服", 3),
        (0, "资深客服", 3),  # 跳级
        (0, "VIP客服", 3),   # 跳级
        (1, "资深客服", 3),
        (1, "VIP客服", 3),   # 跳级
        (2, "VIP客服", 6),
    ]

    promo_results = []
    for from_lv, req_tier, req_months in promo_rules:
        if from_lv != current_level:
            continue
        if current_level >= rank_order.get(req_tier, 0):
            continue

        consecutive = 0
        months_list = []
        for ym, tier in reversed(tier_seq):
            tier_level = rank_order.get(tier, -1)
            req_level = rank_order.get(req_tier, 99)
            if tier_level >= req_level:
                consecutive += 1
                months_list.insert(0, ym)
            else:
                break

        if consecutive > 0:
            promo_results.append({
                "target": req_tier,
                "months": months_list,
                "consecutive": consecutive,
                "required": req_months,
            })

    if promo_results:
        best = max(promo_results, key=lambda x: rank_order.get(x["target"], 0))
        target = best["target"]
        cons = best["consecutive"]
        req = best["required"]
        mons = best["months"]

        if cons >= req:
            note = f"{_fmt_month_range(mons[-req:])}连续{req}月在{target}排名，本月可升岗至{target}"
            return {"type": "up", "target": target, "note": note, "qualified": True, "critical": False}
        else:
            remaining = req - cons
            note = f"已连续{cons}月在{target}排名，再坚持{remaining}月可升岗至{target}"
            if remaining == 1:
                note = "⚠️ " + note
            return {"type": "up", "target": target, "note": note, "qualified": False, "critical": remaining == 1}

    # ── 降岗判断 ──
    if current_level > 0:
        consecutive_down = 0
        down_months = []
        for ym, tier in reversed(tier_seq):
            tier_level = rank_order.get(tier, -1)
            if tier_level < current_level:
                consecutive_down += 1
                down_months.insert(0, ym)
            else:
                break

        DOWN_REQUIRED = 3
        if consecutive_down >= 1:
            new_level = current_level - 1
            target = rank_names[new_level]
            if consecutive_down >= DOWN_REQUIRED:
                note = f"{_fmt_month_range(down_months[-DOWN_REQUIRED:])}连续{DOWN_REQUIRED}月不在{current_rank}排名，本月将降岗至{target}"
                return {"type": "down", "target": target, "note": note, "qualified": True, "critical": False}
            else:
                remaining = DOWN_REQUIRED - consecutive_down
                note = f"已连续{consecutive_down}月不在{current_rank}排名，再持续{remaining}月将降岗至{target}"
                if remaining == 1:
                    note = "⚠️ " + note
                return {"type": "down", "target": target, "note": note, "qualified": False, "critical": remaining == 1}

    return {"type": "keep", "target": "", "note": ""}


# ── 确认岗级变更 ──

def handle_confirm_rank(handler, data):
    """POST /api/performance/confirm-rank
    body: {changes: [{name, new_rank}]}
    """
    changes = data.get("changes", [])
    if not changes:
        return 400, {"success": False, "error": "缺少 changes"}

    db = get_db()
    for c in changes:
        name = c.get("name", "").strip()
        new_rank = c.get("new_rank", "").strip()
        if not name or not new_rank:
            continue
        db.execute("""
            INSERT INTO staff_ranks (name, rank_level) VALUES (?, ?)
            ON CONFLICT(name) DO UPDATE SET rank_level = excluded.rank_level, updated_at = CURRENT_TIMESTAMP
        """, (name, new_rank))
    db.commit()
    return 200, {"success": True, "message": f"已确认 {len(changes)} 条岗级变更"}


# ── 员工岗级管理 ──

def handle_get_staff_ranks(handler, data):
    """GET /api/staff/ranks"""
    db = get_db()
    rows = db.execute("SELECT * FROM staff_ranks ORDER BY name").fetchall()
    result = [
        {
            "name": r["name"],
            "rank_level": r["rank_level"],
            "is_excluded": bool(r["is_excluded"]),
            "updated_at": r["updated_at"],
        }
        for r in rows
    ]
    return 200, {"success": True, "ranks": result}


def handle_update_staff_ranks(handler, data):
    """POST /api/staff/ranks
    body: {ranks: [{name, rank_level, is_excluded}]}
    """
    ranks = data.get("ranks", [])
    if not ranks:
        return 400, {"success": False, "error": "缺少 ranks"}

    db = get_db()
    for r in ranks:
        name = r.get("name", "").strip()
        rank_level = r.get("rank_level", "客服专员")
        is_excluded = 1 if r.get("is_excluded") else 0
        if not name:
            continue
        db.execute("""
            INSERT INTO staff_ranks (name, rank_level, is_excluded) VALUES (?, ?, ?)
            ON CONFLICT(name) DO UPDATE SET
                rank_level = excluded.rank_level,
                is_excluded = excluded.is_excluded,
                updated_at = CURRENT_TIMESTAMP
        """, (name, rank_level, is_excluded))
    db.commit()
    return 200, {"success": True, "message": f"已更新 {len(ranks)} 条员工岗级"}


# ── 年度历史排名 ──

def handle_yearly_history(handler, data):
    """GET /api/performance/yearly?year=&sort_month="""
    year = int(data.get("year", 0))
    sort_month = int(data.get("sort_month", 0))  # 0=最新有数据的月份
    if not year:
        return 400, {"success": False, "error": "缺少 year 参数"}

    db = get_db()

    # 获取该年度所有绩效数据
    score_rows = db.execute("""
        SELECT name, month, score, is_night_shift, COALESCE(historical_tier, '') as historical_tier
        FROM performance_scores
        WHERE year = ?
        ORDER BY month
    """, (year,)).fetchall()

    # 只收集该年度实际有数据的人，并排除固定名单
    all_names = set()
    for r in score_rows:
        if r["name"] not in FIXED_EXCLUDED:
            all_names.add(r["name"])

    # 获取这些人的岗级信息
    staff_map = {}
    if all_names:
        placeholders = ",".join("?" for _ in all_names)
        staff_rows = db.execute(
            f"SELECT name, rank_level FROM staff_ranks WHERE name IN ({placeholders})",
            tuple(all_names)
        ).fetchall()
        staff_map = {r["name"]: r["rank_level"] for r in staff_rows}

    # 获取该年度所有排除记录
    excl_rows = db.execute(
        "SELECT name, month, COALESCE(reason, '') as reason FROM ranking_exclusions WHERE year = ?",
        (year,)
    ).fetchall()
    excl_by_month = {}  # {month: {name: reason}}
    for r in excl_rows:
        excl_by_month.setdefault(r["month"], {})[r["name"]] = r["reason"] or "不参与排名"

    # 计算每个月的排名档位
    monthly_tiers = {}  # {month: {name: tier}}
    for m in range(1, 13):
        month_scores = [r for r in score_rows if r["month"] == m]
        if not month_scores:
            continue
        excl_in_month = excl_by_month.get(m, {})
        ranking = [r for r in month_scores if r["name"] not in FIXED_EXCLUDED and r["score"] > 0
                   and not r["is_night_shift"] and r["name"] not in excl_in_month and not r["historical_tier"]]
        eff = len(ranking)
        ranking_sorted = sorted(ranking, key=lambda x: x["score"], reverse=True)
        vip_q = math.ceil(eff * 0.05)
        senior_q = math.ceil(eff * 0.30) - vip_q
        adv_q = math.ceil(eff * 0.60) - vip_q - senior_q

        monthly_tiers[m] = {}
        for r in month_scores:
            if r["name"] in FIXED_EXCLUDED:
                monthly_tiers[m][r["name"]] = None
            elif r["name"] in excl_in_month:
                monthly_tiers[m][r["name"]] = excl_in_month[r["name"]]
            elif r["score"] == 0:
                monthly_tiers[m][r["name"]] = None
            elif r["is_night_shift"]:
                monthly_tiers[m][r["name"]] = "大夜"
            elif r["historical_tier"]:
                monthly_tiers[m][r["name"]] = r["historical_tier"]
            else:
                idx = next((i for i, rs in enumerate(ranking_sorted) if rs["name"] == r["name"]), -1)
                if idx < vip_q:
                    monthly_tiers[m][r["name"]] = "VIP客服"
                elif idx < vip_q + senior_q:
                    monthly_tiers[m][r["name"]] = "资深客服"
                elif idx < vip_q + senior_q + adv_q:
                    monthly_tiers[m][r["name"]] = "高级客服"
                else:
                    monthly_tiers[m][r["name"]] = "客服专员"

    # 确定排序月份
    available_months = sorted(monthly_tiers.keys(), reverse=True)
    if sort_month > 0 and sort_month in monthly_tiers:
        sort_m = sort_month
    elif available_months:
        sort_m = available_months[0]  # 最新有数据的月份
    else:
        sort_m = 0

    # 查询该年度的状态标注
    status_rows = db.execute("""
        SELECT name, month, note FROM score_status_notes WHERE year = ?
    """, (year,)).fetchall()
    status_map = {}  # {(name, month): note}
    for r in status_rows:
        status_map[(r["name"], r["month"])] = r["note"]

    # 按员工汇总
    employees = []
    for name in all_names:
        row_data = {
            "name": name,
            "rank_level": staff_map.get(name, "客服专员"),
            "months": {},
        }
        for m in range(1, 13):
            if m in monthly_tiers and name in monthly_tiers[m]:
                tier = monthly_tiers[m][name]
                score = next((r["score"] for r in score_rows if r["month"] == m and r["name"] == name), None)
                is_excluded = name in FIXED_EXCLUDED
                note = status_map.get((name, m), "")
                row_data["months"][m] = {
                    "tier": tier,
                    "score": score,
                    "is_excluded": is_excluded,
                    "note": note,
                }
            else:
                row_data["months"][m] = None
        employees.append(row_data)

    # 按指定月份分数降序排列
    def sort_key(emp):
        if sort_m == 0:
            return (0, emp["name"])
        cell = emp["months"].get(sort_m)
        if cell and cell["score"] is not None:
            return (-cell["score"], emp["name"])
        # 该月无数据的排到最后
        return (1, emp["name"])

    employees.sort(key=sort_key)

    return 200, {
        "success": True,
        "year": year,
        "employees": employees,
        "sort_month": sort_m,
        "available_months": available_months,
    }


# ── 通知生成 ──

def handle_notice(handler, data):
    """GET /api/performance/notice?year=&month="""
    year = int(data.get("year", 0))
    month = int(data.get("month", 0))
    if not year or not month:
        return 400, {"success": False, "error": "缺少 year/month 参数"}

    # 调用排名接口获取数据
    _, ranking_result = handle_ranking(handler, data)
    if not _ == 200:
        return _, ranking_result

    ranking_data = ranking_result.get("ranking", [])
    summary = ranking_result.get("summary", {})

    # 收集升级和降级人员
    up_list = []
    down_list = []
    for e in ranking_data:
        if e["change_type"] == "up":
            up_list.append({
                "name": e["name"],
                "current_rank": e["rank_level"],
                "target_rank": e["change_target"],
                "note": e["change_note"],
            })
        elif e["change_type"] == "down":
            down_list.append({
                "name": e["name"],
                "current_rank": e["rank_level"],
                "target_rank": e["change_target"],
                "note": e["change_note"],
            })

    return 200, {
        "success": True,
        "year": year,
        "month": month,
        "summary": summary,
        "up_list": up_list,
        "down_list": down_list,
    }


# ── 状态标注（0分原因）──

def handle_get_status_reasons(handler, data):
    """GET /api/performance/status-reasons — 获取所有已使用过的标注理由"""
    db = get_db()
    rows = db.execute("""
        SELECT DISTINCT note FROM score_status_notes
        WHERE note != '' ORDER BY note
    """).fetchall()
    reasons = [r["note"] for r in rows]
    return 200, {"success": True, "reasons": reasons}


def handle_set_status_note(handler, data):
    """POST /api/performance/status-note
    body: {name, year, month, note}
    """
    name = data.get("name", "").strip()
    year = data.get("year", 0)
    month = data.get("month", 0)
    note = data.get("note", "").strip()

    if not name or not year or not month:
        return 400, {"success": False, "error": "缺少 name/year/month"}

    db = get_db()
    if note:
        db.execute("""
            INSERT INTO score_status_notes (name, year, month, note) VALUES (?, ?, ?, ?)
            ON CONFLICT(name, year, month) DO UPDATE SET note = excluded.note
        """, (name, year, month, note))
    else:
        db.execute("""
            DELETE FROM score_status_notes WHERE name = ? AND year = ? AND month = ?
        """, (name, year, month))
    db.commit()
    return 200, {"success": True, "message": "已保存"}


# ── 每月排除人员管理 ──

def handle_get_exclusions(handler, data):
    """GET /api/performance/exclusions?year=&month="""
    year = int(data.get("year", 0))
    month = int(data.get("month", 0))
    if not year or not month:
        return 400, {"success": False, "error": "缺少 year/month"}

    db = get_db()
    rows = db.execute(
        "SELECT name, COALESCE(reason, '') as reason FROM ranking_exclusions WHERE year = ? AND month = ? ORDER BY name",
        (year, month)
    ).fetchall()
    excluded = [{"name": r["name"], "reason": r["reason"]} for r in rows]

    return 200, {"success": True, "excluded": excluded}


def handle_add_exclusion(handler, data):
    """POST /api/performance/exclusions/add
    body: {name, year, month, reason}
    """
    name = data.get("name", "").strip()
    year = data.get("year", 0)
    month = data.get("month", 0)
    reason = data.get("reason", "").strip()
    if not name or not year or not month:
        return 400, {"success": False, "error": "缺少 name/year/month"}

    db = get_db()
    db.execute("""
        INSERT INTO ranking_exclusions (name, year, month, reason) VALUES (?, ?, ?, ?)
        ON CONFLICT(name, year, month) DO UPDATE SET reason = excluded.reason
    """, (name, year, month, reason))
    db.commit()

    rows = db.execute(
        "SELECT name, COALESCE(reason, '') as reason FROM ranking_exclusions WHERE year = ? AND month = ? ORDER BY name",
        (year, month)
    ).fetchall()
    excluded = [{"name": r["name"], "reason": r["reason"]} for r in rows]
    return 200, {"success": True, "excluded": excluded}


def handle_remove_exclusion(handler, data):
    """POST /api/performance/exclusions/remove
    body: {name, year, month}
    """
    name = data.get("name", "").strip()
    year = data.get("year", 0)
    month = data.get("month", 0)
    if not name or not year or not month:
        return 400, {"success": False, "error": "缺少 name/year/month"}

    db = get_db()
    db.execute(
        "DELETE FROM ranking_exclusions WHERE name = ? AND year = ? AND month = ?",
        (name, year, month)
    )
    db.commit()

    rows = db.execute(
        "SELECT name, COALESCE(reason, '') as reason FROM ranking_exclusions WHERE year = ? AND month = ? ORDER BY name",
        (year, month)
    ).fetchall()
    excluded = [{"name": r["name"], "reason": r["reason"]} for r in rows]
    return 200, {"success": True, "excluded": excluded}
