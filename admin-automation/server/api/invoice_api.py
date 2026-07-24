"""发票报销汇总 API"""
import json
import os
import re
import io
from datetime import datetime

from ..db import get_db

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
INVOICE_DIR = os.path.join(BASE_DIR, "data", "invoices")
MAX_MONTHLY_AMOUNT = 100  # 每月报销上限


def _invoice_log(msg):
    """发票模块日志输出到 stderr（避免被 launcher 丢弃）"""
    import sys
    enc = sys.stderr.encoding or 'utf-8'
    safe = str(msg).encode(enc, errors='replace').decode(enc)
    print(f" {safe}", file=sys.stderr, flush=True)

def _extract_invoice_info_from_pdf(pdf_bytes):
    """从 PDF 中提取发票信息（发票号、姓名、金额）。
    用多种工具提取文本，取能正确识别到姓名的那个结果。"""
    all_texts = _extract_texts_from_pdf_bytes(pdf_bytes)
    if not all_texts:
        return {"invoice_no": "", "name": "", "amount": 0.0}

    _invoice_log(f"可用提取结果: {[t[0] for t in all_texts]}")

    best_result = None

    for tool_name, full_text in all_texts:
        _invoice_log(f"尝试 {tool_name} 提取 (长度: {len(full_text)})")
        result = _parse_invoice_text(full_text)
        if result["name"]:
            _invoice_log(f" {tool_name} 成功识别姓名: {result['name']}")
            result["_tool"] = tool_name
            return result
        if best_result is None:
            best_result = result
            best_result["_tool"] = tool_name

    _invoice_log(f" 所有工具均未识别到姓名，返回 {best_result.get('_tool', '?')} 的结果")
    return best_result or {"invoice_no": "", "name": "", "amount": 0.0}


def _parse_invoice_text(full_text):
    """从 PDF 文本中解析发票信息（发票号、姓名、金额、日期）"""
    result = {"invoice_no": "", "name": "", "amount": 0.0, "invoice_date": ""}

    # 将空白规范化
    collapsed = re.sub(r'\s+', '', full_text)

    # 1. 发票号码（20位数字）
    m = re.search(r'发票号码[：:]?(\d{20})', collapsed)
    if not m:
        m = re.search(r'(\d{20})', collapsed)
    if m:
        result["invoice_no"] = m.group(1) if m.lastindex else m.group()
        _invoice_log(f" 发票号: {result['invoice_no']}")

    # 2. 购买方名称
    # 名称后面紧跟着：统一社会信用代码、纳税人识别号、（个人）、销售方 等
    # 用非贪婪 {2,4}? + 前瞻断言，确保名字捕获在正确边界停止
    name = ""
    # 姓名边界关键字（名称后面紧跟的内容）
    name_boundary = r'(?:统一社会信用代码|纳税人识别号|[（(]个人[）)]|销售方|售|开票日期|项目名称|\d|$|/)'
    nm = re.search(r'名称[：:]?\s*([一-鿿]{2,4}?)' + name_boundary, collapsed)
    if nm:
        raw = nm.group(1)
        if not any(kw in raw for kw in ['发票', '电子', '税务', '公司', '有限', '统一', '信用', '代码', '号码']):
            name = raw

    # 备选：部分发票（如中国电信）购买方没有"名称："标签，姓名出现在日期和运营商公司名之间
    if not name:
        m = re.search(r'\d{4}年\d{1,2}月\d{1,2}日([一-鿿]{2,4})(?:中国电信|中国移动|中国联通)', collapsed)
        if m:
            raw = m.group(1)
            if not any(kw in raw for kw in ['发票', '电子', '税务', '公司', '有限', '统一', '信用', '代码', '号码', '开票人']):
                name = raw
                _invoice_log(f" 备选匹配姓名: {name}")

    if name:
        result["name"] = name
        _invoice_log(f" 姓名: {name}")

    # 3. 价税合计 — 多种格式兼容
    amount_patterns = [
        r'价税合计[（(]?小写[）)]?\s*[¥￥]\s*([\d,]+\.?\d{0,2})',
        r'合计[（(]?小写[）)]?\s*[¥￥]\s*([\d,]+\.?\d{0,2})',
        r'[¥￥]\s*([\d,]+\.\d{2})',
        r'价税合计[（(]?大写[）)]?.*?[¥￥]\s*([\d,]+\.?\d{0,2})',
    ]
    for pat in amount_patterns:
        m = re.search(pat, full_text)
        if m:
            amount_str = m.group(1).replace(',', '')
            try:
                result["amount"] = float(amount_str)
                _invoice_log(f" 金额: {result['amount']}")
            except ValueError:
                pass
            break

    # 4. 开票日期
    dm = re.search(r'开票日期[：:]?\s*(\d{4})年(\d{1,2})月(\d{1,2})日', collapsed)
    if dm:
        result["invoice_date"] = f"{dm.group(1)}-{dm.group(2).zfill(2)}-{dm.group(3).zfill(2)}"
        _invoice_log(f" 开票日期: {result['invoice_date']}")

    return result


def _extract_texts_from_pdf_bytes(pdf_bytes):
    """从 PDF 字节提取文本，返回所有可用工具的提取结果列表 [(tool_name, text), ...]"""
    results = []

    # 方式1: pdfplumber（对中国电子发票表格布局提取最准确）
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            texts = []
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    texts.append(t)
        text = "\n".join(texts)
        if text.strip():
            results.append(("pdfplumber", text))
    except ImportError:
        _invoice_log(" pdfplumber 未安装")
    except Exception as e:
        _invoice_log(f" pdfplumber 提取失败: {e}")

    # 方式2: pymupdf (fitz)
    try:
        import fitz
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        texts = []
        for page in doc:
            t = page.get_text()
            if t:
                texts.append(t)
        doc.close()
        text = "\n".join(texts)
        if text.strip():
            results.append(("pymupdf", text))
    except ImportError:
        _invoice_log(" pymupdf 未安装")
    except Exception as e:
        _invoice_log(f" pymupdf 提取失败: {e}")

    # 方式3: PyPDF2 strict=False
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(io.BytesIO(pdf_bytes), strict=False)
        texts = []
        for page in reader.pages:
            t = page.extract_text()
            if t:
                texts.append(t)
        text = "\n".join(texts)
        if text.strip():
            results.append(("PyPDF2", text))
    except Exception as e:
        _invoice_log(f" PyPDF2 提取失败: {e}")

    # 方式4: 直接从原始字节搜索文本
    try:
        raw_text = pdf_bytes.decode('latin-1')
        import re as _re
        parts = _re.findall(r'\(([^)]+)\)', raw_text)
        text = ''.join(p for p in parts if any('一' <= c <= '鿿' for c in p))
        if text:
            results.append(("raw", text))
    except Exception as e:
        _invoice_log(f" 原始字节搜索失败: {e}")

    return results


def _extract_text_from_pdf_bytes(pdf_bytes):
    """从 PDF 字节提取文本（兼容旧接口，返回第一个可用结果）"""
    results = _extract_texts_from_pdf_bytes(pdf_bytes)
    return results[0][1] if results else ""


def _get_pdf_text(pdf_bytes):
    """提取 PDF 全部文本（用于调试）"""
    return _extract_text_from_pdf_bytes(pdf_bytes)


def _get_invoice_dir(cycle_name, sub=""):
    d = os.path.join(INVOICE_DIR, cycle_name, sub)
    os.makedirs(d, exist_ok=True)
    return d


def _save_upload_file(file_data, save_path):
    """保存上传的文件，file_data 是 cgi.FieldStorage 或 bytes"""
    if isinstance(file_data, bytes):
        with open(save_path, "wb") as f:
            f.write(file_data)
    else:
        with open(save_path, "wb") as f:
            f.write(file_data.file.read())


# ══════════════════════════════════════
# 周期管理
# ══════════════════════════════════════

def handle_list_cycles(handler, _data):
    db = get_db()
    rows = db.execute(
        "SELECT * FROM invoice_cycles ORDER BY year DESC, month_start DESC"
    ).fetchall()
    cycles = [dict(r) for r in rows]
    return 200, {"success": True, "cycles": cycles}


def handle_create_cycle(handler, data):
    db = get_db()
    year = int(data.get("year", 0))
    month_start = int(data.get("month_start", 0))
    month_end = int(data.get("month_end", 0))
    if not year or not month_start or not month_end:
        return 400, {"success": False, "error": "请填写完整的周期信息"}

    # 汇总月份 = 结束月份的下一个月
    summary_month = month_end + 1
    if summary_month > 12:
        summary_month = 1

    cycle_name = f"{year}年{month_start}-{month_end}月话费报销"
    db.execute(
        "INSERT INTO invoice_cycles (cycle_name, year, month_start, month_end, summary_month) VALUES (?, ?, ?, ?, ?)",
        (cycle_name, year, month_start, month_end, summary_month)
    )
    db.commit()

    # 创建文件夹结构
    _get_invoice_dir(cycle_name, "PDF")
    _get_invoice_dir(cycle_name, "PDF+验真")

    return 200, {"success": True, "message": f"周期 {cycle_name} 创建成功"}


def handle_delete_cycle(handler, data):
    db = get_db()
    cycle_id = int(data.get("cycle_id", 0))
    if not cycle_id:
        return 400, {"success": False, "error": "缺少 cycle_id"}

    cycle = db.execute("SELECT * FROM invoice_cycles WHERE id = ?", (cycle_id,)).fetchone()
    if not cycle:
        return 404, {"success": False, "error": "周期不存在"}

    import shutil
    cycle_dir = os.path.join(INVOICE_DIR, cycle["cycle_name"])
    if os.path.isdir(cycle_dir):
        shutil.rmtree(cycle_dir)
        _invoice_log(f"已删除周期目录: {cycle_dir}")

    db.execute("DELETE FROM invoice_records WHERE cycle_id = ?", (cycle_id,))
    db.execute("DELETE FROM invoice_cycles WHERE id = ?", (cycle_id,))
    db.commit()
    return 200, {"success": True, "message": "删除成功"}


# ══════════════════════════════════════
# 发票上传（员工端，无需登录）
# ══════════════════════════════════════

def handle_upload_invoice(handler, data):
    """员工上传发票 PDF。无需登录。
    系统自动从 PDF 提取姓名、发票号、金额。
    """
    cycle_id = int(data.get("cycle_id", 0))
    if not cycle_id:
        return 400, {"success": False, "error": "缺少 cycle_id"}

    db = get_db()
    cycle = db.execute("SELECT * FROM invoice_cycles WHERE id = ?", (cycle_id,)).fetchone()
    if not cycle:
        return 404, {"success": False, "error": "周期不存在"}
    cycle_name = cycle["cycle_name"]

    pdf_dir = _get_invoice_dir(cycle_name, "PDF")

    extracted = {"name": "", "month1": {}, "month2": {}}

    # ── 处理第1个月 PDF ──
    pdf1 = data.get("pdf1")
    pdf1_bytes = None
    if pdf1 and hasattr(pdf1, 'filename') and pdf1.filename:
        pdf1_bytes = pdf1.file.read()
        info = _extract_invoice_info_from_pdf(pdf1_bytes)
        extracted["month1"] = info
        if info.get("name") and not extracted["name"]:
            extracted["name"] = info["name"]

    # ── 处理第2个月 PDF ──
    pdf2 = data.get("pdf2")
    pdf2_bytes = None
    if pdf2 and hasattr(pdf2, 'filename') and pdf2.filename:
        pdf2_bytes = pdf2.file.read()
        info = _extract_invoice_info_from_pdf(pdf2_bytes)
        extracted["month2"] = info
        if info.get("name") and not extracted["name"]:
            extracted["name"] = info["name"]

    # ── 如果 PDF 未提取到姓名 ──
    name = extracted["name"] or data.get("name", "").strip()
    if not name:
        # 保存 PDF 到临时位置以便排查（加时间戳，保留每次失败记录）
        debug_dir = _get_invoice_dir(cycle_name, "_debug")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        debug_info = {"pdf1_text": "", "pdf2_text": ""}
        if pdf1_bytes:
            fn = f"pdf1_{ts}.pdf"
            with open(os.path.join(debug_dir, fn), "wb") as f:
                f.write(pdf1_bytes)
            debug_info["pdf1_text"] = _get_pdf_text(pdf1_bytes)
        if pdf2_bytes:
            fn = f"pdf2_{ts}.pdf"
            with open(os.path.join(debug_dir, fn), "wb") as f:
                f.write(pdf2_bytes)
            debug_info["pdf2_text"] = _get_pdf_text(pdf2_bytes)
        return 400, {
            "success": False,
            "error": "未能从 PDF 中识别姓名",
            "debug": debug_info
        }

    invoice_no1 = extracted["month1"].get("invoice_no", "")
    invoice_no2 = extracted["month2"].get("invoice_no", "")
    invoice_date1 = extracted["month1"].get("invoice_date", "")
    invoice_date2 = extracted["month2"].get("invoice_date", "")
    month1_amount = min(extracted["month1"].get("amount", 0) or 0, MAX_MONTHLY_AMOUNT)
    month2_amount = min(extracted["month2"].get("amount", 0) or 0, MAX_MONTHLY_AMOUNT)
    total = month1_amount + month2_amount

    # ── 保存 PDF 文件 ──
    pdf1_filename = ""
    pdf2_filename = ""
    verify_dir = _get_invoice_dir(cycle_name, f"PDF+验真/{name}")
    if pdf1_bytes:
        pdf1_filename = f"{name}+{invoice_no1}.pdf" if invoice_no1 else f"{name}+月1.pdf"
        with open(os.path.join(pdf_dir, pdf1_filename), "wb") as f:
            f.write(pdf1_bytes)
        # 同步复制到 PDF+验真
        with open(os.path.join(verify_dir, pdf1_filename), "wb") as f:
            f.write(pdf1_bytes)
    if pdf2_bytes:
        pdf2_filename = f"{name}+{invoice_no2}.pdf" if invoice_no2 else f"{name}+月2.pdf"
        with open(os.path.join(pdf_dir, pdf2_filename), "wb") as f:
            f.write(pdf2_bytes)
        # 同步复制到 PDF+验真
        with open(os.path.join(verify_dir, pdf2_filename), "wb") as f:
            f.write(pdf2_bytes)

    # ── 话费账单明细（可选） ──
    list_filename = ""
    inv_list = data.get("invoice_list")
    if inv_list and hasattr(inv_list, 'filename') and inv_list.filename:
        ext = os.path.splitext(inv_list.filename)[1] or ".png"
        list_filename = f"{name}话费账单明细{ext}"
        with open(os.path.join(verify_dir, list_filename), "wb") as f:
            f.write(inv_list.file.read())

    # ── 写入/更新数据库 ──
    existing = db.execute(
        "SELECT id, month1_amount, month2_amount, invoice_no1, invoice_no2, verification1_filename, verification2_filename FROM invoice_records WHERE cycle_id = ? AND name = ?",
        (cycle_id, name)
    ).fetchone()

    if existing:
        # 只更新实际提供了数据的字段，避免覆盖已有数据
        fields = []
        values = []
        if pdf1_bytes:
            fields.append("month1_amount=?"); values.append(month1_amount)
            fields.append("pdf1_filename=?"); values.append(pdf1_filename)
            fields.append("invoice_no1=?"); values.append(invoice_no1)
            fields.append("invoice_date1=?"); values.append(invoice_date1)
        if pdf2_bytes:
            fields.append("month2_amount=?"); values.append(month2_amount)
            fields.append("pdf2_filename=?"); values.append(pdf2_filename)
            fields.append("invoice_no2=?"); values.append(invoice_no2)
            fields.append("invoice_date2=?"); values.append(invoice_date2)
        # total 根据现有数据重新计算
        existing_m1 = existing["month1_amount"] or 0
        existing_m2 = existing["month2_amount"] or 0
        new_m1 = round(month1_amount if pdf1_bytes else existing_m1, 2)
        new_m2 = round(month2_amount if pdf2_bytes else existing_m2, 2)
        fields.append("total_amount=?"); values.append(round(new_m1 + new_m2, 2))
        if list_filename:
            fields.append("invoice_list_filename=?"); values.append(list_filename)
        # 根据实际验真状态计算 status，不盲目重置已验证的记录
        need1 = bool(invoice_no1) if pdf1_bytes else bool(existing["invoice_no1"])
        need2 = bool(invoice_no2) if pdf2_bytes else bool(existing["invoice_no2"])
        has1 = bool(existing["verification1_filename"])
        has2 = bool(existing["verification2_filename"])
        all_verified = (not need1 or has1) and (not need2 or has2)
        fields.append("status=?"); values.append('verified' if all_verified else 'pending_verification')
        fields.append("uploaded_at=CURRENT_TIMESTAMP")
        values.append(existing["id"])
        db.execute(f"UPDATE invoice_records SET {', '.join(fields)} WHERE id=?", values)
    else:
        db.execute(
            """INSERT INTO invoice_records
               (cycle_id, name, month1_amount, month2_amount, total_amount,
                pdf1_filename, pdf2_filename, invoice_no1, invoice_no2,
                invoice_date1, invoice_date2,
                invoice_list_filename, status)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending_verification')""",
            (cycle_id, name, month1_amount, month2_amount, total,
             pdf1_filename, pdf2_filename,
             invoice_no1, invoice_no2,
             invoice_date1, invoice_date2,
             list_filename)
        )
    db.commit()

    # 获取记录ID
    if existing:
        record_id = existing["id"]
    else:
        record_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

    return 200, {
        "success": True,
        "message": f"{name} 发票提交成功",
        "record_id": record_id,
        "name": name,
        "invoice_no1": invoice_no1,
        "invoice_no2": invoice_no2,
        "invoice_date1": invoice_date1,
        "invoice_date2": invoice_date2,
        "month1_amount": month1_amount,
        "month2_amount": month2_amount,
        "total": total,
    }


# ══════════════════════════════════════
# 记录管理（管理员端）
# ══════════════════════════════════════

def handle_list_records(handler, data):
    cycle_id = int(data.get("cycle_id", 0))
    if not cycle_id:
        return 400, {"success": False, "error": "缺少 cycle_id"}
    db = get_db()
    rows = db.execute(
        "SELECT * FROM invoice_records WHERE cycle_id = ? ORDER BY id",
        (cycle_id,)
    ).fetchall()
    records = []
    for r in rows:
        d = dict(r)
        d.pop("invoice_no1", None)
        d.pop("invoice_no2", None)
        records.append(d)
    stats = {
        "total": len(records),
        "verified": sum(1 for r in records if r["status"] == "verified"),
        "pending": sum(1 for r in records if r["status"] == "pending_verification"),
        "draft": sum(1 for r in records if r["status"] == "draft"),
        "total_amount": sum(r["total_amount"] or 0 for r in records),
    }
    return 200, {"success": True, "records": records, "stats": stats}


def handle_update_record(handler, data):
    record_id = int(data.get("record_id", 0))
    if not record_id:
        return 400, {"success": False, "error": "缺少 record_id"}
    db = get_db()
    updates = {}
    for field in ["bank_name", "month1_amount", "month2_amount", "status", "name"]:
        if field in data:
            updates[field] = data[field]
    if "month1_amount" in updates:
        updates["month1_amount"] = min(float(updates["month1_amount"]), MAX_MONTHLY_AMOUNT)
    if "month2_amount" in updates:
        updates["month2_amount"] = min(float(updates["month2_amount"]), MAX_MONTHLY_AMOUNT)
    if updates:
        set_clause = ", ".join(f"{k}=?" for k in updates)
        vals = list(updates.values()) + [record_id]
        if "month1_amount" in updates or "month2_amount" in updates:
            # 重新计算 total
            record = db.execute("SELECT * FROM invoice_records WHERE id=?", (record_id,)).fetchone()
            m1 = updates.get("month1_amount", record["month1_amount"] or 0)
            m2 = updates.get("month2_amount", record["month2_amount"] or 0)
            db.execute(f"UPDATE invoice_records SET {set_clause}, total_amount=? WHERE id=?",
                       vals[:len(updates)] + [m1 + m2] + [record_id])
        else:
            db.execute(f"UPDATE invoice_records SET {set_clause} WHERE id=?", vals)
        db.commit()
    return 200, {"success": True, "message": "更新成功"}


def handle_delete_record(handler, data):
    """管理员删除记录，同时删除该员工上传的所有文件"""
    record_id = int(data.get("record_id", 0))
    if not record_id:
        return 400, {"success": False, "error": "缺少 record_id"}

    db = get_db()
    record = db.execute("SELECT * FROM invoice_records WHERE id = ?", (record_id,)).fetchone()
    if not record:
        return 404, {"success": False, "error": "记录不存在"}

    cycle = db.execute("SELECT * FROM invoice_cycles WHERE id = ?", (record["cycle_id"],)).fetchone()
    if not cycle:
        return 404, {"success": False, "error": "周期不存在"}

    import shutil
    cycle_name = cycle["cycle_name"]
    name = record["name"]

    # 删除 PDF 文件
    for pdf_field in ["pdf1_filename", "pdf2_filename"]:
        fn = record[pdf_field] or ""
        if fn:
            pdf_path = os.path.join(INVOICE_DIR, cycle_name, "PDF", fn)
            if os.path.isfile(pdf_path):
                os.remove(pdf_path)
                _invoice_log(f"已删除: {pdf_path}")

    # 删除验真目录
    verify_dir = os.path.join(INVOICE_DIR, cycle_name, "PDF+验真", name)
    if os.path.isdir(verify_dir):
        shutil.rmtree(verify_dir)
        _invoice_log(f"已删除目录: {verify_dir}")

    db.execute("DELETE FROM invoice_records WHERE id = ?", (record_id,))
    db.commit()
    return 200, {"success": True, "message": f"{name} 的记录已删除"}


def handle_verify_record(handler, data):
    """员工确认验真"""
    record_id = int(data.get("record_id", 0))
    verify_code = data.get("verify_code", "").strip()
    if not record_id:
        return 400, {"success": False, "error": "缺少 record_id"}

    db = get_db()
    record = db.execute("SELECT * FROM invoice_records WHERE id = ?", (record_id,)).fetchone()
    if not record:
        return 404, {"success": False, "error": "记录不存在"}

    db.execute(
        "UPDATE invoice_records SET status='verified', verification1_filename=?, uploaded_at=CURRENT_TIMESTAMP WHERE id=?",
        (verify_code or "", record_id)
    )
    db.commit()
    return 200, {"success": True, "message": "验真确认成功"}


def handle_verify_upload(handler, data):
    """上传验真截图确认。month=1 或 2，分别存到 verification1/2_filename。"""
    record_id = int(data.get("record_id", 0))
    month = int(data.get("month", 1))
    if not record_id:
        return 400, {"success": False, "error": "缺少 record_id"}

    db = get_db()
    record = db.execute("SELECT * FROM invoice_records WHERE id = ?", (record_id,)).fetchone()
    if not record:
        return 404, {"success": False, "error": "记录不存在"}

    cycle = db.execute("SELECT * FROM invoice_cycles WHERE id = ?", (record["cycle_id"],)).fetchone()
    if not cycle:
        return 404, {"success": False, "error": "周期不存在"}

    screenshot = data.get("screenshot")
    if not screenshot or not hasattr(screenshot, 'filename') or not screenshot.filename:
        return 400, {"success": False, "error": "请上传验真截图"}

    verify_dir = _get_invoice_dir(cycle["cycle_name"], f"PDF+验真/{record['name']}")
    ext = os.path.splitext(screenshot.filename)[1] or ".png"
    invoice_month = cycle["month_start"] if month == 1 else cycle["month_end"]
    filename = f"{record['name']}{invoice_month}月发票验真{ext}"
    filepath = os.path.join(verify_dir, filename)

    with open(filepath, "wb") as f:
        f.write(screenshot.file.read())

    # 同时将对应的发票 PDF 复制到 PDF+验真 目录
    pdf_dir = _get_invoice_dir(cycle["cycle_name"], "PDF")
    pdf_filename = record["pdf1_filename"] if month == 1 else record["pdf2_filename"]
    if pdf_filename:
        src = os.path.join(pdf_dir, pdf_filename)
        dst = os.path.join(verify_dir, pdf_filename)
        if os.path.exists(src) and not os.path.exists(dst):
            import shutil
            shutil.copy2(src, dst)

    col = "verification1_filename" if month == 1 else "verification2_filename"
    db.execute(
        f"UPDATE invoice_records SET {col}=? WHERE id=?",
        (filename, record_id)
    )
    db.commit()  # 先提交，确保后续 SELECT 能读到最新数据

    # 所有有发票的月份都验真 → verified
    try:
        record2 = db.execute("SELECT * FROM invoice_records WHERE id = ?", (record_id,)).fetchone()
        need1 = bool(record2["invoice_no1"])
        need2 = bool(record2["invoice_no2"])
        has1 = bool(record2["verification1_filename"])
        has2 = bool(record2["verification2_filename"])
        all_done = (not need1 or has1) and (not need2 or has2)
        _invoice_log(f"verify status check: need1={need1} need2={need2} has1={has1} has2={has2} all_done={all_done}")
        if all_done:
            db.execute("UPDATE invoice_records SET status='verified' WHERE id=?", (record_id,))
            _invoice_log(f"status -> verified for record {record_id}")
        db.commit()
    except Exception as e:
        _invoice_log(f"verify status update error: {e}")
        db.commit()

    return 200, {"success": True, "message": f"第{month}月验真截图已提交"}


# ══════════════════════════════════════
# 导出
# ══════════════════════════════════════

def handle_export_summary(handler, data):
    """导出 ZIP（Excel 统计表 + PDF 文件夹原件）"""
    cycle_id = int(data.get("cycle_id", 0))
    if not cycle_id:
        return 400, {"success": False, "error": "缺少 cycle_id"}

    db = get_db()
    cycle = db.execute("SELECT * FROM invoice_cycles WHERE id = ?", (cycle_id,)).fetchone()
    if not cycle:
        return 404, {"success": False, "error": "周期不存在"}

    records = db.execute(
        "SELECT * FROM invoice_records WHERE cycle_id = ? ORDER BY id",
        (cycle_id,)
    ).fetchall()

    import openpyxl
    import zipfile
    from email.header import Header

    cycle_name = cycle["cycle_name"]
    zip_buf = io.BytesIO()

    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:

        # ── 1. 生成 Excel 统计表 ──
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "票据登记表"

        font_normal = openpyxl.styles.Font(name='微软雅黑', size=10)
        font_bold = openpyxl.styles.Font(name='微软雅黑', size=10, bold=True)
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )
        center_align = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        center_wrap = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 标题行
        ws.merge_cells('A1:H1')
        ws['A1'] = '票据登记表'
        ws['A1'].font = font_bold
        ws['A1'].alignment = center_align
        ws['A1'].border = thin_border

        # 表头
        headers = ['序号', '部门', '姓名', '报销票据类别',
                   f'{cycle["month_start"]}月票据金额', f'{cycle["month_end"]}月票据金额',
                   '票据合计金额', '备注']
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=i, value=h)
            cell.font = font_bold
            cell.alignment = center_align
            cell.border = thin_border

        # 数据行
        for idx, r in enumerate(records, 1):
            row_num = idx + 2
            m1 = r['month1_amount'] or 0
            m2 = r['month2_amount'] or 0

            # 序号
            c = ws.cell(row=row_num, column=1, value=idx)
            c.font = font_bold
            c.alignment = center_align
            c.border = thin_border

            # 部门
            c = ws.cell(row=row_num, column=2, value='客服中心')
            c.font = font_normal
            c.alignment = center_wrap
            c.border = thin_border

            # 姓名
            c = ws.cell(row=row_num, column=3, value=r['name'])
            c.font = font_normal
            c.alignment = center_align
            c.border = thin_border

            # 报销票据类别
            c = ws.cell(row=row_num, column=4, value='费用')
            c.font = font_normal
            c.alignment = center_align
            c.border = thin_border

            # X月票据金额
            c = ws.cell(row=row_num, column=5, value=m1)
            c.font = font_normal
            c.alignment = center_align
            c.border = thin_border
            c.number_format = '0.00_ '

            # Y月票据金额
            c = ws.cell(row=row_num, column=6, value=m2)
            c.font = font_normal
            c.alignment = center_align
            c.border = thin_border
            c.number_format = '0.00_ '

            # 票据合计金额（公式）
            c = ws.cell(row=row_num, column=7)
            c.value = f'=E{row_num}+F{row_num}'
            c.font = font_normal
            c.alignment = center_align
            c.border = thin_border
            c.number_format = '0.00_ '

            # 备注
            c = ws.cell(row=row_num, column=8, value=f'{cycle["year"]}年{cycle["month_start"]}-{cycle["month_end"]}月份话费')
            c.font = font_normal
            c.alignment = center_wrap
            c.border = thin_border

        # 列宽
        col_widths = {'A': 6, 'B': 10, 'C': 10, 'D': 14, 'E': 14, 'F': 14, 'G': 14, 'H': 24}
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        excel_buf = io.BytesIO()
        wb.save(excel_buf)
        excel_buf.seek(0)

        outer = f'{cycle["month_start"]}-{cycle["month_end"]}月话费报销-在线客服'

        excel_name = f'{cycle["year"]}年{cycle["month_start"]}月-{cycle["month_end"]}月的票据统计表-在线客服.xlsx'
        zf.writestr(f'{outer}/{excel_name}', excel_buf.getvalue())

        # ── 2. 打包 PDF 文件夹 ──
        invoice_dir = os.path.join(INVOICE_DIR, cycle_name)
        if os.path.isdir(invoice_dir):
            for root, dirs, files in os.walk(invoice_dir):
                # 跳过调试目录
                dirs[:] = [d for d in dirs if d != '_debug']
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, invoice_dir)
                    zf.write(file_path, f'{outer}/{arcname}')

    zip_buf.seek(0)

    zip_filename = f'{cycle["year"]}年{cycle["month_start"]}-{cycle["month_end"]}月话费报销.zip'

    handler.send_response(200)
    handler.send_header("Content-Type", "application/zip")
    handler.send_header("Content-Disposition",
                        f'attachment; filename="{Header(zip_filename, "utf-8").encode()}"')
    handler.send_header("Access-Control-Allow-Origin", "*")
    handler.send_header("Access-Control-Expose-Headers", "Content-Disposition")
    handler.end_headers()
    handler.wfile.write(zip_buf.getvalue())
    return None
